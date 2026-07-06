import os
import re
import uuid
import json
import random
import base64
import unicodedata
import html
import mimetypes
import posixpath
import zipfile
import xml.etree.ElementTree as ET
import docx2txt
import jieba
from typing import Optional
from datetime import timedelta, timezone, datetime
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, Form
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from app.database import get_db
from app.models.qa_feedback import QaFeedback
from app.models.qa_history import QaSession, QaMessage
from app.schemas.qa_feedback import QaFeedbackCreate, QaFeedbackOut
from app.api.auth import get_current_user, oauth2_scheme

router = APIRouter()

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "qa_uploads")
BEIJING_TZ = timezone(timedelta(hours=8))


def _to_beijing_iso(dt):
    if not dt:
        return None
    return dt.replace(tzinfo=timezone.utc).astimezone(BEIJING_TZ).isoformat(timespec="seconds")


def _get_user_id_from_token(token: str, db: Session):
    if not token:
        return None
    from jose import jwt
    from app.api.auth import SECRET_KEY, ALGORITHM
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if username:
            from app.crud.user import get_user as crud_get_user
            u = crud_get_user(db, username=username)
            if u:
                return u.id
    except Exception:
        pass
    return None


def _save_qa_history(db: Session, session_type: str, question: str, answer_data: dict, user_id: int, session_id: Optional[int] = None, sources: list = None, search_hit: int = 0, relevance_score: float = 0.0):
    title = question[:80] + ("..." if len(question) > 80 else "")
    sess = None
    if session_id:
        sess = db.query(QaSession).filter(QaSession.id == session_id, QaSession.user_id == user_id).first()
    if sess:
        sess.updated_at = datetime.utcnow()
    else:
        sess = QaSession(user_id=user_id, session_type=session_type, title=title)
        db.add(sess)
        db.flush()

    db.add(QaMessage(session_id=sess.id, role="user", content=question))
    db.add(QaMessage(
        session_id=sess.id,
        role="assistant",
        content=answer_data.get("answer", ""),
        sources=json.dumps(sources or answer_data.get("sources", []), ensure_ascii=False),
        search_hit=search_hit,
        relevance_score=relevance_score,
    ))
    db.commit()
    return sess.id


class GeneralQAInput(BaseModel):
    question: str
    knowledge_ids: list = []
    session_id: Optional[int] = None


QA_STOPWORDS = {
    "呢", "吗", "呀", "啊", "的", "了", "着", "过",
    "请问", "这个", "那个", "这些", "那些",
}


def _normalize_text(text: str):
    return re.sub(r"\s+", " ", (text or "")).strip()


_jieba_initialized = False


def _ensure_jieba():
    global _jieba_initialized
    if not _jieba_initialized:
        jieba.setLogLevel(20)
        _jieba_initialized = True


def _tokenize(text: str):
    _ensure_jieba()
    normalized = _normalize_text(text).lower()
    tokens = _jieba_cut(normalized)
    tokens = [t for t in tokens if len(t) >= 2 and t not in QA_STOPWORDS]
    if not tokens:
        tokens = [t for t in _jieba_cut(normalized) if len(t) >= 1 and t not in QA_STOPWORDS]
    return tokens


def _jieba_cut(text: str):
    pure_chinese = re.sub(r"[^\u4e00-\u9fff]+", " ", text)
    words = []
    if pure_chinese.strip():
        words.extend([w for w in jieba.cut(pure_chinese) if w.strip() and len(w.strip()) >= 1])
    alphanum = re.findall(r"[a-z0-9_\-\.]+", text.lower())
    words.extend([w for w in alphanum if len(w) >= 2])
    return words


def _quick_tokenize(text: str):
    normalized = _normalize_text(text).lower()
    parts = re.findall(r"[\u4e00-\u9fff]{1,}|[a-z0-9_\-\.]+", normalized)
    return [p for p in parts if p and p not in QA_STOPWORDS]


def _char_ngrams(text: str, n: int = 2):
    pure = re.sub(r"\s+", "", _normalize_text(text).lower())
    if len(pure) < n:
        return {pure} if pure else set()
    return {pure[i:i + n] for i in range(len(pure) - n + 1)}


def _split_content_to_chunks(content: str, chunk_size: int = 500, overlap: int = 120):
    text = _normalize_text(content)
    if not text:
        return []
    if len(text) <= chunk_size:
        return [text]
    chunks = []
    start = 0
    step = max(chunk_size - overlap, 120)
    while start < len(text):
        end = min(len(text), start + chunk_size)
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= len(text):
            break
        start += step
    return chunks


def _score_chunk(question: str, chunk: str, title: str = ""):
    q_tokens = set(_tokenize(question))
    c_tokens = set(_tokenize(chunk))
    title_tokens = set(_tokenize(title))
    token_overlap = len(q_tokens & c_tokens)
    title_overlap = len(q_tokens & title_tokens)

    q_ngrams = _char_ngrams(question)
    c_ngrams = _char_ngrams(chunk)
    ngram_overlap = len(q_ngrams & c_ngrams)

    keyword_hits = 0
    for token in q_tokens:
        if len(token) >= 2:
            keyword_hits += chunk.lower().count(token)

    score = token_overlap * 3.5 + title_overlap * 2.0 + min(keyword_hits, 10) * 0.8 + min(ngram_overlap, 30) * 0.25
    return round(score, 4)


def _rank_document_chunks(question: str, documents: list, limit: int = 5):
    ranked = []
    for doc in documents:
        content = _normalize_text(doc.get("content", ""))
        if not content:
            continue
        chunks = _split_content_to_chunks(content)
        for idx, chunk in enumerate(chunks):
            score = _score_chunk(question, chunk, doc.get("title", ""))
            if score <= 0:
                continue
            ranked.append({
                "document_id": doc.get("document_id"),
                "title": doc.get("title") or "未命名文档",
                "chunk": chunk,
                "score": score,
                "chunk_index": idx,
            })
    ranked.sort(key=lambda item: item["score"], reverse=True)
    return ranked[:limit]


def _needs_clarification(question: str, ranked_sources: list):
    tokens = _tokenize(question)
    if len(_normalize_text(question)) <= 2:
        return True
    if not ranked_sources:
        return True
    top_score = ranked_sources[0].get("score", 0)
    if top_score < 0.5:
        return True
    return False


def _build_clarification_answer(question: str, ranked_sources: list):
    if not ranked_sources:
        return {
            "answer": f"我暂时没有在已选资料中定位到与“{question}”直接相关的内容。请补充更具体的信息，例如产品名称、章节名、参数名、步骤名或场景。",
            "source": "",
        }

    titles = []
    for item in ranked_sources[:3]:
        title = item.get("title")
        if title and title not in titles:
            titles.append(title)

    hints = "、".join(titles)
    return {
        "answer": f"您的问题还不够具体，我已在这些资料里找到可能相关的内容：{hints}。请补充想确认的对象、指标、步骤或章节，我再基于对应片段给出准确答案。",
        "source": hints,
    }


def _build_context_from_sources(ranked_sources: list, max_chars: int = 8000):
    parts = []
    total = 0
    normalized_sources = []
    seen_titles = set()
    for item in ranked_sources:
        snippet = item.get("chunk", "")
        title = item.get("title") or "未命名文档"
        block = f"--- 文档: {title} / 片段 {item.get('chunk_index', 0) + 1} ---\n{snippet}"
        if total + len(block) > max_chars and parts:
            break
        parts.append(block)
        total += len(block)
        if title and title not in seen_titles:
            seen_titles.add(title)
            normalized_sources.append({
                "document_id": item.get("document_id"),
                "title": title,
                "snippet": snippet[:120],
                "score": item.get("score", 0),
            })
    return "\n\n".join(parts), normalized_sources


def _extract_file_content(file_path, file_type):
    file_type = file_type.lower() if file_type else ""
    try:
        if file_type in ["txt", "md", "markdown", "json", "xml", "html", "css", "js", "py"]:
            with open(file_path, "r", encoding="utf-8") as f:
                return f.read()
        elif file_type == "pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            return "\n".join([page.extract_text() or "" for page in reader.pages])
        elif file_type == "docx":
            return docx2txt.process(file_path) or ""
        elif file_type in ["xlsx", "xls"]:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            rows = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows.append(f"[Sheet: {sheet_name}]")
                for row in ws.iter_rows(values_only=True):
                    rows.append(" | ".join([str(c) if c is not None else "" for c in row]))
            return "\n".join(rows)
        else:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
    except Exception as e:
        return f"[文件解析失败: {str(e)}]"


def _extract_embedded_images(file_path, file_type):
    file_type = file_type.lower() if file_type else ""
    image_entries = []

    try:
        if file_type in ["docx", "xlsx", "xls"]:
            media_prefixes = ["word/media/", "xl/media/"]
            with zipfile.ZipFile(file_path) as zf:
                for name in sorted(zf.namelist()):
                    if not any(name.startswith(prefix) for prefix in media_prefixes):
                        continue
                    mime, _ = mimetypes.guess_type(name)
                    if not mime or not mime.startswith("image/"):
                        continue
                    raw = zf.read(name)
                    image_entries.append({
                        "name": os.path.basename(name),
                        "mime": mime,
                        "data_url": f"data:{mime};base64,{base64.b64encode(raw).decode('ascii')}",
                    })
        elif file_type == "pdf":
            try:
                import fitz
            except Exception:
                return []

            seen = set()
            doc = fitz.open(file_path)
            try:
                for page_index in range(len(doc)):
                    page = doc[page_index]
                    for img_index, img in enumerate(page.get_images(full=True), start=1):
                        xref = img[0]
                        if xref in seen:
                            continue
                        seen.add(xref)
                        extracted = doc.extract_image(xref)
                        raw = extracted.get("image")
                        ext = extracted.get("ext", "png")
                        if not raw:
                            continue
                        mime = mimetypes.guess_type(f"image.{ext}")[0] or "image/png"
                        image_entries.append({
                            "name": f"page-{page_index + 1}-image-{img_index}.{ext}",
                            "mime": mime,
                            "data_url": f"data:{mime};base64,{base64.b64encode(raw).decode('ascii')}",
                        })
            finally:
                doc.close()
    except Exception:
        return []

    return image_entries


def _data_url(raw, name):
    mime, _ = mimetypes.guess_type(name)
    mime = mime or "image/png"
    return f"data:{mime};base64,{base64.b64encode(raw).decode('ascii')}"


def _build_docx_preview_html(file_path):
    ns = {
        "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "v": "urn:schemas-microsoft-com:vml",
    }

    with zipfile.ZipFile(file_path) as zf:
        rels = {}
        rel_root = ET.fromstring(zf.read("word/_rels/document.xml.rels"))
        for rel in rel_root:
            rid = rel.attrib.get("Id")
            target = rel.attrib.get("Target", "")
            if rid and target:
                rels[rid] = posixpath.normpath("word/" + target.lstrip("/")) if not target.startswith("word/") else target

        doc_root = ET.fromstring(zf.read("word/document.xml"))

        def render_run(run):
            parts = []
            for node in run.iter():
                tag = node.tag.split("}")[-1]
                if tag == "t" and node.text:
                    parts.append(html.escape(node.text))
                elif tag == "tab":
                    parts.append("&emsp;")
                elif tag == "br":
                    parts.append("<br>")
                elif tag == "blip":
                    rid = node.attrib.get(f"{{{ns['r']}}}embed") or node.attrib.get(f"{{{ns['r']}}}link")
                    media_path = rels.get(rid)
                    if media_path and media_path in zf.namelist():
                        name = os.path.basename(media_path)
                        parts.append(f'<img class="doc-preview-inline-image" src="{_data_url(zf.read(media_path), name)}" alt="{html.escape(name)}">')
                elif tag == "imagedata":
                    rid = node.attrib.get(f"{{{ns['r']}}}id")
                    media_path = rels.get(rid)
                    if media_path and media_path in zf.namelist():
                        name = os.path.basename(media_path)
                        parts.append(f'<img class="doc-preview-inline-image" src="{_data_url(zf.read(media_path), name)}" alt="{html.escape(name)}">')
            return "".join(parts)

        def render_paragraph(p):
            content = "".join(render_run(r) for r in p.findall("w:r", ns))
            return f"<p>{content or '&nbsp;'}</p>"

        body = doc_root.find("w:body", ns)
        blocks = []
        for child in list(body or []):
            tag = child.tag.split("}")[-1]
            if tag == "p":
                blocks.append(render_paragraph(child))
            elif tag == "tbl":
                rows = []
                for tr in child.findall("w:tr", ns):
                    cells = []
                    for tc in tr.findall("w:tc", ns):
                        cell_html = "".join(render_paragraph(p) for p in tc.findall("w:p", ns))
                        cells.append(f"<td>{cell_html}</td>")
                    rows.append("<tr>" + "".join(cells) + "</tr>")
                blocks.append('<table class="doc-preview-table">' + "".join(rows) + "</table>")

    return '<div class="doc-preview-html doc-preview-docx">' + "".join(blocks) + "</div>"


def _build_pdf_preview_html(file_path):
    try:
        import fitz
    except Exception:
        return ""

    pages = []
    doc = fitz.open(file_path)
    try:
        for page_index, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5), alpha=False)
            data_url = f"data:image/png;base64,{base64.b64encode(pix.tobytes('png')).decode('ascii')}"
            pages.append(f'<figure class="doc-preview-page"><img src="{data_url}" alt="第 {page_index} 页"><figcaption>第 {page_index} 页</figcaption></figure>')
    finally:
        doc.close()
    return '<div class="doc-preview-html doc-preview-pdf">' + "".join(pages) + "</div>"


def _build_xlsx_preview_html(file_path):
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        anchored_images = {}
        for img in getattr(ws, "_images", []):
            try:
                row = img.anchor._from.row + 1
                col = img.anchor._from.col + 1
                anchored_images.setdefault((row, col), []).append(_data_url(img._data(), getattr(img, "path", "image.png")))
            except Exception:
                continue

        rows = []
        image_rows = [r for r, _ in anchored_images.keys()]
        image_cols = [c for _, c in anchored_images.keys()]
        max_row = max([ws.max_row] + image_rows)
        max_col = max([ws.max_column] + image_cols)
        for r in range(1, max_row + 1):
            cells = []
            for c in range(1, max_col + 1):
                value = ws.cell(r, c).value
                content = html.escape(str(value)) if value is not None else ""
                for data_url in anchored_images.get((r, c), []):
                    content += f'<img class="doc-preview-inline-image" src="{data_url}" alt="sheet image">'
                cells.append(f"<td>{content}</td>")
            rows.append("<tr>" + "".join(cells) + "</tr>")
        sheets.append(f'<section class="doc-preview-sheet"><h4>{html.escape(ws.title)}</h4><table class="doc-preview-table">' + "".join(rows) + "</table></section>")
    return '<div class="doc-preview-html doc-preview-xlsx">' + "".join(sheets) + "</div>"


def _build_preview_html(file_path, file_type, text):
    file_type = file_type.lower() if file_type else ""
    try:
        if file_type == "docx":
            return _build_docx_preview_html(file_path)
        if file_type == "pdf":
            return _build_pdf_preview_html(file_path)
        if file_type in ["xlsx", "xls"]:
            return _build_xlsx_preview_html(file_path)
    except Exception:
        pass
    return f'<pre class="doc-preview">{html.escape(text[:10000])}</pre>'


def _is_valid_answer(text):
    if not text or len(text.strip()) < 20:
        return False
    short_patterns = ["OK", "ok", "好的", "收到", "明白", "已处理"]
    stripped = text.strip().strip('"')
    if stripped in short_patterns:
        return False
    return True


def _call_ai_qa(question, context, source_titles: Optional[list] = None):
    source_hint = "、".join(source_titles or [])
    try:
        from app.utils.ai_client import ai_client
        enhanced_question = question if not source_hint else f"{question}\n\n优先参考资料：{source_hint}"
        result = ai_client.qa_answer(enhanced_question, context)
        if result and result.get("answer") and _is_valid_answer(result["answer"]) and result["answer"] != "文档中未找到相关信息":
            return {
                "answer": result.get("answer", ""),
                "source": result.get("source", source_hint)
            }
    except Exception:
        pass

    try:
        from app.utils.ai_client import ai_client
        prompt = f"""基于以下文档片段回答问题：

文档片段：
{context[:8000]}

问题：{question}

候选来源：{source_hint or '未提供'}

请按照以下要求回答：
1. 回答必须基于提供的文档片段，禁止编造信息
2. 信息不足时明确说明"文档中未找到相关信息"
3. 回答末尾附上最相关的来源名称或片段编号
4. 严格使用文档中的原始术语，不得自行改写成近义词（如文档写"主机"则必须用"主机"，不能写成"主持人"、"电脑"等）
5. 保持文档中的产品名、型号、参数、单位等专有信息完全不变"""
        messages = [{"role": "user", "content": prompt}]
        proxy_result = ai_client.chat(messages, max_tokens=2048)
        if proxy_result and _is_valid_answer(proxy_result):
            return {"answer": proxy_result, "source": source_hint}
    except Exception:
        pass

    snippet_preview = context[:1200].strip()
    if snippet_preview:
        return {
            "answer": f"AI 引擎暂时不可用，以下是文档中与您问题相关的原始内容片段，供参考：\n\n{snippet_preview}",
            "source": source_hint
        }

    return {"answer": f"根据当前资料，我暂时无法对“{question}”给出高置信度结论。请补充更具体的对象、参数、章节或步骤名称。", "source": source_hint}


def _random_hex():
    return hex(random.getrandbits(32))[2:]


def _parse_json_array(text: str):
    if not text:
        return []

    def _clean(s):
        return ''.join(c for c in s if not unicodedata.category(c).startswith('P')).strip()

    text = text.strip()
    m = re.search(r"\[[\s\S]*\]", text)
    if not m:
        return []
    try:
        arr = json.loads(m.group(0))
        if isinstance(arr, list):
            cleaned = []
            for item in arr:
                s = _clean(str(item))
                if s:
                    cleaned.append(s)
            return cleaned
    except Exception:
        pass
    lines = [line.strip().lstrip("0123456789.、- ") for line in text.split("\n") if line.strip()]
    return [_clean(line) for line in lines if len(line) > 3]


def _generate_suggestions(context: str = "", question: str = "", answer: str = "", max_count: int = 4, timeout: float = 3.0):
    if question and answer:
        prompt = f"""你是一个善于引导对话的助手。根据以下问答内容，以用户的视角生成{max_count}条用户可能想继续追问的问题（本次随机种子：{_random_hex()}）。

用户刚问了：{question}
AI的回答摘要：{answer[:2000]}

要求：
1. 每条问题必须从用户视角出发，使用"我"、"我想知道"、"能帮我"等第一人称语气
2. 问题要具体、自然，与回答中提到的具体知识点、术语、概念紧密关联
3. 问题要有吸引力，让人想继续深入了解
4. 问题尽量多样化，覆盖不同方面
5. 问题文本中不要包含任何标点符号（如引号、逗号、句号、问号、顿号、书名号等），纯文字即可

请直接返回JSON数组：
["用户视角的问题1", "用户视角的问题2"]"""
    else:
        prompt = f"""你是知识库的导读助手。以下是从知识库中抽取的文档片段，请仔细阅读后，找出其中用户最可能关心的具体话题，以用户第一人称生成{max_count}条推荐问题（本次随机种子：{_random_hex()}）。

知识库文档片段：
{context[:6000]}

要求：
1. 问题必须紧密关联文档中提到的具体知识点、术语、概念、规范或产品名称
2. 问题尽量多样化，覆盖不同方面，避免每次生成相同或相似的问题
3. 使用第一人称：我、我想知道、我该如何、为什么我的、能帮我解释一下
4. 语气自然，像真实用户面对知识库时会提出的问题
5. 问题文本中不要包含任何标点符号（如引号、逗号、句号、问号、顿号、书名号、括号等），纯文字即可

请直接返回JSON数组：
["提及具体知识点的用户视角问题1", "提及具体知识点的用户视角问题2", "提及具体知识点的用户视角问题3"]"""

    def _call():
        try:
            from app.utils.ai_client import ai_client
            messages = [{"role": "user", "content": prompt}]
            result = ai_client.chat(messages, max_tokens=512, temperature=0.8)
            parsed = _parse_json_array(result)
            return parsed[:max_count] if parsed else []
        except Exception:
            return []

    try:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_call)
            return future.result(timeout=timeout)
    except FutureTimeoutError:
        return []


# ─── 知识库问答 ────────────────────────────────────────────

@router.post("/general")
async def ask_general_question(
    input_data: GeneralQAInput,
    db: Session = Depends(get_db),
    token: str = Depends(oauth2_scheme),
):
    question = (input_data.question or "").strip()
    knowledge_ids = input_data.knowledge_ids or []
    user_id = _get_user_id_from_token(token, db)

    documents = []

    from app.models.knowledge import KnowledgeFile, Folder
    from app.crud.knowledge import get_folder_files

    def _collect_files_recursive(folder_id):
        files = []
        folder = db.query(Folder).filter(Folder.id == folder_id).first()
        if folder:
            files.extend(get_folder_files(db, folder_id))
            for child in folder.children:
                files.extend(_collect_files_recursive(child.id))
        return files

    if knowledge_ids:
        seen = set()
        for fid in knowledge_ids:
            try:
                all_files = _collect_files_recursive(int(fid))
                for f in all_files:
                    fid_unique = f.get("id")
                    if fid_unique in seen:
                        continue
                    seen.add(fid_unique)
                    file_path = f.get("file_path", "")
                    file_name = f.get("name", "")
                    file_type = f.get("file_type", "")
                    if file_path and os.path.exists(file_path):
                        content = _extract_file_content(file_path, file_type)
                        if _normalize_text(content):
                            documents.append({
                                "document_id": fid_unique,
                                "title": file_name,
                                "content": content,
                            })
            except Exception:
                continue

    if not documents:
        result = {
            "answer": "未选择知识库或所选知识库中没有文档内容。请在左侧选择一个包含文档的知识库。",
            "source": ""
        }
        sources = []
        search_hit = 0
        relevance_score = 0.0
    else:
        ranked_sources = _rank_document_chunks(question, documents)
        top_score = ranked_sources[0]["score"] if ranked_sources else 0.0
        needs_clarify = _needs_clarification(question, ranked_sources)
        search_hit = 1 if (ranked_sources and not needs_clarify) else 0
        relevance_score = round(top_score, 4)
        if needs_clarify:
            result = _build_clarification_answer(question, ranked_sources)
            _, sources = _build_context_from_sources(ranked_sources)
        else:
            context, sources = _build_context_from_sources(ranked_sources)
            result = _call_ai_qa(question, context, [item.get("title") for item in sources])

    response_data = {
        "question": question,
        "answer": result.get("answer", ""),
        "sources": sources,
        "suggestions": _generate_suggestions(question=question, answer=result.get("answer", ""), max_count=3, timeout=10.0)
    }

    session_id = None
    if user_id:
        session_id = _save_qa_history(db, "general", question, result, user_id, input_data.session_id, sources, search_hit, relevance_score)
    response_data["session_id"] = session_id

    return response_data


@router.get("/suggestions/initial")
async def get_initial_suggestions(
    kb_ids: str = Query("", description="已勾选的知识库ID列表，逗号分隔"),
    limit: int = Query(4, description="返回数量，默认4，最大8"),
    db: Session = Depends(get_db),
    token: str = Depends(oauth2_scheme),
):
    limit = min(max(limit, 1), 8)
    kb_id_list = [x.strip() for x in kb_ids.split(",") if x.strip()] if kb_ids else []

    if not kb_id_list:
        return {"code": 0, "data": {"suggestions": [], "refreshable": False}}

    from app.models.knowledge import KnowledgeFile, Folder
    from app.crud.knowledge import get_folder_files

    def _collect_files_recursive(folder_id):
        files = []
        folder = db.query(Folder).filter(Folder.id == folder_id).first()
        if folder:
            files.extend(get_folder_files(db, folder_id))
            for child in folder.children:
                files.extend(_collect_files_recursive(child.id))
        return files

    documents = []
    seen = set()
    for fid_str in kb_id_list:
        try:
            all_files = _collect_files_recursive(int(fid_str))
            for f in all_files:
                fid_unique = f.get("id")
                if fid_unique in seen:
                    continue
                seen.add(fid_unique)
                file_path = f.get("file_path", "")
                file_name = f.get("name", "")
                file_type = f.get("file_type", "")
                if file_path and os.path.exists(file_path):
                    content = _extract_file_content(file_path, file_type)
                    if _normalize_text(content):
                        documents.append({
                            "document_id": fid_unique,
                            "title": file_name,
                            "content": content,
                        })
        except Exception:
            continue

    if not documents:
        return {"code": 0, "data": {"suggestions": [], "refreshable": False}}

    context_parts = []
    total_len = 0
    doc_names = []
    for doc in documents:
        title = doc.get("title", "")
        if title:
            doc_names.append(title)
        snippet = doc.get("content", "")[:1500]
        part = f"### 文档：{title}\n{snippet}\n"
        if total_len + len(part) > 8000:
            context_parts.append(part[:8000 - total_len])
            break
        context_parts.append(part)
        total_len += len(part)

    doc_list_str = "\n".join([f"- {n}" for n in doc_names[:10]])
    context = f"已选知识库包含以下文档：\n{doc_list_str}\n\n详细内容：\n" + "\n".join(context_parts)
    suggestions = _generate_suggestions(context=context, max_count=limit, timeout=15.0)

    return {
        "code": 0,
        "data": {
            "suggestions": suggestions,
            "refreshable": len(suggestions) > 0
        }
    }


# ─── 文档对话 ────────────────────────────────────────────

@router.post("/docs/chat")
async def ask_document_question(
    question: str = Form(...),
    files: list[UploadFile] = File(...),
    session_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    token: str = Depends(oauth2_scheme),
):
    if not os.path.exists(UPLOAD_DIR):
        os.makedirs(UPLOAD_DIR)

    documents = []
    file_names = []

    for file in files:
        if not file.filename:
            continue
        ext = os.path.splitext(file.filename)[1] if file.filename else ""
        safe_name = f"{uuid.uuid4()}{ext}"
        file_path = os.path.join(UPLOAD_DIR, safe_name)

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        file_type = ext[1:] if ext else "unknown"
        text = _extract_file_content(file_path, file_type)
        if _normalize_text(text):
            documents.append({
                "document_id": safe_name,
                "title": file.filename,
                "content": text,
            })
            file_names.append(file.filename)

        try:
            os.remove(file_path)
        except Exception:
            pass

    if not documents:
        return {"question": question, "answer": "未能从上传的文件中提取到文字内容，请确认文件格式。", "sources": [], "session_id": None}

    ranked_sources = _rank_document_chunks(question, documents)
    top_score = ranked_sources[0]["score"] if ranked_sources else 0.0
    needs_clarify = _needs_clarification(question, ranked_sources)
    search_hit = 1 if (ranked_sources and not needs_clarify) else 0
    relevance_score = round(top_score, 4)
    if needs_clarify:
        result = _build_clarification_answer(question, ranked_sources)
        _, sources = _build_context_from_sources(ranked_sources)
    else:
        context, sources = _build_context_from_sources(ranked_sources)
        result = _call_ai_qa(question, context, [item.get("title") for item in sources])

    user_id = _get_user_id_from_token(token, db)

    new_session_id = None
    if user_id:
        new_session_id = _save_qa_history(db, "doc", question, result, user_id, session_id, sources, search_hit, relevance_score)

    return {
        "question": question,
        "answer": result.get("answer", ""),
        "files": file_names,
        "sources": sources,
        "session_id": new_session_id
    }


# ─── 文档预览 ──────────────────────────────────────────────

@router.post("/docs/preview")
async def preview_document_content(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(400, "未选择文件")

    if not os.path.exists(UPLOAD_DIR):
        os.makedirs(UPLOAD_DIR)

    ext = os.path.splitext(file.filename)[1] if file.filename else ""
    safe_name = f"{uuid.uuid4()}{ext}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)

    content_bytes = await file.read()
    with open(file_path, "wb") as f:
        f.write(content_bytes)

    file_type = ext[1:] if ext else "unknown"
    try:
        text = _extract_file_content(file_path, file_type)
        preview_html = _build_preview_html(file_path, file_type, text)
    except Exception:
        text = "无法解析该文件，可能是格式不被支持或文件损坏。"
        preview_html = f'<pre class="doc-preview">{html.escape(text)}</pre>'

    try:
        os.remove(file_path)
    except Exception:
        pass

    return {
        "filename": file.filename,
        "content": text[:10000],
        "truncated": len(text) > 10000,
        "images": [],
        "preview_html": preview_html,
    }


# ─── 文档对话首次上传 ─────────────────────────────────────

@router.post("/docs/upload")
async def upload_docs_for_chat(files: list[UploadFile] = File(...)):
    if not os.path.exists(UPLOAD_DIR):
        os.makedirs(UPLOAD_DIR)

    results = []
    for file in files:
        if not file.filename:
            continue
        ext = os.path.splitext(file.filename)[1] if file.filename else ""
        safe_name = f"{uuid.uuid4()}{ext}"
        file_path = os.path.join(UPLOAD_DIR, safe_name)

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        file_type = ext[1:] if ext else "unknown"
        text = _extract_file_content(file_path, file_type)

        results.append({
            "temp_id": safe_name,
            "name": file.filename,
            "file_type": file_type,
            "content": text,
        })

    return {"files": results}


# ─── 反馈系统 ────────────────────────────────────────────

@router.post("/feedback")
async def submit_feedback(
    feedback: QaFeedbackCreate,
    db: Session = Depends(get_db),
):
    fb = QaFeedback(
        question=feedback.question,
        answer=feedback.answer,
        rating=feedback.rating,
        feedback_text=feedback.feedback_text,
    )
    db.add(fb)
    db.commit()
    db.refresh(fb)
    return {"message": "反馈已提交", "id": fb.id}


@router.get("/feedback")
async def get_feedbacks(
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可查看反馈列表")
    total = db.query(QaFeedback).count()
    items = db.query(QaFeedback).order_by(QaFeedback.created_at.desc()).offset(skip).limit(limit).all()
    return {
        "total": total,
        "items": [
            {
                "id": item.id,
                "question": item.question,
                "answer": item.answer,
                "rating": item.rating,
                "feedback_text": item.feedback_text,
                "resolved": item.resolved,
                "created_at": _to_beijing_iso(item.created_at),
            }
            for item in items
        ]
    }


@router.get("/feedback/unread-count")
async def get_unread_feedback_count(
    db: Session = Depends(get_db),
    token: str = Depends(oauth2_scheme),
):
    from jose import jwt, JWTError
    from app.api.auth import SECRET_KEY, ALGORITHM
    if not token:
        return {"count": 0}
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if not username:
            return {"count": 0}
    except JWTError:
        return {"count": 0}

    from app.crud.user import get_user as crud_get_user
    user = crud_get_user(db, username=username)
    if not user or user.role != "admin":
        return {"count": 0}

    count = db.query(QaFeedback).filter(QaFeedback.resolved == False).count()
    return {"count": count}


@router.put("/feedback/{feedback_id}/resolve")
async def resolve_feedback(
    feedback_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    fb = db.query(QaFeedback).filter(QaFeedback.id == feedback_id).first()
    if not fb:
        raise HTTPException(status_code=404, detail="反馈不存在")
    fb.resolved = True
    db.commit()
    return {"message": "反馈已标记为已处理"}


# ─── 历史记录 ─────────────────────────────────────────────

@router.get("/history/sessions")
async def get_qa_sessions(
    type: str = Query("all"),
    db: Session = Depends(get_db),
    token: str = Depends(oauth2_scheme),
):
    user_id = _get_user_id_from_token(token, db)
    if not user_id:
        return {"sessions": []}

    q = db.query(QaSession).filter(QaSession.user_id == user_id)
    if type == "general":
        q = q.filter(QaSession.session_type == "general")
    elif type == "doc":
        q = q.filter(QaSession.session_type.in_(["doc", "manual"]))
    sessions = q.order_by(QaSession.updated_at.desc()).all()

    return {
        "sessions": [
            {
                "id": s.id,
                "session_type": s.session_type,
                "title": s.title,
                "created_at": _to_beijing_iso(s.created_at),
                "updated_at": _to_beijing_iso(s.updated_at),
                "message_count": len(s.messages) if s.messages else 0,
            }
            for s in sessions
        ]
    }


@router.get("/history/sessions/{session_id}")
async def get_qa_session_detail(
    session_id: int,
    db: Session = Depends(get_db),
    token: str = Depends(oauth2_scheme),
):
    user_id = _get_user_id_from_token(token, db)
    if not user_id:
        raise HTTPException(status_code=401, detail="请先登录")

    sess = db.query(QaSession).filter(QaSession.id == session_id, QaSession.user_id == user_id).first()
    if not sess:
        raise HTTPException(status_code=404, detail="会话不存在")

    messages = []
    for m in sess.messages:
        sources = []
        try:
            sources = json.loads(m.sources or "[]")
        except Exception:
            sources = []
        messages.append({
            "id": m.id,
            "role": m.role,
            "content": m.content,
            "sources": sources,
            "rating": m.rating,
            "created_at": _to_beijing_iso(m.created_at),
        })

    return {
        "session": {
            "id": sess.id,
            "session_type": sess.session_type,
            "title": sess.title,
            "created_at": _to_beijing_iso(sess.created_at),
            "updated_at": _to_beijing_iso(sess.updated_at),
        },
        "messages": messages,
    }


@router.delete("/history/sessions/{session_id}")
async def delete_qa_session(
    session_id: int,
    db: Session = Depends(get_db),
    token: str = Depends(oauth2_scheme),
):
    user_id = _get_user_id_from_token(token, db)
    if not user_id:
        raise HTTPException(status_code=401, detail="请先登录")

    sess = db.query(QaSession).filter(QaSession.id == session_id, QaSession.user_id == user_id).first()
    if not sess:
        raise HTTPException(status_code=404, detail="会话不存在")

    db.delete(sess)
    db.commit()
    return {"message": "已删除"}


# ─── 问答看板（管理员） ──────────────────────────────

@router.get("/dashboard")
async def get_qa_dashboard(
    start_date: str = Query(None, description="起始日期 YYYY-MM-DD"),
    end_date: str = Query(None, description="结束日期 YYYY-MM-DD"),
    period: str = Query("yesterday", description="快捷时间范围: today, yesterday, this_week, this_month, last_7_days, last_30_days"),
    user_name: str = Query(None, description="筛选用户"),
    session_type: str = Query(None, description="筛选场域"),
    rating: int = Query(None, description="筛选评分"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可访问")

    now = datetime.now(BEIJING_TZ)
    today = now.date()

    PERIOD_LABELS = {
        "today": "今日",
        "yesterday": "昨日",
        "this_week": "本周",
        "this_month": "本月",
        "last_7_days": "近7天",
        "last_30_days": "近30天",
    }
    COMPARE_LABELS = {
        "today": "较昨日",
        "yesterday": "较前一日",
        "this_week": "较上周",
        "this_month": "较上月",
        "last_7_days": "较前7天",
        "last_30_days": "较前30天",
    }

    def _period_range(p: str):
        if p == "today":
            s = datetime(today.year, today.month, today.day, tzinfo=BEIJING_TZ)
            e = now
            cs = s - timedelta(days=1)
            ce = s - timedelta(seconds=1)
            return s, e, cs, ce
        elif p == "yesterday":
            s = datetime(today.year, today.month, today.day, tzinfo=BEIJING_TZ) - timedelta(days=1)
            e = datetime(today.year, today.month, today.day, tzinfo=BEIJING_TZ) - timedelta(seconds=1)
            cs = s - timedelta(days=1)
            ce = s - timedelta(seconds=1)
            return s, e, cs, ce
        elif p == "this_week":
            weekday = today.weekday()
            s = datetime(today.year, today.month, today.day, tzinfo=BEIJING_TZ) - timedelta(days=weekday)
            e = now
            cs = s - timedelta(days=7)
            ce = s - timedelta(seconds=1)
            return s, e, cs, ce
        elif p == "this_month":
            s = datetime(today.year, today.month, 1, tzinfo=BEIJING_TZ)
            e = now
            if today.month == 1:
                cs = datetime(today.year - 1, 12, 1, tzinfo=BEIJING_TZ)
            else:
                cs = datetime(today.year, today.month - 1, 1, tzinfo=BEIJING_TZ)
            ce = datetime(today.year, today.month, 1, tzinfo=BEIJING_TZ) - timedelta(seconds=1)
            return s, e, cs, ce
        elif p == "last_7_days":
            s = datetime(today.year, today.month, today.day, tzinfo=BEIJING_TZ) - timedelta(days=6)
            e = now
            cs = s - timedelta(days=7)
            ce = s - timedelta(seconds=1)
            return s, e, cs, ce
        elif p == "last_30_days":
            s = datetime(today.year, today.month, today.day, tzinfo=BEIJING_TZ) - timedelta(days=29)
            e = now
            cs = s - timedelta(days=30)
            ce = s - timedelta(seconds=1)
            return s, e, cs, ce
        else:
            s = datetime(today.year, today.month, today.day, tzinfo=BEIJING_TZ) - timedelta(days=1)
            e = datetime(today.year, today.month, today.day, tzinfo=BEIJING_TZ) - timedelta(seconds=1)
            cs = s - timedelta(days=1)
            ce = s - timedelta(seconds=1)
            return s, e, cs, ce

    period_label = PERIOD_LABELS.get(period, "昨日")
    compare_label = COMPARE_LABELS.get(period, "较前一日")
    overview_start, overview_end, compare_start, compare_end = _period_range(period)

    def _count_active_users(s, e):
        return db.query(func.count(func.distinct(QaSession.user_id))).filter(
            QaSession.created_at >= s, QaSession.created_at <= e,
        ).scalar() or 0

    def _count_conversations(s, e):
        return db.query(func.count(QaSession.id)).filter(
            QaSession.created_at >= s, QaSession.created_at <= e,
        ).scalar() or 0

    def _get_hit_rate(s, e):
        total = db.query(func.count(QaMessage.id)).filter(
            QaMessage.role == "assistant",
            QaMessage.created_at >= s, QaMessage.created_at <= e,
        ).scalar() or 0
        hits = db.query(func.count(QaMessage.id)).filter(
            QaMessage.role == "assistant",
            QaMessage.search_hit == 1,
            QaMessage.created_at >= s, QaMessage.created_at <= e,
        ).scalar() or 0
        return round(hits / total * 100, 1) if total > 0 else 0.0

    active_users = _count_active_users(overview_start, overview_end)
    conversations = _count_conversations(overview_start, overview_end)
    compare_active_users = _count_active_users(compare_start, compare_end)
    compare_conversations = _count_conversations(compare_start, compare_end)
    hit_rate = _get_hit_rate(overview_start, overview_end)
    compare_hit_rate = _get_hit_rate(compare_start, compare_end)

    # ── 图表数据 ──
    if start_date and end_date:
        try:
            sd = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=BEIJING_TZ)
            ed = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=BEIJING_TZ)
        except ValueError:
            raise HTTPException(status_code=400, detail="日期格式错误，应为 YYYY-MM-DD")
    else:
        ed = datetime.now(BEIJING_TZ)
        sd = ed - timedelta(days=29)

    daily_users = db.query(
        func.date(QaSession.created_at).label("d"),
        func.count(func.distinct(QaSession.user_id)).label("c"),
    ).filter(
        QaSession.created_at >= sd, QaSession.created_at <= ed,
    ).group_by("d").order_by("d").all()
    active_users_chart = [{"date": str(row.d), "count": row.c} for row in daily_users]

    daily_msgs = db.query(
        func.date(QaMessage.created_at).label("d"),
        func.count(QaMessage.id).label("c"),
    ).filter(
        QaMessage.role == "user",
        QaMessage.created_at >= sd, QaMessage.created_at <= ed,
    ).group_by("d").order_by("d").all()
    conversations_chart = [{"date": str(row.d), "count": row.c} for row in daily_msgs]

    daily_hits = db.query(
        func.date(QaMessage.created_at).label("d"),
        func.count(QaMessage.id).label("total"),
        func.sum(QaMessage.search_hit).label("hits"),
    ).filter(
        QaMessage.role == "assistant",
        QaMessage.created_at >= sd, QaMessage.created_at <= ed,
    ).group_by("d").order_by("d").all()
    hit_rate_chart = []
    for row in daily_hits:
        rate = round((row.hits or 0) / row.total * 100, 1) if row.total > 0 else 0.0
        hit_rate_chart.append({"date": str(row.d), "rate": rate})

    daily_type_hits = db.query(
        func.date(QaMessage.created_at).label("d"),
        QaSession.session_type,
        func.count(QaMessage.id).label("total"),
        func.sum(QaMessage.search_hit).label("hits"),
    ).join(QaSession, QaMessage.session_id == QaSession.id).filter(
        QaMessage.role == "assistant",
        QaMessage.created_at >= sd, QaMessage.created_at <= ed,
    ).group_by("d", QaSession.session_type).order_by("d").all()

    def _build_type_hit_chart(daily_rows, stype):
        result = []
        idx = 0
        for row in daily_hits:
            d = str(row.d)
            rate = 0.0
            while idx < len(daily_rows) and str(daily_rows[idx].d) < d:
                idx += 1
            if idx < len(daily_rows) and str(daily_rows[idx].d) == d and daily_rows[idx].session_type == stype:
                r = daily_rows[idx]
                rate = round((r.hits or 0) / r.total * 100, 1) if r.total > 0 else 0.0
                idx += 1
            result.append({"date": d, "rate": rate})
        return result

    general_hit_rate_chart = _build_type_hit_chart(daily_type_hits, "general")
    manual_hit_rate_chart = _build_type_hit_chart(daily_type_hits, "manual")

    # ── 明细列表 ──
    detail_query = db.query(
        QaMessage.id, QaMessage.session_id, QaMessage.content,
        QaMessage.sources, QaMessage.rating, QaMessage.created_at,
        QaSession.session_type, QaSession.title,
    ).join(QaSession, QaMessage.session_id == QaSession.id).filter(
        QaMessage.role == "user"
    )

    user_map = {}
    if user_name:
        from app.models.user import User
        detail_query = detail_query.join(User, QaSession.user_id == User.id)
        detail_query = detail_query.filter(
            (User.username.contains(user_name)) | (User.display_name.contains(user_name))
        )
        users = db.query(User.id, User.username, User.display_name).all()
        user_map = {u.id: (u.display_name or u.username) for u in users}
    else:
        from app.models.user import User
        users = db.query(User.id, User.username, User.display_name).all()
        user_map = {u.id: (u.display_name or u.username) for u in users}

    if session_type:
        detail_query = detail_query.filter(QaSession.session_type == session_type)
    if rating is not None:
        detail_query = detail_query.filter(QaMessage.rating == rating)
    if start_date and end_date:
        detail_query = detail_query.filter(QaMessage.created_at >= sd, QaMessage.created_at <= ed)

    total = detail_query.count()
    records = detail_query.order_by(QaMessage.created_at.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    session_ids = list(set(r.session_id for r in records))
    session_user_map = {}
    if session_ids:
        sessions_with_user = db.query(QaSession.id, QaSession.user_id).filter(
            QaSession.id.in_(session_ids)
        ).all()
        session_user_map = {s.id: s.user_id for s in sessions_with_user}

    ai_replies = {}
    if session_ids:
        replies = db.query(
            QaMessage.session_id, QaMessage.id, QaMessage.content, QaMessage.created_at, QaMessage.search_hit
        ).filter(
            QaMessage.role == "assistant",
            QaMessage.session_id.in_(session_ids),
        ).order_by(QaMessage.created_at.asc()).all()
        session_replies = {}
        for r in replies:
            session_replies.setdefault(r.session_id, []).append(r)
        for rec in records:
            session_reps = session_replies.get(rec.session_id, [])
            best_reply = None
            for rep in session_reps:
                if rep.created_at > rec.created_at:
                    best_reply = rep
                    break
            if best_reply:
                ai_replies[rec.id] = best_reply

    items = []
    for r in records:
        uid = session_user_map.get(r.session_id)
        user_name_display = user_map.get(uid, "未知")
        reply = ai_replies.get(r.id)
        answer_text = reply.content if reply else ""
        search_hit_val = bool(reply.search_hit) if reply else False
        success = bool(answer_text and "未检索到" not in answer_text and "定位到" not in answer_text)
        items.append({
            "id": r.id,
            "session_id": r.session_id,
            "user_name": user_name_display,
            "session_type": r.session_type or "general",
            "session_title": r.title or "",
            "question": r.content,
            "answer": answer_text,
            "sources": r.sources or "[]",
            "rating": r.rating,
            "success": success,
            "search_hit": search_hit_val,
            "created_at": _to_beijing_iso(r.created_at),
        })

    return {
        "overview": {
            "active_users": active_users,
            "conversations": conversations,
            "active_users_trend": active_users - compare_active_users,
            "conversations_trend": conversations - compare_conversations,
            "hit_rate": hit_rate,
            "hit_rate_trend": round(hit_rate - compare_hit_rate, 1),
            "period_label": period_label,
            "compare_label": compare_label,
        },
        "charts": {
            "active_users": active_users_chart,
            "conversations": conversations_chart,
            "hit_rate": hit_rate_chart,
            "general_hit_rate": general_hit_rate_chart,
            "manual_hit_rate": manual_hit_rate_chart,
        },
        "period": period,
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }
