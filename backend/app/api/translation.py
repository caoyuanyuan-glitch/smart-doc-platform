from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Form, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
import os
import io
import csv
import json
import zipfile
import tempfile
import sys
import threading
import xml.etree.ElementTree as ET
import lxml.etree as ET_LXML
import re
import unicodedata
from difflib import SequenceMatcher

from app.database import get_db, SessionLocal
from app.schemas.translation import (
    TranslationRequest,
    TranslationResponse,
    MemoryEntry,
    MemoryEntryOut,
    MemoryFileEntryRequest,
    TranslationDocOut,
)
from app.crud.document import get_document, get_documents
from app.crud.review import get_reviews
from app.crud.memory import (
    create_memory_entry, get_memory_entries, get_memory_entry,
    delete_memory_entry, get_memory_banks
)
from app.models.knowledge import KnowledgeFile
from app.models.memory import MemoryBank
from app.models.translation_doc import TranslationDoc
from app.utils.document_parser import parse_file, parse_xlsx_textual_content
from app.utils.ai_client import ai_client
from app.utils.file_utils import read_file_safe

router = APIRouter()

UPLOAD_DIR = "./static/uploads/translation"
TRANSLATION_OUTPUT_DIR = "./static/translations"
UNSUPPORTED_TRANSLATION_EXTENSIONS = {".dita", ".zip"}
WRITABLE_MEMORY_FILE_TYPES = {"xlsx", "xlsm", "xltx", "xltm"}

_translate_tasks = {}
_translate_tasks_lock = threading.Lock()
_thread_locals = threading.local()
_memory_file_cache = {}
_memory_file_cache_lock = threading.Lock()
SUPPORTED_AI_MODELS = {"kimi", "deepseek", "arkclaw", "mcai", "proxy"}
WORD_CONNECTORS = {"-", "_", ".", "/", ":", "+", "'", "’", "%", "#", "&"}


def _read_text_file_with_fallback(file_path: str) -> str:
    return read_file_safe(file_path)


def _normalize_ai_model(model: str) -> str:
    normalized = (model or "").strip().lower()
    if normalized in SUPPORTED_AI_MODELS:
        return normalized
    resolved = ai_client.resolve_translation_model(normalized)
    return resolved or "kimi"


def _reset_translation_usage_stats():
    _thread_locals.translation_usage_stats = {
        "ai_char_count": 0,
        "memory_char_count": 0,
        "ai_word_count": 0,
        "memory_word_count": 0,
    }


def _get_translation_usage_stats():
    stats = getattr(_thread_locals, "translation_usage_stats", None)
    if stats is None:
        stats = {
            "ai_char_count": 0,
            "memory_char_count": 0,
            "ai_word_count": 0,
            "memory_word_count": 0,
        }
        _thread_locals.translation_usage_stats = stats
    return stats


def _is_cjk_char(char: str) -> bool:
    return bool(re.match(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]", char))


def _is_word_core_char(char: str) -> bool:
    if not char or char.isspace() or _is_cjk_char(char):
        return False
    category = unicodedata.category(char)
    return category.startswith("L") or category.startswith("N")


def _consume_word_token(text: str, start: int) -> int:
    index = start
    while index < len(text):
        current = text[index]
        if _is_word_core_char(current):
            index += 1
            continue
        if current in WORD_CONNECTORS:
            prev_char = text[index - 1] if index - 1 >= start else ""
            next_char = text[index + 1] if index + 1 < len(text) else ""
            if _is_word_core_char(prev_char) and _is_word_core_char(next_char):
                index += 1
                continue
        break
    return index


def _count_text_units(text: str) -> int:
    normalized = unicodedata.normalize("NFKC", str(text or ""))
    if not normalized.strip():
        return 0

    count = 0
    index = 0
    while index < len(normalized):
        char = normalized[index]
        if char.isspace():
            index += 1
            continue
        if _is_cjk_char(char):
            count += 1
            index += 1
            continue
        if _is_word_core_char(char):
            index = _consume_word_token(normalized, index)
            count += 1
            continue
        index += 1
    return count


def _record_translation_usage(source: str, text: str):
    if source not in {"ai", "memory"}:
        return
    if text is None:
        return
    stats = _get_translation_usage_stats()
    stats[f"{source}_char_count"] += len(str(text))
    stats[f"{source}_word_count"] += _count_text_units(text)


def _record_passthrough_usage(text: str):
    _record_translation_usage("memory", text)


def _snapshot_translation_usage_stats():
    stats = _get_translation_usage_stats()
    return {
        "ai_char_count": int(stats.get("ai_char_count") or 0),
        "memory_char_count": int(stats.get("memory_char_count") or 0),
        "ai_word_count": int(stats.get("ai_word_count") or 0),
        "memory_word_count": int(stats.get("memory_word_count") or 0),
    }


def _normalize_usage_counts(source_count: int, ai_count: int, memory_count: int):
    source_count = max(0, int(source_count or 0))
    ai_count = max(0, int(ai_count or 0))
    memory_count = max(0, int(memory_count or 0))
    ai_count = min(ai_count, source_count)
    uncategorized_count = max(0, source_count - ai_count - memory_count)
    memory_count = min(source_count - ai_count, memory_count + uncategorized_count)
    return {
        "source_count": source_count,
        "ai_count": ai_count,
        "memory_count": memory_count,
    }


def _normalize_doc_char_counts(doc):
    normalized = _normalize_usage_counts(
        getattr(doc, "source_char_count", 0),
        getattr(doc, "ai_char_count", 0),
        getattr(doc, "memory_char_count", 0),
    )
    return {
        "source_char_count": normalized["source_count"],
        "ai_char_count": normalized["ai_count"],
        "memory_char_count": normalized["memory_count"],
    }


def _load_doc_source_text(doc) -> str:
    if getattr(doc, "file_type", "") == "text":
        return getattr(doc, "original_content", "") or ""
    filename = (getattr(doc, "filename", "") or "").strip()
    if not filename:
        return getattr(doc, "original_content", "") or ""
    file_path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(file_path):
        return getattr(doc, "original_content", "") or ""
    ext = os.path.splitext(filename)[1].lower()
    try:
        if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return parse_xlsx_textual_content(file_path)
        if ext == ".txt":
            return _read_text_file_with_fallback(file_path)
        return parse_file(file_path)
    except Exception:
        return getattr(doc, "original_content", "") or ""


def _normalize_doc_word_counts(doc):
    stored_source = max(0, int(getattr(doc, "source_word_count", 0) or 0))
    stored_ai = max(0, int(getattr(doc, "ai_word_count", 0) or 0))
    stored_memory = max(0, int(getattr(doc, "memory_word_count", 0) or 0))
    if stored_source or stored_ai or stored_memory:
        normalized = _normalize_usage_counts(stored_source, stored_ai, stored_memory)
        return {
            "source_word_count": normalized["source_count"],
            "ai_word_count": normalized["ai_count"],
            "memory_word_count": normalized["memory_count"],
            "dirty": stored_source != normalized["source_count"] or stored_ai != normalized["ai_count"] or stored_memory != normalized["memory_count"],
        }

    source_text = _load_doc_source_text(doc)
    source_word_count = _count_text_units(source_text)
    legacy = _normalize_doc_char_counts(doc)
    ai_word_count = 0
    if source_word_count > 0 and legacy["source_char_count"] > 0:
        ai_word_count = min(source_word_count, round(source_word_count * legacy["ai_char_count"] / legacy["source_char_count"]))
    memory_word_count = max(0, source_word_count - ai_word_count)
    normalized = _normalize_usage_counts(source_word_count, ai_word_count, memory_word_count)
    return {
        "source_word_count": normalized["source_count"],
        "ai_word_count": normalized["ai_count"],
        "memory_word_count": normalized["memory_count"],
        "dirty": any([normalized["source_count"], normalized["ai_count"], normalized["memory_count"]]),
    }


def _normalize_doc_usage_counts(doc):
    source_char_count = max(0, int(getattr(doc, "source_char_count", 0) or 0))
    ai_char_count = max(0, int(getattr(doc, "ai_char_count", 0) or 0))
    memory_char_count = max(0, int(getattr(doc, "memory_char_count", 0) or 0))

    ai_char_count = min(ai_char_count, source_char_count)
    uncategorized_char_count = max(0, source_char_count - ai_char_count - memory_char_count)
    memory_char_count = min(source_char_count - ai_char_count, memory_char_count + uncategorized_char_count)

    return {
        "source_char_count": source_char_count,
        "ai_char_count": ai_char_count,
        "memory_char_count": memory_char_count,
    }


def _summarize_docs(docs):
    normalized_usage = [_normalize_doc_word_counts(doc) for doc in docs]
    return {
        "doc_count": len(docs),
        "doc_word_count": sum(item["source_word_count"] for item in normalized_usage),
        "ai_word_count": sum(item["ai_word_count"] for item in normalized_usage),
        "memory_word_count": sum(item["memory_word_count"] for item in normalized_usage),
    }


def _clone_zipinfo(zinfo: zipfile.ZipInfo) -> zipfile.ZipInfo:
    cloned = zipfile.ZipInfo(filename=zinfo.filename, date_time=zinfo.date_time)
    cloned.compress_type = zinfo.compress_type
    cloned.comment = zinfo.comment
    cloned.extra = zinfo.extra
    cloned.create_system = zinfo.create_system
    cloned.create_version = zinfo.create_version
    cloned.extract_version = zinfo.extract_version
    cloned.flag_bits = zinfo.flag_bits
    cloned.volume = zinfo.volume
    cloned.internal_attr = zinfo.internal_attr
    cloned.external_attr = zinfo.external_attr
    return cloned


def _normalize_match_text(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t\u00a0]+", " ", text)
    text = re.sub(r"\s*\n\s*", "\n", text)
    return text.strip().lower()


def _normalize_compact_text(value: str) -> str:
    return re.sub(r"[\s\-_,.;:!?()\[\]{}<>/\\|，。；：！？（）【】《》、·•*×]+", "", _normalize_match_text(value))


def _text_matches_source_language(text: str, source_lang: str) -> bool:
    normalized = (text or "").strip()
    if not normalized:
        return False

    lang = (source_lang or "").lower()
    if lang == "zh":
        return bool(re.search(r"[\u4e00-\u9fff]", normalized))
    if lang == "ja":
        return bool(re.search(r"[\u3040-\u30ff\u4e00-\u9fff]", normalized))
    if lang == "ko":
        return bool(re.search(r"[\uac00-\ud7af]", normalized))
    if lang in {"en", "fr", "de", "es", "ru"}:
        return bool(re.search(r"[A-Za-zÀ-ÖØ-öø-ÿА-Яа-я]", normalized))
    return bool(re.search(r"\w", normalized, re.UNICODE))


def _normalize_memory_lookup_key(value: str) -> str:
    text = _normalize_compact_text(value)
    text = re.sub(r"^[\(（\[]?[0-9一二三四五六七八九十]+[\)）\]]?[、.．:：-]*", "", text)
    return text


def _memory_unit_kind(value: str) -> str:
    text = (value or "").strip()
    compact = _normalize_memory_lookup_key(text)
    if not compact:
        return "empty"
    if re.search(r"[\n。！？!?；;]", text):
        return "sentence"
    if re.search(r"[，,：:]", text) and len(compact) >= 8:
        return "sentence"
    if len(compact) >= 12:
        return "sentence"
    return "term"


def _memory_candidate_eligible(source_text: str, candidate_text: str) -> bool:
    source_compact = _normalize_memory_lookup_key(source_text)
    candidate_compact = _normalize_memory_lookup_key(candidate_text)
    if not source_compact or not candidate_compact:
        return False
    if source_compact == candidate_compact:
        return True

    source_kind = _memory_unit_kind(source_text)
    candidate_kind = _memory_unit_kind(candidate_text)
    if source_kind != candidate_kind:
        return False

    source_len = len(source_compact)
    candidate_len = len(candidate_compact)
    shorter = min(source_len, candidate_len)
    longer = max(source_len, candidate_len)
    length_ratio = shorter / longer if longer else 0

    if source_kind == "term":
        return length_ratio >= 0.75 and longer <= max(shorter + 4, shorter * 2)
    return length_ratio >= 0.65


def _memory_similarity(source_text: str, candidate_text: str) -> float:
    normalized_source = _normalize_match_text(source_text)
    normalized_candidate = _normalize_match_text(candidate_text)
    compact_source = _normalize_memory_lookup_key(source_text)
    compact_candidate = _normalize_memory_lookup_key(candidate_text)
    if not normalized_source or not normalized_candidate:
        return 0.0
    if normalized_source == normalized_candidate or compact_source == compact_candidate:
        return 1.0

    ratio = SequenceMatcher(None, normalized_source, normalized_candidate).ratio()
    compact_ratio = SequenceMatcher(None, compact_source, compact_candidate).ratio() if compact_source and compact_candidate else 0.0
    if compact_source and compact_candidate and (compact_source in compact_candidate or compact_candidate in compact_source):
        contain_ratio = min(len(compact_source), len(compact_candidate)) / max(len(compact_source), len(compact_candidate))
        compact_ratio = max(compact_ratio, contain_ratio)
    return max(ratio, compact_ratio)


def _memory_char_match_score(source_text: str, candidate_text: str) -> float:
    compact_source = _normalize_memory_lookup_key(source_text)
    compact_candidate = _normalize_memory_lookup_key(candidate_text)
    if not compact_source or not compact_candidate:
        return 0.0
    matched_chars = sum(block.size for block in SequenceMatcher(None, compact_source, compact_candidate).get_matching_blocks())
    return matched_chars / len(compact_source)


def _apply_memory_translation_preserving_unmatched(source_text: str, candidate_source: str, translated_text: str) -> str:
    source = (source_text or "").strip()
    candidate = (candidate_source or "").strip()
    if not source or not candidate:
        return translated_text

    if source == candidate or _normalize_memory_lookup_key(source) == _normalize_memory_lookup_key(candidate):
        return translated_text
    if candidate in source:
        return source.replace(candidate, translated_text, 1)
    if source in candidate:
        return translated_text

    pieces = []
    inserted_translation = False
    for tag, source_start, source_end, _, _ in SequenceMatcher(None, source, candidate).get_opcodes():
        if tag == "equal":
            if not inserted_translation:
                pieces.append(translated_text)
                inserted_translation = True
            continue
        if tag in {"insert", "replace"}:
            pieces.append(source[source_start:source_end])

    if not inserted_translation:
        pieces.insert(0, translated_text)
    return "".join(pieces) or translated_text


def _match_memory_candidates(source_text: str, candidates, threshold: float = 0.8, preserve_sentence_unmatched: bool = True):
    normalized_source = _normalize_match_text(source_text)
    compact_source = _normalize_memory_lookup_key(source_text)
    if not normalized_source:
        return None

    for entry in candidates:
        candidate_source = entry["source_text"]
        if _normalize_match_text(candidate_source) == normalized_source:
            return entry["translated_text"]
        if compact_source and _normalize_memory_lookup_key(candidate_source) == compact_source:
            return entry["translated_text"]

    lines = [line.strip() for line in re.split(r"\n+", source_text or "") if line.strip()]
    if len(lines) > 1:
        translated_lines = []
        for line in lines:
            line_match = _match_memory_candidates(line, candidates, threshold=0.95)
            if not line_match:
                translated_lines = []
                break
            translated_lines.append(line_match)
        if translated_lines and len(translated_lines) == len(lines):
            return "\n".join(translated_lines)

    best_match = None
    best_score = threshold
    for entry in candidates:
        if not _memory_candidate_eligible(source_text, entry["source_text"]):
            continue
        source_kind = _memory_unit_kind(source_text)
        exact_match = (
            _normalize_match_text(entry["source_text"]) == normalized_source
            or (compact_source and _normalize_memory_lookup_key(entry["source_text"]) == compact_source)
        )
        if source_kind == "sentence" and not preserve_sentence_unmatched and not exact_match:
            continue
        score = max(
            _memory_similarity(source_text, entry["source_text"]),
            _memory_char_match_score(source_text, entry["source_text"]),
        )
        if score >= best_score:
            best_score = score
            best_match = _apply_memory_translation_preserving_unmatched(
                source_text,
                entry["source_text"],
                entry["translated_text"],
            )

    return best_match


def _split_text_for_memory_segments(text: str):
    segments = []
    pattern = re.compile(r"([^\n。！？!?；;.:：，,]+[。！？!?；;.:：，,]?|\n+)")
    for piece in pattern.findall(text or ""):
        if piece:
            segments.append(piece)
    return segments or [text]


def _split_text_for_memory_clauses(text: str):
    parts = []
    current = []
    delimiters = set("\n。！？!?；;.:：，,、")
    for char in text or "":
        current.append(char)
        if char in delimiters:
            parts.append("".join(current))
            current = []
    if current:
        parts.append("".join(current))
    return parts or [text]


def _translate_by_memory_segments(source_text: str, candidates, threshold: float = 0.8):
    segments = _split_text_for_memory_segments(source_text)
    translated_segments = []
    matched_count = 0

    for segment in segments:
        if not segment or not segment.strip():
            translated_segments.append(segment)
            continue

        stripped_segment = segment.strip()
        trailing = segment[len(segment.rstrip()):]
        leading_len = len(segment) - len(segment.lstrip())
        leading = segment[:leading_len]

        matched = _match_memory_candidates(stripped_segment, candidates, threshold=threshold)
        if not matched:
            return None, 0

        translated_segments.append(f"{leading}{matched}{trailing}")
        matched_count += 1

    if matched_count == 0:
        return None, 0
    return "".join(translated_segments), matched_count


def _translate_by_memory_clauses(source_text: str, candidates, threshold: float = 0.8):
    clauses = _split_text_for_memory_clauses(source_text)
    translated_clauses = []
    matched_count = 0

    for clause in clauses:
        if not clause or not clause.strip():
            translated_clauses.append(clause)
            continue

        body = clause.strip()
        prefix_match = re.match(r"^[\(（\[]?[0-9一二三四五六七八九十]+[\)）\]]?[、.．:：-]*", body)
        prefix = prefix_match.group(0) if prefix_match else ""
        content = body[len(prefix):].strip() if prefix else body
        suffix = ""
        if content and content[-1] in "。！？!?；;.:：，,、":
            suffix = content[-1]
            content = content[:-1].strip()

        matched = _match_memory_candidates(content, candidates, threshold=threshold) if content else None
        if matched:
            translated_clauses.append(f"{prefix}{matched}{suffix}")
            matched_count += 1
        else:
            translated_clauses.append(clause)

    if matched_count == 0:
        return None, 0
    return "".join(translated_clauses), matched_count


def _prefer_memory_first(engine: str) -> bool:
    return engine == "memory" or (engine == "hybrid" and bool(_get_memory_bank() or _get_memory_file_id()))


def _translate_text_items(texts, engine: str, model: str, source_lang: str, target_lang: str, db: Session):
    if not texts:
        return []

    translated_texts = [None] * len(texts)
    pending_indexes = list(range(len(texts)))

    if _prefer_memory_first(engine):
        allow_partial = engine == "memory"
        next_pending = []
        for index in pending_indexes:
            result, hit = translate_with_memory(
                texts[index],
                source_lang,
                target_lang,
                db,
                bank=_get_memory_bank(),
                memory_file_id=_get_memory_file_id(),
                allow_partial=allow_partial,
            )
            if hit:
                translated_texts[index] = result
                _record_translation_usage("memory", texts[index])
            else:
                next_pending.append(index)
        pending_indexes = next_pending

    if engine == "memory":
        for index in pending_indexes:
            translated_texts[index] = texts[index]
            _record_passthrough_usage(texts[index])
        return translated_texts

    if pending_indexes:
        sep = "\n---DOCSEG---\n"
        batch_size = 5
        for batch_start in range(0, len(pending_indexes), batch_size):
            batch_end = min(batch_start + batch_size, len(pending_indexes))
            batch_indexes = pending_indexes[batch_start:batch_end]
            if len(batch_indexes) == 1:
                translated_texts[batch_indexes[0]] = _do_translate(texts[batch_indexes[0]], engine, model, source_lang, target_lang, db)
            else:
                batch_texts = [texts[i] for i in batch_indexes]
                combined = sep.join(batch_texts)
                translated_combined = _do_translate(combined, engine, model, source_lang, target_lang, db)
                translated_parts = [part.strip() for part in translated_combined.split(sep)]
                if len(translated_parts) != len(batch_texts):
                    for bi, idx in enumerate(batch_indexes):
                        if bi < len(translated_parts):
                            translated_texts[idx] = translated_parts[bi]
                        else:
                            translated_texts[idx] = _do_translate(texts[idx], engine, model, source_lang, target_lang, db)
                else:
                    for bi, idx in enumerate(batch_indexes):
                        translated_texts[idx] = translated_parts[bi]

    return translated_texts


def _cell_looks_non_translatable(text: str, source_lang: str) -> bool:
    stripped = (text or "").strip()
    if not stripped:
        return True
    if re.fullmatch(r"[\d\s.\-_/+()]+", stripped):
        return True
    if source_lang == "zh" and not re.search(r"[\u4e00-\u9fff]", stripped):
        return True
    return False


def _sanitize_excel_sheet_title(title: str) -> str:
    sanitized = re.sub(r"[:\\/?*\[\]]", "_", str(title or "").strip())
    sanitized = re.sub(r"\s+", " ", sanitized).strip().strip("'")
    return (sanitized or "Sheet")[:31]


def _dedupe_excel_sheet_title(title: str, used_titles) -> str:
    if title not in used_titles:
        return title

    base_title = title[:31].rstrip()
    counter = 2
    while True:
        suffix = f" ({counter})"
        candidate = f"{base_title[:31 - len(suffix)].rstrip()}{suffix}"
        if candidate not in used_titles:
            return candidate
        counter += 1


def _translate_excel_header_footer_text(text: str, engine: str, model: str,
                                        source_lang: str, target_lang: str, db: Session) -> str:
    stripped = (text or "").strip()
    if not stripped or _cell_looks_non_translatable(stripped, source_lang):
        _record_passthrough_usage(stripped or text)
        return text

    markers = []

    def protect(match):
        markers.append(match.group(0))
        return f"__XLSHF{len(markers) - 1}__"

    protected = re.sub(r"&&|&\[[^\]]+\]|&[A-Za-z]", protect, stripped)
    translated = _do_translate(protected, engine, model, source_lang, target_lang, db)
    for index, marker in enumerate(markers):
        translated = translated.replace(f"__XLSHF{index}__", marker)
    return translated


def _filename_looks_non_translatable(filename: str, source_lang: str) -> bool:
    stripped = (filename or "").strip()
    if not stripped:
        return True
    if re.search(r"[\u4e00-\u9fff]", stripped):
        return False
    if source_lang == "zh":
        return True
    if source_lang in ("en", "fr", "de", "es"):
        return False
    return True


def _get_memory_bank():
    return getattr(_thread_locals, 'memory_bank', None)


def _set_memory_bank(bank: str):
    _thread_locals.memory_bank = bank if bank else None


def _get_memory_file_id():
    return getattr(_thread_locals, 'memory_file_id', None)


def _set_memory_file_id(file_id):
    _thread_locals.memory_file_id = file_id


def _normalize_header_key(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", str(value or "").strip().lower())


def _parse_memory_text_entries(raw_text: str):
    entries = []
    delimiters = ["\t", "=>", "->", "→", "|", "｜"]
    for raw_line in raw_text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        for delimiter in delimiters:
            if delimiter in line:
                source_text, translated_text = [part.strip() for part in line.split(delimiter, 1)]
                if source_text and translated_text:
                    entries.append({
                        "source_text": source_text,
                        "translated_text": translated_text
                    })
                break
    return entries


def _normalize_memory_entry(row: dict):
    if not isinstance(row, dict):
        return None

    normalized_row = {_normalize_header_key(key): value for key, value in row.items()}
    source_keys = ["source", "sourcetext", "src", "原文", "源文", "text", "zh", "zh-cn", "zhcn", "cn", "中文", "chinese"]
    target_keys = ["target", "translatedtext", "translation", "译文", "targettext", "result", "en", "en-us", "enus", "英文", "english"]
    source_lang_keys = ["sourcelang", "srclang", "源语言"]
    target_lang_keys = ["targetlang", "tgtlang", "目标语言"]

    def pick_value(keys):
        for key in keys:
            normalized_key = _normalize_header_key(key)
            if normalized_key in normalized_row:
                value = normalized_row[normalized_key]
                if value is not None and str(value).strip() != "":
                    return str(value).strip()
        return None

    source_text = pick_value(source_keys)
    translated_text = pick_value(target_keys)
    if not source_text or not translated_text:
        return None

    return {
        "source_text": source_text,
        "translated_text": translated_text,
        "source_lang": pick_value(source_lang_keys),
        "target_lang": pick_value(target_lang_keys),
    }


def _load_memory_entries_from_json(file_path: str):
    data = json.loads(_read_text_file_with_fallback(file_path))

    if isinstance(data, dict):
        for key in ["entries", "items", "data", "records"]:
            if isinstance(data.get(key), list):
                data = data[key]
                break

    if not isinstance(data, list):
        return []

    entries = []
    for item in data:
        normalized = _normalize_memory_entry(item)
        if normalized:
            entries.append(normalized)
    return entries


def _load_memory_entries_from_delimited_file(file_path: str, delimiter: str = None):
    content = _read_text_file_with_fallback(file_path)
    sample = content[:2048]
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            delimiter = dialect.delimiter
        except Exception:
            delimiter = ","
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    entries = []
    for row in reader:
        normalized = _normalize_memory_entry(row)
        if normalized:
            entries.append(normalized)
    return entries


def _load_memory_entries_from_excel(file_path: str):
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    entries = []
    for row_values in rows[1:]:
        row = {headers[index]: row_values[index] for index in range(min(len(headers), len(row_values)))}
        normalized = _normalize_memory_entry(row)
        if normalized:
            entries.append(normalized)
    return entries


def _load_memory_file_entries(file_path: str, file_type: str):
    ext = file_type.lower()
    if ext == "json":
        return _load_memory_entries_from_json(file_path)
    if ext in ["csv", "tsv"]:
        return _load_memory_entries_from_delimited_file(file_path, "\t" if ext == "tsv" else None)
    if ext in ["xlsx", "xlsm", "xltx", "xltm"]:
        return _load_memory_entries_from_excel(file_path)

    try:
        raw_text = parse_file(file_path)
    except Exception:
        raw_text = ""
    return _parse_memory_text_entries(raw_text or "")


def _get_cached_memory_file_entries(memory_file: KnowledgeFile):
    cache_key = (
        memory_file.id,
        memory_file.updated_at.isoformat() if memory_file.updated_at else "",
        memory_file.file_path,
    )
    with _memory_file_cache_lock:
        if cache_key in _memory_file_cache:
            return _memory_file_cache[cache_key]

    entries = _load_memory_file_entries(memory_file.file_path, memory_file.file_type or "")
    with _memory_file_cache_lock:
        _memory_file_cache[cache_key] = entries
    return entries


def _invalidate_memory_file_cache(memory_file_id: int):
    with _memory_file_cache_lock:
        stale_keys = [key for key in _memory_file_cache if key[0] == memory_file_id]
        for key in stale_keys:
            _memory_file_cache.pop(key, None)


def _append_memory_entry_to_excel(memory_file: KnowledgeFile, source_text: str, translated_text: str,
                                  source_lang: str, target_lang: str):
    from openpyxl import load_workbook

    workbook = load_workbook(memory_file.file_path)
    worksheet = workbook.active

    headers = []
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        headers = [worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)]
        headers = [str(header).strip() if header is not None else "" for header in headers]

    if not any(headers):
        headers = ["source_text", "translated_text"]
        for index, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=index).value = header

    source_keywords = ["source", "source_text", "sourcetext", "src", "原文", "源文", "text"]
    translated_keywords = ["target", "translated_text", "translatedtext", "translation", "译文", "targettext", "result"]
    normalized_headers = {
        _normalize_header_key(header): index
        for index, header in enumerate(headers, start=1)
        if _normalize_header_key(header)
    }

    source_column = 1
    translated_column = 2

    next_row = worksheet.max_row + 1 if worksheet.max_row else 2
    worksheet.cell(row=next_row, column=source_column).value = source_text
    worksheet.cell(row=next_row, column=translated_column).value = translated_text

    workbook.save(memory_file.file_path)
    workbook.close()


def _search_memory_file(db: Session, memory_file_id: int, source_text: str, source_lang: str, target_lang: str,
                        threshold: float = 0.8, preserve_sentence_unmatched: bool = True):
    if not memory_file_id:
        return None

    memory_file = db.query(KnowledgeFile).filter(KnowledgeFile.id == memory_file_id).first()
    if not memory_file or not os.path.exists(memory_file.file_path):
        return None

    entries = _get_cached_memory_file_entries(memory_file)
    candidates = []
    for entry in entries:
        entry_source_lang = (entry.get("source_lang") or source_lang).strip()
        entry_target_lang = (entry.get("target_lang") or target_lang).strip()
        if entry_source_lang != source_lang or entry_target_lang != target_lang:
            continue
        candidates.append(entry)

    return _match_memory_candidates(source_text, candidates, threshold=threshold, preserve_sentence_unmatched=preserve_sentence_unmatched)


def _collect_memory_candidates(db: Session, source_lang: str, target_lang: str,
                               bank: str = None, memory_file_id: int = None):
    candidates = []
    seen = set()

    if memory_file_id:
        memory_file = db.query(KnowledgeFile).filter(KnowledgeFile.id == memory_file_id).first()
        if memory_file and os.path.exists(memory_file.file_path):
            for entry in _get_cached_memory_file_entries(memory_file):
                entry_source_lang = (entry.get("source_lang") or source_lang).strip()
                entry_target_lang = (entry.get("target_lang") or target_lang).strip()
                if entry_source_lang != source_lang or entry_target_lang != target_lang:
                    continue
                source_text_value = (entry.get("source_text") or "").strip()
                translated_text_value = (entry.get("translated_text") or "").strip()
                if not source_text_value or not translated_text_value:
                    continue
                dedupe_key = (source_text_value, translated_text_value)
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                candidates.append({
                    "source_text": source_text_value,
                    "translated_text": translated_text_value,
                })

    query = db.query(MemoryBank).filter(
        MemoryBank.source_lang == source_lang,
        MemoryBank.target_lang == target_lang,
    )
    if bank:
        query = query.filter(MemoryBank.tags == bank)

    for entry in query.all():
        source_text_value = (entry.source_text or "").strip()
        translated_text_value = (entry.translated_text or "").strip()
        if not source_text_value or not translated_text_value:
            continue
        dedupe_key = (source_text_value, translated_text_value)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        candidates.append({
            "source_text": source_text_value,
            "translated_text": translated_text_value,
        })

    return candidates


def _find_memory_glossary(source_text: str, candidates, max_entries: int = 20):
    glossary = []
    normalized_source = (source_text or "").strip()
    if not normalized_source:
        return glossary

    compact_source = _normalize_memory_lookup_key(source_text)
    for entry in sorted(candidates, key=lambda item: len(item["source_text"]), reverse=True):
        source_term = entry["source_text"]
        translated_term = entry["translated_text"]
        normalized_term = _normalize_match_text(source_term)
        compact_term = _normalize_memory_lookup_key(source_term)
        if normalized_term == normalized_source or (compact_term and compact_term == compact_source):
            return [{"source_text": source_term, "translated_text": translated_term, "full_match": True}]
        if len(source_term) >= 2 and (source_term in source_text or normalized_term in normalized_source or (compact_term and compact_term in compact_source)):
            glossary.append({"source_text": source_term, "translated_text": translated_term, "full_match": False})
        if len(glossary) >= max_entries:
            break
    return glossary


def _apply_memory_glossary(source_text: str, glossary):
    translated = source_text
    replaced = False
    for entry in sorted(glossary, key=lambda item: len(item["source_text"]), reverse=True):
        source_term = entry["source_text"]
        translated_term = entry["translated_text"]
        if source_term and source_term in translated:
            translated = translated.replace(source_term, translated_term)
            replaced = True
    return translated, replaced


def _do_translate(text: str, engine: str, model: str, source_lang: str, target_lang: str, db: Session) -> str:
    result = None
    if engine in ["memory", "hybrid"]:
        r, hit = translate_with_memory(
            text,
            source_lang,
            target_lang,
            db,
            bank=_get_memory_bank(),
            memory_file_id=_get_memory_file_id(),
            allow_partial=engine == "memory",
        )
        if hit:
            result = r
            _record_translation_usage("memory", text)
    if engine in ["ai", "hybrid"] and not result:
        result = translate_with_ai(text, model, source_lang, target_lang)
        if result == text:
            _record_passthrough_usage(text)
        elif result:
            _record_translation_usage("ai", text)
    if not result:
        if engine == "memory":
            _record_passthrough_usage(text)
            return text
        raise HTTPException(status_code=500, detail="翻译失败")
    return result


def _translate_dita_xml(fpath: str, engine: str, model: str, source_lang: str, target_lang: str, db: Session) -> tuple:
    """Returns (translated_content, all_original_parts, all_translated_parts)"""
    all_original_parts = []
    all_translated_parts = []

    with open(fpath, "rb") as f:
        raw_bytes = f.read()
    raw_str = raw_bytes.decode("utf-8")

    ns_decls = re.findall(r'xmlns:(\w+)="([^"]*)"', raw_str)
    for prefix, uri in ns_decls:
        if prefix:
            ET_LXML.register_namespace(prefix, uri)

    parser = ET_LXML.XMLParser(remove_blank_text=False, strip_cdata=False)
    tree = ET_LXML.parse(fpath, parser)
    root = tree.getroot()

    def _translate_element(elem):
        if elem.text and elem.text.strip():
            orig = elem.text
            stripped = orig.strip()
            translated = _do_translate(stripped, engine, model, source_lang, target_lang, db)
            elem.text = orig.replace(stripped, translated, 1)
            all_original_parts.append(stripped)
            all_translated_parts.append(translated)
        for child in elem:
            _translate_element(child)
        if elem.tail and elem.tail.strip():
            orig = elem.tail
            stripped = orig.strip()
            translated = _do_translate(stripped, engine, model, source_lang, target_lang, db)
            elem.tail = orig.replace(stripped, translated, 1)
            all_original_parts.append(stripped)
            all_translated_parts.append(translated)

    _translate_element(root)

    doctype = tree.docinfo.doctype
    result_bytes = ET_LXML.tostring(
        tree,
        xml_declaration=True,
        encoding="UTF-8",
        doctype=doctype,
        pretty_print=False,
    )
    return result_bytes.decode("utf-8"), all_original_parts, all_translated_parts


def _translate_xlf_xml(fpath: str, engine: str, model: str, source_lang: str, target_lang: str, db: Session) -> tuple:
    all_original_parts = []
    all_translated_parts = []

    tree = ET.parse(fpath)
    root_xliff = tree.getroot()
    xliff_ns = "urn:oasis:names:tc:xliff:document:1.2"
    ET.register_namespace("", xliff_ns)

    for trans_unit in root_xliff.iter(f"{{{xliff_ns}}}trans-unit"):
        source_el = trans_unit.find(f"{{{xliff_ns}}}source")
        if source_el is None:
            continue

        target_parts = []
        text_segments = []

        if source_el.text and source_el.text.strip():
            seg = source_el.text.strip()
            text_segments.append(seg)
            target_parts.append(_do_translate(seg, engine, model, source_lang, target_lang, db))

        for child in list(source_el):
            child_tail = child.tail
            child.tail = None
            tag_str = ET.tostring(child, encoding="unicode").strip()
            child.tail = child_tail
            target_parts.append(tag_str)
            if child_tail and child_tail.strip():
                seg = child_tail.strip()
                text_segments.append(seg)
                target_parts.append(_do_translate(seg, engine, model, source_lang, target_lang, db))

        if not target_parts:
            continue

        target_content = "".join(target_parts)
        target_el = trans_unit.find(f"{{{xliff_ns}}}target")
        target_str = f"<target>{target_content}</target>"
        try:
            target_fragment = ET.fromstring(target_str)
        except Exception:
            continue
        if target_el is not None:
            trans_unit.remove(target_el)
        trans_unit.append(target_fragment)
        target_fragment.set("state", "translated")

        all_original_parts.append(" ".join(text_segments))
        all_translated_parts.append(target_content)

    return ET.tostring(root_xliff, encoding="unicode"), all_original_parts, all_translated_parts


def _translate_idml(fpath: str, engine: str, model: str, source_lang: str, target_lang: str, db: Session) -> tuple:
    """Translate IDML (InDesign Markup Language) preserving formatting."""
    IDML_NS = "http://ns.adobe.com/AdobeInDesign/idml/1.0/packaging"
    ET_LXML.register_namespace("idPkg", IDML_NS)

    in_buf = io.BytesIO()
    with open(fpath, "rb") as f:
        in_buf.write(f.read())
    in_buf.seek(0)

    all_original = []
    all_translated = []
    texts_to_translate = []
    content_targets = []
    modified_entries = {}

    def _is_meaningful_text(text: str) -> bool:
        stripped = text.strip()
        if not stripped:
            return False
        cleaned = re.sub(r'[\d\s.,;:!?\-+*/=()\[\]{}<>@#$%^&|\\/\'"]+', '', stripped)
        if len(cleaned) < 2:
            return False
        return True

    with zipfile.ZipFile(in_buf, 'r') as z_in:
        for zinfo in z_in.infolist():
            name = zinfo.filename
            if not name.endswith(".xml"):
                continue
            raw = z_in.read(zinfo)
            try:
                root = ET_LXML.fromstring(raw)
            except Exception:
                continue

            try:
                content_elems = root.xpath("//*[local-name()='Content']")
            except Exception:
                continue

            has_text = False
            for elem in content_elems:
                if isinstance(elem, ET_LXML._Element) and elem.text and _is_meaningful_text(elem.text):
                    texts_to_translate.append(elem.text.strip())
                    content_targets.append(elem)
                    has_text = True
            if has_text:
                modified_entries[name] = root

    if not texts_to_translate:
        return in_buf.read(), [], []

    translated_parts = _translate_text_items(texts_to_translate, engine, model, source_lang, target_lang, db)
    while len(translated_parts) < len(texts_to_translate):
        translated_parts.append(texts_to_translate[len(translated_parts)])
    for text_idx, translated in enumerate(translated_parts):
        all_original.append(texts_to_translate[text_idx])
        all_translated.append(translated)
        content_targets[text_idx].text = translated

    out_buf = io.BytesIO()
    in_buf.seek(0)
    with zipfile.ZipFile(in_buf, 'r') as z_in:
        with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as z_out:
            for zinfo in z_in.infolist():
                raw = z_in.read(zinfo)
                name = zinfo.filename
                if name in modified_entries:
                    raw = ET_LXML.tostring(modified_entries[name], xml_declaration=True, encoding="UTF-8", pretty_print=False)
                output_info = _clone_zipinfo(zinfo)
                if name == "mimetype":
                    output_info.compress_type = zipfile.ZIP_STORED
                z_out.writestr(output_info, raw)
    out_buf.seek(0)

    return out_buf.read(), all_original, all_translated


def _translate_pptx_xml(fpath: str, engine: str, model: str, source_lang: str, target_lang: str, db: Session) -> tuple:
    """Returns (translated_bytes, all_original_parts, all_translated_parts)"""
    PPTX_NS = {
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
        "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    }
    DRAWING_NS = PPTX_NS["a"]
    CHART_NS = PPTX_NS["c"]

    for prefix, uri in PPTX_NS.items():
        ET_LXML.register_namespace(prefix, uri)

    in_buf = io.BytesIO()
    with open(fpath, "rb") as f:
        in_buf.write(f.read())
    in_buf.seek(0)

    pptx_original = []
    pptx_translated = []
    texts_to_translate = []
    xml_roots = {}
    text_targets = []

    def is_translatable_pptx_xml(name: str) -> bool:
        return (
            name.startswith("ppt/slides/slide")
            or name.startswith("ppt/notesSlides/notesSlide")
            or name.startswith("ppt/slideMasters/slideMaster")
            or name.startswith("ppt/slideLayouts/slideLayout")
            or name.startswith("ppt/notesMasters/notesMaster")
            or name.startswith("ppt/handoutMasters/handoutMaster")
            or name.startswith("ppt/diagrams/data")
            or name.startswith("ppt/diagrams/drawing")
            or name.startswith("ppt/charts/chart")
        )

    def collect_text_target(elem):
        if elem.text and elem.text.strip():
            stripped = elem.text.strip()
            if _text_matches_source_language(stripped, source_lang):
                texts_to_translate.append(stripped)
                text_targets.append((elem, stripped))

    def preserve_short_fragment(text: str) -> str | None:
        if (source_lang or "").lower() != "zh":
            return None
        compact = _normalize_memory_lookup_key(text)
        chinese_chars = re.findall(r"[\u4e00-\u9fff]", text or "")
        if len(chinese_chars) != 1 or compact != chinese_chars[0]:
            return None

        result, hit = translate_with_memory(
            text,
            source_lang,
            target_lang,
            db,
            bank=_get_memory_bank(),
            memory_file_id=_get_memory_file_id(),
            allow_partial=False,
        )
        if hit:
            _record_translation_usage("memory", text)
        return result if hit else text

    with zipfile.ZipFile(in_buf, 'r') as z_in:
        for zinfo in z_in.infolist():
            name = zinfo.filename
            if not name.endswith(".xml") or "/_rels/" in name:
                continue
            if not is_translatable_pptx_xml(name):
                continue

            raw = z_in.read(zinfo)
            root = ET_LXML.fromstring(raw if isinstance(raw, bytes) else raw.encode("utf-8"))
            xml_roots[name] = root
            for t_elem in root.iter(f"{{{DRAWING_NS}}}t"):
                collect_text_target(t_elem)
            if name.startswith("ppt/charts/chart"):
                for v_elem in root.iter(f"{{{CHART_NS}}}v"):
                    collect_text_target(v_elem)

    if texts_to_translate:
        translated_parts = [None] * len(texts_to_translate)
        pending_texts = []
        pending_indexes = []
        for index, text in enumerate(texts_to_translate):
            preserved = preserve_short_fragment(text)
            if preserved is None:
                pending_texts.append(text)
                pending_indexes.append(index)
            else:
                translated_parts[index] = preserved

        if pending_texts:
            pending_translations = _translate_text_items(pending_texts, engine, model, source_lang, target_lang, db)
            for pending_index, translated in zip(pending_indexes, pending_translations):
                translated_parts[pending_index] = translated

        for i, (t_elem, original_text) in enumerate(text_targets):
            if i < len(translated_parts):
                translated = translated_parts[i] or original_text
                pptx_original.append(texts_to_translate[i])
                pptx_translated.append(translated)
                t_elem.text = (t_elem.text or original_text).replace(original_text, translated, 1)

    out_buf = io.BytesIO()
    in_buf.seek(0)
    with zipfile.ZipFile(in_buf, 'r') as z_in:
        with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as z_out:
            for zinfo in z_in.infolist():
                raw = z_in.read(zinfo)
                name = zinfo.filename
                if name in xml_roots:
                    raw = ET_LXML.tostring(xml_roots[name], xml_declaration=True, encoding="UTF-8", pretty_print=False)
                z_out.writestr(zinfo, raw)
    out_buf.seek(0)

    return out_buf.read(), pptx_original, pptx_translated


def _apply_translated_text_to_docx_paragraph(para, translated: str):
    runs = para.runs
    if not runs:
        return
    if len(runs) == 1:
        runs[0].text = translated
        return

    total_orig = sum(len(r.text or "") for r in runs)
    if total_orig == 0:
        runs[0].text = translated
        for r in runs[1:]:
            r.text = ""
        return

    pos = 0
    for r in runs:
        orig_len = len(r.text or "")
        chunk_len = max(1, int(len(translated) * orig_len / total_orig))
        chunk = translated[pos:pos + chunk_len]
        r.text = chunk
        pos += chunk_len
    if pos < len(translated):
        runs[-1].text += translated[pos:]


def _collect_docx_container_paragraphs(container, paragraphs_to_translate):
    for para in getattr(container, "paragraphs", []):
        full_text = para.text
        if full_text.strip():
            paragraphs_to_translate.append(para)

    for table in getattr(container, "tables", []):
        for row in table.rows:
            for cell in row.cells:
                _collect_docx_container_paragraphs(cell, paragraphs_to_translate)


def _translate_docx(fpath: str, engine: str, model: str, source_lang: str, target_lang: str, db: Session) -> tuple:
    """Translate DOCX preserving formatting (paragraphs, runs, tables, headers, footers)."""
    from docx import Document

    doc = Document(fpath)
    all_original = []
    all_translated = []

    paragraphs_to_translate = []
    seen_story_parts = set()

    def add_container(container):
        part_name = str(getattr(getattr(container, "part", None), "partname", ""))
        if part_name and part_name in seen_story_parts:
            return
        if part_name:
            seen_story_parts.add(part_name)
        _collect_docx_container_paragraphs(container, paragraphs_to_translate)

    add_container(doc)
    for section in doc.sections:
        for container in [
            section.header,
            section.footer,
            section.first_page_header,
            section.first_page_footer,
            section.even_page_header,
            section.even_page_footer,
        ]:
            add_container(container)

    texts = [p.text for p in paragraphs_to_translate]
    if texts:
        translated_texts = _translate_text_items(texts, engine, model, source_lang, target_lang, db)
        for i, para in enumerate(paragraphs_to_translate):
            if i < len(translated_texts):
                translated = translated_texts[i]
                all_original.append(texts[i])
                all_translated.append(translated)
                _apply_translated_text_to_docx_paragraph(para, translated)

    out_buf = io.BytesIO()
    doc.save(out_buf)
    out_buf.seek(0)
    return out_buf.read(), all_original, all_translated


def _translate_xlsx(fpath: str, engine: str, model: str, source_lang: str, target_lang: str, db: Session) -> tuple:
    """Translate XLSX preserving all formatting, formulas, and structure."""
    from openpyxl import load_workbook

    wb = load_workbook(fpath)
    all_original = []
    all_translated = []
    text_indexes = {}
    unique_texts = []
    cell_entries = []
    used_sheet_titles = set()

    for ws in wb.worksheets:
        translated_title = ws.title
        if not _filename_looks_non_translatable(ws.title, source_lang):
            translated_title = _do_translate(ws.title.strip(), engine, model, source_lang, target_lang, db)
        translated_title = _sanitize_excel_sheet_title(translated_title)
        translated_title = _dedupe_excel_sheet_title(translated_title, used_sheet_titles)
        ws.title = translated_title
        used_sheet_titles.add(translated_title)

        for header_footer_name in ["oddHeader", "oddFooter", "evenHeader", "evenFooter", "firstHeader", "firstFooter"]:
            header_footer = getattr(ws, header_footer_name, None)
            if not header_footer:
                continue
            for part_name in ["left", "center", "right"]:
                part = getattr(header_footer, part_name, None)
                original_text = getattr(part, "text", None)
                if not original_text or not str(original_text).strip():
                    continue
                translated_text = _translate_excel_header_footer_text(
                    str(original_text),
                    engine,
                    model,
                    source_lang,
                    target_lang,
                    db,
                )
                if translated_text != original_text:
                    part.text = translated_text
                all_original.append(str(original_text))
                all_translated.append(str(part.text or translated_text))

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    text = cell.value.strip()
                    if _cell_looks_non_translatable(text, source_lang):
                        cell_entries.append((cell, text, None))
                        continue
                    if text not in text_indexes:
                        text_indexes[text] = len(unique_texts)
                        unique_texts.append(text)
                    cell_entries.append((cell, text, text_indexes[text]))

    translated_texts = []
    if unique_texts:
        translated_texts = _translate_text_items(unique_texts, engine, model, source_lang, target_lang, db)

    for cell, original_text, translated_index in cell_entries:
        translated_text = original_text if translated_index is None else translated_texts[translated_index]
        if translated_index is None:
            _record_passthrough_usage(original_text)
        all_original.append(original_text)
        all_translated.append(translated_text)
        cell.value = translated_text

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.read(), all_original, all_translated


def _translate_pdf(fpath: str, engine: str, model: str, source_lang: str, target_lang: str, db: Session) -> tuple:
    """Translate PDF and rebuild an approximate layout-preserved PDF."""
    import fitz

    all_original = []
    all_translated = []
    doc = fitz.open(fpath)
    block_separator = "\n---PDFBLOCK---\n"

    try:
        for page in doc:
            blocks = []
            for block in page.get_text("blocks", sort=True):
                if len(block) < 5:
                    continue
                x0, y0, x1, y1, text = block[:5]
                block_type = block[6] if len(block) > 6 else 0
                cleaned_text = (text or "").strip()
                if block_type != 0 or not cleaned_text:
                    continue
                rect = fitz.Rect(x0, y0, x1, y1)
                if rect.is_empty or rect.width < 4 or rect.height < 4:
                    continue
                blocks.append({
                    "rect": rect,
                    "text": cleaned_text,
                    "line_count": max(1, len([line for line in cleaned_text.splitlines() if line.strip()])),
                })

            if not blocks:
                all_original.append("")
                all_translated.append("")
                continue

            original_blocks = [item["text"] for item in blocks]
            translated_blocks = _translate_text_items(original_blocks, engine, model, source_lang, target_lang, db)

            page_original = []
            page_translated = []

            for block_info, translated_text in zip(blocks, translated_blocks):
                rect = block_info["rect"]
                final_text = (translated_text or "").strip()
                page_original.append(block_info["text"])
                page_translated.append(final_text)
                page.add_redact_annot(rect, fill=(1, 1, 1))

            page.apply_redactions()

            for block_info, translated_text in zip(blocks, translated_blocks):
                rect = block_info["rect"]
                final_text = (translated_text or "").strip()

                line_count = max(block_info["line_count"], len([line for line in final_text.splitlines() if line.strip()]))
                font_size = min(16, max(7, rect.height / max(line_count * 1.35, 1)))
                rc = -1
                while font_size >= 5:
                    rc = page.insert_textbox(
                        rect,
                        final_text,
                        fontname="china-s",
                        fontsize=font_size,
                        color=(0, 0, 0),
                        align=fitz.TEXT_ALIGN_LEFT,
                        lineheight=1.1,
                        overlay=True,
                    )
                    if rc >= 0:
                        break
                    font_size -= 0.5

                if rc < 0 and final_text:
                    clipped_text = final_text[: max(1, int(len(final_text) * 0.85))] + "..."
                    page.insert_textbox(
                        rect,
                        clipped_text,
                        fontname="china-s",
                        fontsize=5,
                        color=(0, 0, 0),
                        align=fitz.TEXT_ALIGN_LEFT,
                        lineheight=1.05,
                        overlay=True,
                    )

            all_original.append("\n".join(page_original))
            all_translated.append("\n".join(page_translated))

        out_buf = io.BytesIO()
        doc.save(out_buf, garbage=3, deflate=True)
        out_buf.seek(0)
        return out_buf.read(), all_original, all_translated
    finally:
        doc.close()


def _translate_markdown(fpath: str, engine: str, model: str, source_lang: str, target_lang: str, db: Session) -> tuple:
    """Translate Markdown preserving all syntax (headings, links, code blocks, tables, lists)."""
    import re as md_re

    with open(fpath, "r", encoding="utf-8") as f:
        content = f.read()

    code_blocks = []
    inline_codes = []
    links = []
    images = []

    def _preserve_code_blocks(text):
        def _repl(m):
            code_blocks.append(m.group(0))
            return f"%%CODEBLOCK{len(code_blocks) - 1}%%"
        return md_re.sub(r'```[\s\S]*?```', _repl, text)

    def _preserve_inline_code(text):
        def _repl(m):
            inline_codes.append(m.group(0))
            return f"%%INLINECODE{len(inline_codes) - 1}%%"
        return md_re.sub(r'`[^`]+`', _repl, text)

    def _preserve_links(text):
        def _repl(m):
            links.append(m.group(0))
            return f"%%LINK{len(links) - 1}%%"
        return md_re.sub(r'\[([^\]]*)\]\([^)]+\)', _repl, text)

    def _preserve_images(text):
        def _repl(m):
            images.append(m.group(0))
            return f"%%IMAGE{len(images) - 1}%%"
        return md_re.sub(r'!\[([^\]]*)\]\([^)]+\)', _repl, text)

    protected = content
    protected = _preserve_code_blocks(protected)
    protected = _preserve_inline_code(protected)
    protected = _preserve_images(protected)
    protected = _preserve_links(protected)

    translated = _do_translate(protected, engine, model, source_lang, target_lang, db)

    for i, cb in enumerate(code_blocks):
        translated = translated.replace(f"%%CODEBLOCK{i}%%", cb)
    for i, ic in enumerate(inline_codes):
        translated = translated.replace(f"%%INLINECODE{i}%%", ic)
    for i, img in enumerate(images):
        translated = translated.replace(f"%%IMAGE{i}%%", img)
    for i, link in enumerate(links):
        translated = translated.replace(f"%%LINK{i}%%", link)

    return translated.encode("utf-8"), [content], [translated]


def _looks_like_hallucination(result: str, original: str) -> bool:
    if not result or not original:
        return False
    hallucination_signals = [
        "翻译规则",
        "Translation Rules",
        "Translation begins",
        "你是一个专业",
        "请将以下",
        "原文：",
        "只输出翻译",
        "严格将原文",
        "Only output the translated",
        "must be strictly translated",
    ]
    result_lower = result.lower()
    for signal in hallucination_signals:
        if signal.lower() in result_lower:
            return True
    if len(result) > max(len(original) * 5, 500):
        return True
    return False


def translate_with_ai(content: str, model: str, source_lang: str, target_lang: str, glossary=None) -> str:
    model = _normalize_ai_model(model)
    lang_names = {
        "zh": "中文", "en": "英文", "ja": "日文", "ko": "韩文",
        "fr": "法文", "de": "德文", "es": "西班牙文", "ru": "俄文"
    }
    src_name = lang_names.get(source_lang, source_lang)
    tgt_name = lang_names.get(target_lang, target_lang)

    glossary_lines = ""
    if glossary:
        glossary_pairs = []
        for entry in glossary:
            source_text = (entry.get("source_text") or "").strip()
            translated_text = (entry.get("translated_text") or "").strip()
            if source_text and translated_text:
                glossary_pairs.append(f"- {source_text} => {translated_text}")
        if glossary_pairs:
            glossary_lines = "\n5. 必须优先采用以下术语映射，保持术语译法一致\n\n术语映射：\n" + "\n".join(glossary_pairs) + "\n"

    prompt = f"""你是一个专业的技术文档翻译引擎。请将以下{src_name}文本翻译为{tgt_name}。

翻译规则：
1. 必须严格将原文内容翻译为{tgt_name}，输出必须是纯{tgt_name}文本，不得包含{src_name}
2. 保持原文的段落结构、编号、列表格式完全不变
3. 使用专业准确的技术术语
4. 只输出翻译后的{tgt_name}结果，不要添加任何解释、注释或原始文本
{glossary_lines}

原文：
{content[:8000]}"""
    messages = [{"role": "user", "content": prompt}]

    if model == "kimi":
        result = ai_client.call_kimi(messages, max_tokens=4096)
        if result is None:
            result = ai_client.chat(messages, max_tokens=4096, fallback=True)
    elif model == "deepseek":
        result = ai_client.call_deepseek(messages, max_tokens=4096)
        if result is None:
            result = ai_client.chat(messages, max_tokens=4096, fallback=True)
    elif model == "arkclaw":
        result = ai_client.call_arkclaw(messages, max_tokens=4096)
        if result is None:
            result = ai_client.chat(messages, max_tokens=4096, fallback=True)
    else:
        result = ai_client.chat(messages, max_tokens=4096, fallback=True)

    if result is None:
        available = ", ".join(ai_client.available_providers()) or "none"
        provider_errors = " | ".join(ai_client.last_provider_errors())
        detail = (
            f"AI翻译引擎不可用。当前已配置提供商: {available}。"
            "请检查 `/api/translation/providers/status`，并确认 KIMI_API_KEY、DEEPSEEK_API_KEY 或 ARKCLAW_API_KEY 已注入当前服务进程。"
        )
        if provider_errors:
            detail += f" 最近调用错误: {provider_errors}"
        raise HTTPException(
            status_code=500,
            detail=detail
        )

    if _looks_like_hallucination(result, content):
        return content

    return result


def translate_with_memory(content: str, source_lang: str, target_lang: str,
                          db: Session, bank: str = None, memory_file_id: int = None,
                          allow_partial: bool = True) -> tuple:
    candidates = _collect_memory_candidates(db, source_lang, target_lang, bank=bank, memory_file_id=memory_file_id)

    file_result = _search_memory_file(
        db,
        memory_file_id,
        content,
        source_lang,
        target_lang,
        preserve_sentence_unmatched=allow_partial,
    )
    if file_result:
        return file_result, True
    memory_result = _match_memory_candidates(content, candidates, threshold=0.88, preserve_sentence_unmatched=allow_partial)
    if memory_result:
        return memory_result, True

    memory_result = _match_memory_candidates(content, candidates, threshold=0.8, preserve_sentence_unmatched=allow_partial)
    if memory_result:
        return memory_result, True

    if not allow_partial:
        return None, False

    segmented_result, matched_segments = _translate_by_memory_segments(content, candidates, threshold=0.8)
    if segmented_result and matched_segments > 0:
        return segmented_result, True

    clause_result, matched_clauses = _translate_by_memory_clauses(content, candidates, threshold=0.8)
    if clause_result and matched_clauses > 0:
        return clause_result, True

    glossary = _find_memory_glossary(content, candidates)
    if glossary:
        if glossary[0].get("full_match"):
            return glossary[0]["translated_text"], True
        replaced_text, replaced = _apply_memory_glossary(content, glossary)
        if replaced:
            return replaced_text, True
    return None, False


def _translate_filename(filename: str, source_lang: str, target_lang: str, model: str = None) -> str:
    model = _normalize_ai_model(model)
    if source_lang == target_lang:
        return filename
    if _filename_looks_non_translatable(filename, source_lang):
        return filename

    lang_names = {
        "zh": "Chinese", "en": "English", "ja": "Japanese", "ko": "Korean",
        "fr": "French", "de": "German", "es": "Spanish", "ru": "Russian"
    }
    src_name = lang_names.get(source_lang, source_lang)
    tgt_name = lang_names.get(target_lang, target_lang)

    prompt = f'Translate this filename from {src_name} to {tgt_name}. Output ONLY the translated filename, no other text: "{filename}"'
    messages = [{"role": "user", "content": prompt}]
    result = None
    if model == "kimi":
        result = ai_client.call_kimi(messages, max_tokens=500)
    elif model == "deepseek":
        result = ai_client.call_deepseek(messages, max_tokens=500)
    elif model == "arkclaw":
        result = ai_client.call_arkclaw(messages, max_tokens=500)

    if result is None:
        result = ai_client.chat(messages, max_tokens=500, fallback=True)

    if result:
        translated = result.strip().strip('"').strip("'").replace("/", "_").replace("\\", "_").replace(":", "_")
        if translated and translated != filename and len(translated) <= max(len(filename) * 2, 120) and "does not contain" not in translated.lower():
            return translated

    return filename


def _run_translate_thread(doc_id: int, file_path: str, ext: str, filename: str,
                          engine: str, model: str, source_lang: str, target_lang: str,
                          memory_bank: str = "", memory_file_id: int = None):
    """Background thread that performs the actual translation and updates the DB record."""
    db = SessionLocal()
    _set_memory_bank(memory_bank)
    _set_memory_file_id(memory_file_id)
    _reset_translation_usage_stats()
    try:
        with _translate_tasks_lock:
            _translate_tasks[doc_id] = {"status": "processing", "error": None}

        if not os.path.exists(TRANSLATION_OUTPUT_DIR):
            os.makedirs(TRANSLATION_OUTPUT_DIR, exist_ok=True)

        base_name = os.path.splitext(filename)[0]
        translated_filename = _translate_filename(base_name, source_lang, target_lang, model)
        output_filename = f"{translated_filename}{ext}"
        output_path = os.path.join(TRANSLATION_OUTPUT_DIR, output_filename)

        all_original_parts = []
        all_translated_parts = []

        if ext == ".zip":
            with tempfile.TemporaryDirectory() as tmpdir:
                translated_files = []
                with zipfile.ZipFile(file_path, 'r') as zf:
                    for zinfo in zf.infolist():
                        if zinfo.is_dir():
                            continue
                        zip_name = zinfo.filename
                        safe_fname = f"_tmp_{len(translated_files)}{os.path.splitext(zip_name)[1]}"
                        safe_fpath = os.path.join(tmpdir, safe_fname)
                        try:
                            with open(safe_fpath, "wb") as sf:
                                sf.write(zf.read(zinfo))
                        except Exception:
                            continue
                        _, name_ext = os.path.splitext(zip_name)
                        ext_lower = name_ext.lower()
                        if ext_lower == ".dita":
                            try:
                                translated_content, orig, trans = _translate_dita_xml(safe_fpath, engine, model, source_lang, target_lang, db)
                                all_original_parts.extend(orig)
                                all_translated_parts.extend(trans)
                            except Exception:
                                continue
                        elif ext_lower == ".xlf":
                            try:
                                translated_content, orig, trans = _translate_xlf_xml(safe_fpath, engine, model, source_lang, target_lang, db)
                                all_original_parts.extend(orig)
                                all_translated_parts.extend(trans)
                            except Exception:
                                continue
                        elif ext_lower == ".pptx":
                            try:
                                translated_content, orig, trans = _translate_pptx_xml(safe_fpath, engine, model, source_lang, target_lang, db)
                                all_original_parts.append("\n".join(orig))
                                all_translated_parts.append("\n".join(trans))
                            except Exception:
                                continue
                        elif ext_lower == ".idml":
                            try:
                                translated_content, orig, trans = _translate_idml(safe_fpath, engine, model, source_lang, target_lang, db)
                                all_original_parts.append("\n".join(orig))
                                all_translated_parts.append("\n".join(trans))
                            except Exception:
                                continue
                        elif ext_lower in (".docx", ".doc"):
                            try:
                                translated_content, orig, trans = _translate_docx(safe_fpath, engine, model, source_lang, target_lang, db)
                                all_original_parts.extend(orig)
                                all_translated_parts.extend(trans)
                            except Exception:
                                continue
                        elif ext_lower in (".xlsx", ".xls"):
                            try:
                                translated_content, orig, trans = _translate_xlsx(safe_fpath, engine, model, source_lang, target_lang, db)
                                all_original_parts.extend(orig)
                                all_translated_parts.extend(trans)
                            except Exception:
                                continue
                        elif ext_lower == ".md":
                            try:
                                translated_content, orig, trans = _translate_markdown(safe_fpath, engine, model, source_lang, target_lang, db)
                                all_original_parts.extend(orig)
                                all_translated_parts.extend(trans)
                            except Exception:
                                continue
                        elif ext_lower == ".pdf":
                            try:
                                translated_content, orig, trans = _translate_pdf(safe_fpath, engine, model, source_lang, target_lang, db)
                                all_original_parts.extend(orig)
                                all_translated_parts.extend(trans)
                            except Exception:
                                continue
                        else:
                            try:
                                file_content = parse_file(safe_fpath)
                                if not file_content.strip():
                                    continue
                            except Exception:
                                continue
                            translated_content = _do_translate(file_content, engine, model, source_lang, target_lang, db)
                            all_original_parts.append(file_content)
                            all_translated_parts.append(translated_content)
                        dir_part, name_part = os.path.split(zip_name)
                        name_base = os.path.splitext(name_part)[0]
                        new_name = _translate_filename(name_base, source_lang, target_lang, model)
                        new_rel = os.path.join(dir_part, f"{new_name}{name_ext}") if dir_part else f"{new_name}{name_ext}"
                        translated_files.append((new_rel, translated_content))
                        if os.path.exists(safe_fpath):
                            os.remove(safe_fpath)
                if not translated_files:
                    raise Exception("ZIP包中未找到可翻译的文件")
                with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as out_zf:
                    for rel_path, content in translated_files:
                        if isinstance(content, bytes):
                            out_zf.writestr(rel_path, content)
                        else:
                            out_zf.writestr(rel_path, content.encode("utf-8"))
        elif ext == ".dita":
            translated_xml, orig, trans = _translate_dita_xml(file_path, engine, model, source_lang, target_lang, db)
            all_original_parts.extend(orig)
            all_translated_parts.extend(trans)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(translated_xml)
        elif ext == ".xlf":
            translated_xml, orig, trans = _translate_xlf_xml(file_path, engine, model, source_lang, target_lang, db)
            all_original_parts.extend(orig)
            all_translated_parts.extend(trans)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(translated_xml)
        elif ext == ".pptx":
            translated_bytes, orig, trans = _translate_pptx_xml(file_path, engine, model, source_lang, target_lang, db)
            all_original_parts.append("\n".join(orig))
            all_translated_parts.append("\n".join(trans))
            with open(output_path, "wb") as f:
                f.write(translated_bytes)
        elif ext == ".idml":
            translated_bytes, orig, trans = _translate_idml(file_path, engine, model, source_lang, target_lang, db)
            all_original_parts.append("\n".join(orig))
            all_translated_parts.append("\n".join(trans))
            with open(output_path, "wb") as f:
                f.write(translated_bytes)
        elif ext in (".docx", ".doc"):
            translated_bytes, orig, trans = _translate_docx(file_path, engine, model, source_lang, target_lang, db)
            all_original_parts.extend(orig)
            all_translated_parts.extend(trans)
            with open(output_path, "wb") as f:
                f.write(translated_bytes)
        elif ext in (".xlsx", ".xls"):
            translated_bytes, orig, trans = _translate_xlsx(file_path, engine, model, source_lang, target_lang, db)
            all_original_parts.extend(orig)
            all_translated_parts.extend(trans)
            with open(output_path, "wb") as f:
                f.write(translated_bytes)
        elif ext == ".md":
            translated_bytes, orig, trans = _translate_markdown(file_path, engine, model, source_lang, target_lang, db)
            all_original_parts.extend(orig)
            all_translated_parts.extend(trans)
            with open(output_path, "wb") as f:
                f.write(translated_bytes)
        elif ext == ".pdf":
            translated_bytes, orig, trans = _translate_pdf(file_path, engine, model, source_lang, target_lang, db)
            all_original_parts.extend(orig)
            all_translated_parts.extend(trans)
            with open(output_path, "wb") as f:
                f.write(translated_bytes)
        elif ext == ".txt":
            content = _read_text_file_with_fallback(file_path)
            if not content.strip():
                raise Exception("无法从文件中提取文本内容")
            translated = _do_translate(content, engine, model, source_lang, target_lang, db)
            all_original_parts.append(content)
            all_translated_parts.append(translated)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(translated)
        else:
            content = parse_file(file_path)
            if not content.strip():
                raise Exception("无法从文件中提取文本内容")
            translated = _do_translate(content, engine, model, source_lang, target_lang, db)
            all_original_parts.append(content)
            all_translated_parts.append(translated)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(translated)

        original_content = "\n\n".join(all_original_parts)
        translated_content = "\n\n".join(all_translated_parts)

        doc = db.query(TranslationDoc).filter(TranslationDoc.id == doc_id).first()
        if doc:
            usage_stats = _snapshot_translation_usage_stats()
            doc.translated_filename = output_filename
            doc.original_content = original_content[:8000]
            doc.translated_content = translated_content[:8000]
            doc.original_preview = original_content[:500] + "..." if len(original_content) > 500 else original_content
            doc.translated_preview = translated_content[:500] + "..." if len(translated_content) > 500 else translated_content
            total_chars = len(original_content)
            total_words = _count_text_units(original_content)
            doc.source_char_count = total_chars
            doc.ai_char_count = min(total_chars, usage_stats["ai_char_count"])
            doc.memory_char_count = min(total_chars, usage_stats["memory_char_count"])
            doc.source_word_count = total_words
            doc.ai_word_count = min(total_words, usage_stats["ai_word_count"])
            doc.memory_word_count = min(total_words, usage_stats["memory_word_count"])
            db.commit()

        with _translate_tasks_lock:
            _translate_tasks[doc_id] = {"status": "completed", "error": None, "translated_filename": output_filename}

    except Exception as e:
        error_msg = str(e)
        doc = db.query(TranslationDoc).filter(TranslationDoc.id == doc_id).first()
        if doc:
            doc.translated_filename = f"ERROR:{error_msg[:200]}"
            db.commit()
        with _translate_tasks_lock:
            _translate_tasks[doc_id] = {"status": "error", "error": error_msg}
    finally:
        db.close()


@router.post("/translate", response_model=TranslationResponse)
async def translate_text(req: TranslationRequest, db: Session = Depends(get_db)):
    req.model = _normalize_ai_model(req.model)
    translated = None
    from_memory = False
    from_ai = False
    engine = req.engine
    _set_memory_bank(req.memory_bank)
    _set_memory_file_id(req.memory_file_id)
    source_char_count = len(req.content)
    source_word_count = _count_text_units(req.content)

    if engine in ["memory", "hybrid"]:
        result, hit = translate_with_memory(
            req.content,
            req.source_lang,
            req.target_lang,
            db,
            bank=req.memory_bank,
            memory_file_id=req.memory_file_id,
            allow_partial=req.engine == "memory",
        )
        if hit:
            translated = result
            from_memory = True
            if engine == "memory":
                doc_record = TranslationDoc(
                    filename=f"text_{datetime.now().strftime('%Y%m%d%H%M%S')}",
                    file_type="text",
                    source_lang=req.source_lang,
                    target_lang=req.target_lang,
                    engine=engine,
                    model=req.model,
                    original_content=req.content[:8000],
                    translated_content=translated[:8000],
                    source_char_count=source_char_count,
                    ai_char_count=0,
                    memory_char_count=source_char_count,
                    source_word_count=source_word_count,
                    ai_word_count=0,
                    memory_word_count=source_word_count,
                )
                db.add(doc_record)
                db.commit()
                return TranslationResponse(
                    original=req.content,
                    translated=translated,
                    engine_used="memory",
                    from_memory=True,
                    from_ai=False
                )

    if engine in ["ai", "hybrid"] and not translated:
        translated = translate_with_ai(req.content, req.model, req.source_lang, req.target_lang)
        from_ai = True

        if engine == "hybrid":
            create_memory_entry(
                db=db,
                source_text=req.content,
                translated_text=translated,
                source_lang=req.source_lang,
                target_lang=req.target_lang,
                tags=req.memory_bank or ""
            )

    if not translated:
        raise HTTPException(status_code=500, detail="翻译失败，AI和记忆库引擎均未返回结果")

    ai_chars = source_char_count if from_ai else 0
    memory_chars = source_char_count if from_memory else 0
    ai_words = source_word_count if from_ai else 0
    memory_words = source_word_count if from_memory else 0
    doc_record = TranslationDoc(
        filename=f"text_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        file_type="text",
        source_lang=req.source_lang,
        target_lang=req.target_lang,
        engine=engine,
        model=req.model,
        original_content=req.content[:8000],
        translated_content=translated[:8000],
        source_char_count=source_char_count,
        ai_char_count=ai_chars,
        memory_char_count=memory_chars,
        source_word_count=source_word_count,
        ai_word_count=ai_words,
        memory_word_count=memory_words,
    )
    db.add(doc_record)
    db.commit()

    return TranslationResponse(
        original=req.content,
        translated=translated,
        engine_used=engine,
        from_memory=from_memory,
        from_ai=from_ai
    )


@router.get("/stats")
async def get_translation_stats(batch_id: str = Query(None), db: Session = Depends(get_db)):
    all_docs = db.query(TranslationDoc).all()
    docs_updated = False
    for doc in all_docs:
        word_counts = _normalize_doc_word_counts(doc)
        if word_counts["dirty"]:
            doc.source_word_count = word_counts["source_word_count"]
            doc.ai_word_count = word_counts["ai_word_count"]
            doc.memory_word_count = word_counts["memory_word_count"]
            docs_updated = True
    if docs_updated:
        db.commit()

    text_docs = [d for d in all_docs if d.file_type == "text"]
    file_docs = [d for d in all_docs if d.file_type != "text"]

    text_word_count = sum(_normalize_doc_word_counts(doc)["source_word_count"] for doc in text_docs)
    all_usage = [_normalize_doc_word_counts(doc) for doc in all_docs]
    ai_word_count = sum(item["ai_word_count"] for item in all_usage)
    memory_word_count = sum(item["memory_word_count"] for item in all_usage)

    latest_batch_id = (batch_id or "").strip() or None
    if not latest_batch_id:
        latest_batch_doc = (
            db.query(TranslationDoc)
            .filter(TranslationDoc.file_type != "text", TranslationDoc.batch_id != "")
            .order_by(TranslationDoc.created_at.desc())
            .first()
        )
        latest_batch_id = latest_batch_doc.batch_id if latest_batch_doc else None

    current_upload = {
        "batch_id": latest_batch_id,
        "doc_count": 0,
        "doc_word_count": 0,
        "ai_word_count": 0,
        "memory_word_count": 0,
    }
    if latest_batch_id:
        batch_docs = (
            db.query(TranslationDoc)
            .filter(TranslationDoc.file_type != "text", TranslationDoc.batch_id == latest_batch_id)
            .all()
        )
        current_upload = {"batch_id": latest_batch_id, **_summarize_docs(batch_docs)}

    overall_docs = _summarize_docs(file_docs)

    return {
        "text_word_count": text_word_count,
        "doc_count": overall_docs["doc_count"],
        "doc_word_count": overall_docs["doc_word_count"],
        "ai_word_count": ai_word_count,
        "memory_word_count": memory_word_count,
        "current_upload": current_upload,
    }


@router.get("/providers/status")
async def get_translation_provider_status():
    return ai_client.provider_status()


@router.post("/translate/file")
async def translate_file(
    file: UploadFile = File(...),
    engine: str = Form("hybrid"),
    model: str = Form("kimi"),
    source_lang: str = Form("zh"),
    target_lang: str = Form("en"),
    memory_bank: str = Form(""),
    memory_file_id: int = Form(None),
    batch_id: str = Form(""),
    db: Session = Depends(get_db)
):
    model = _normalize_ai_model(model)
    if not os.path.exists(UPLOAD_DIR):
        os.makedirs(UPLOAD_DIR, exist_ok=True)

    filename = file.filename
    file_path = os.path.join(UPLOAD_DIR, filename)

    try:
        with open(file_path, "wb") as buffer:
            while chunk := await file.read(1024 * 1024):
                buffer.write(chunk)
    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=f"文件保存失败: {str(e)}")

    ext = os.path.splitext(filename)[1].lower()
    if ext in UNSUPPORTED_TRANSLATION_EXTENSIONS:
        raise HTTPException(status_code=400, detail="当前文件格式暂不支持文档翻译，请改用 Word、Excel、PPT、PDF、Markdown、TXT 或 XLF 文件")

    doc = TranslationDoc(
        filename=filename,
        file_type=ext.replace(".", ""),
        source_lang=source_lang,
        target_lang=target_lang,
        engine=engine,
        model=model if engine != "memory" else "memory",
        original_content="",
        translated_content="",
        translated_filename="",
        original_preview="",
        translated_preview="",
        batch_id=(batch_id or "").strip(),
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    thread = threading.Thread(
        target=_run_translate_thread,
        args=(doc.id, file_path, ext, filename, engine, model, source_lang, target_lang, memory_bank, memory_file_id),
        daemon=True
    )
    thread.start()

    return {
        "message": "翻译任务已提交",
        "doc_id": doc.id,
        "original_filename": filename,
        "download_url": f"/api/translation/download/{doc.id}"
    }


@router.get("/translate/file/{doc_id}/status")
async def get_translate_file_status(doc_id: int, db: Session = Depends(get_db)):
    doc = db.query(TranslationDoc).filter(TranslationDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    orig_filename = doc.filename or ""

    with _translate_tasks_lock:
        task = _translate_tasks.get(doc_id)

    if task:
        return {
            "doc_id": doc_id,
            "status": task["status"],
            "error": task.get("error"),
            "translated_filename": task.get("translated_filename", ""),
            "original_filename": orig_filename,
            "download_url": f"/api/translation/download/{doc_id}" if task["status"] == "completed" else None,
        }

    fn = doc.translated_filename or ""
    if fn == "":
        if doc.created_at and datetime.utcnow() - doc.created_at > timedelta(minutes=10):
            error_msg = "翻译任务已中断，请重新提交文件翻译"
            doc.translated_filename = f"ERROR:{error_msg}"
            db.commit()
            return {"doc_id": doc_id, "status": "error", "error": error_msg, "original_filename": orig_filename}
        return {"doc_id": doc_id, "status": "processing", "error": None, "original_filename": orig_filename}
    if fn.startswith("ERROR:"):
        return {"doc_id": doc_id, "status": "error", "error": fn[6:], "original_filename": orig_filename}

    return {
        "doc_id": doc_id,
        "status": "completed",
        "translated_filename": fn,
        "original_filename": orig_filename,
        "download_url": f"/api/translation/download/{doc_id}",
        "error": None
    }


@router.get("/download/{doc_id}")
async def download_translated_file(doc_id: int, db: Session = Depends(get_db)):
    doc = db.query(TranslationDoc).filter(TranslationDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Translation doc not found")

    output_filename = doc.translated_filename
    if not output_filename or output_filename.startswith("ERROR:"):
        raise HTTPException(status_code=404, detail="翻译尚未完成或失败")

    output_path = os.path.join(TRANSLATION_OUTPUT_DIR, output_filename)

    if os.path.exists(output_path):
        media_type = "application/octet-stream"
        if output_filename.endswith(".zip"):
            media_type = "application/zip"
        elif output_filename.endswith(".xlf"):
            media_type = "application/xliff+xml"
        elif output_filename.endswith(".idml"):
            media_type = "application/vnd.adobe.indesign-idml-package"
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type=media_type
        )

    raise HTTPException(status_code=404, detail="文件不存在")


@router.get("/reviewed-docs")
async def get_reviewed_documents(db: Session = Depends(get_db)):
    reviews = get_reviews(db)
    reviewed_doc_ids = set()
    for review in reviews:
        if review.status == "completed":
            reviewed_doc_ids.add(review.document_id)

    documents = get_documents(db, limit=500)
    result = []
    for doc in documents:
        if doc.id in reviewed_doc_ids:
            result.append({
                "id": doc.id,
                "filename": doc.filename,
                "file_type": doc.file_type,
                "preview": doc.preview,
                "created_at": doc.created_at.isoformat() if doc.created_at else None
            })

    return result


@router.get("/memory/banks")
async def get_banks(db: Session = Depends(get_db)):
    """Return list of distinct memory bank names (tags) available for translation."""
    banks = get_memory_banks(db)
    return {"banks": banks}


@router.get("/memory", response_model=list[MemoryEntryOut])
async def list_memory_entries(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, le=500),
    keyword: str = Query(None),
    db: Session = Depends(get_db)
):
    return get_memory_entries(db, skip=skip, limit=limit, keyword=keyword)


@router.post("/memory", response_model=MemoryEntryOut)
async def add_memory_entry(entry: MemoryEntry, db: Session = Depends(get_db)):
    return create_memory_entry(
        db=db,
        source_text=entry.source_text,
        translated_text=entry.translated_text,
        source_lang=entry.source_lang,
        target_lang=entry.target_lang,
        tags=entry.tags
    )


@router.post("/memory/file-entry")
async def add_memory_file_entry(entry: MemoryFileEntryRequest, db: Session = Depends(get_db)):
    memory_file = db.query(KnowledgeFile).filter(KnowledgeFile.id == entry.memory_file_id).first()
    if not memory_file:
        raise HTTPException(status_code=404, detail="记忆库文件不存在")

    if not os.path.exists(memory_file.file_path):
        raise HTTPException(status_code=404, detail="记忆库源文件不存在")

    file_type = str(memory_file.file_type or "").lower()
    if file_type not in WRITABLE_MEMORY_FILE_TYPES:
        raise HTTPException(status_code=400, detail="当前仅支持写入 Excel 记忆库文件（xlsx、xlsm、xltx、xltm）")

    source_text = (entry.source_text or "").strip()
    translated_text = (entry.translated_text or "").strip()
    if not source_text or not translated_text:
        raise HTTPException(status_code=400, detail="原文和译文不能为空")

    try:
        _append_memory_entry_to_excel(
            memory_file=memory_file,
            source_text=source_text,
            translated_text=translated_text,
            source_lang=entry.source_lang,
            target_lang=entry.target_lang,
        )
        memory_file.file_size = os.path.getsize(memory_file.file_path)
        memory_file.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(memory_file)
        _invalidate_memory_file_cache(memory_file.id)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"写入记忆库文件失败: {str(exc)}")

    return {
        "message": "已写入记忆库文件",
        "memory_file_id": memory_file.id,
        "filename": memory_file.filename,
        "file_type": file_type,
    }


@router.delete("/memory/{entry_id}")
async def remove_memory_entry(entry_id: int, db: Session = Depends(get_db)):
    entry = delete_memory_entry(db, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Memory entry not found")
    return {"message": "Memory entry deleted successfully"}


@router.get("/docs", response_model=list[TranslationDocOut])
async def list_translation_docs(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, le=500),
    db: Session = Depends(get_db)
):
    docs = db.query(TranslationDoc).order_by(TranslationDoc.created_at.desc()).offset(skip).limit(limit).all()
    return docs


@router.get("/docs/{doc_id}", response_model=TranslationDocOut)
async def get_translation_doc(doc_id: int, db: Session = Depends(get_db)):
    doc = db.query(TranslationDoc).filter(TranslationDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Translation doc not found")
    return doc


@router.delete("/docs/{doc_id}")
async def delete_translation_doc(doc_id: int, db: Session = Depends(get_db)):
    doc = db.query(TranslationDoc).filter(TranslationDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Translation doc not found")
    db.delete(doc)
    db.commit()
    return {"message": "Translation doc deleted successfully"}
