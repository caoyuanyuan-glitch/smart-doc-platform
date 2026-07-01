import csv
import io
import json
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.crud import polish_learning_rule as crud
from app.schemas.polish_learning_rule import (
    PolishLearningRuleCreate,
    PolishLearningRuleUpdate,
    PolishLearningRuleOut,
    PolishLearningRuleBatchImport,
    PolishLearningRuleBatchDelete,
)

router = APIRouter(prefix="/api/polish-rules", tags=["润色规则管理"])

# 合法的规则分类值
VALID_RULE_TYPES = {"system_rule", "replacement_rule", "forbidden_rule", "sentence_applicability_rule", "imperative_rule", "format_rule"}
FIELD_MAP = {
    "规则名称": "rule_name",
    "规则分类": "rule_type",
    "匹配模式": "match_pattern",
    "替换文本": "replacement_text",
    "说明": "description",
    "优先级": "priority_level",
    "是否启用": "enabled",
}
RULE_TYPE_LABELS = {
    "系统规则": "system_rule",
    "术语替换": "replacement_rule",
    "禁止规则": "forbidden_rule",
    "句式适用": "sentence_applicability_rule",
    "祈使句规则": "imperative_rule",
    "格式规则": "format_rule",
}


def normalize_rule_type(value: str) -> str:
    if not value:
        return value
    text = str(value).strip()
    return RULE_TYPE_LABELS.get(text, text)


def normalize_rule_row(row: dict) -> dict:
    for chinese, english in FIELD_MAP.items():
        if chinese in row:
            row[english] = row.pop(chinese)

    rule_type = normalize_rule_type(row.get("rule_type", ""))
    row["rule_type"] = rule_type
    if rule_type and rule_type not in VALID_RULE_TYPES:
        raise HTTPException(status_code=400, detail=f"无效的规则分类「{rule_type}」，可选: 系统规则、术语替换、禁止规则、句式适用、祈使句规则、格式规则")

    if "priority_level" in row and row.get("priority_level") not in (None, ""):
        row["priority_level"] = int(row.get("priority_level", 0))
    else:
        row["priority_level"] = 0

    if "enabled" in row:
        enabled = row.get("enabled")
        if isinstance(enabled, bool):
            row["enabled"] = enabled
        else:
            row["enabled"] = str(enabled).strip().lower() in ("true", "1", "yes", "是", "启用")
    else:
        row["enabled"] = True

    row["trigger_count"] = int(row.get("trigger_count", 0) or 0)
    return row


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.get("/", response_model=dict)
def list_rules(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    rule_type: str = Query(None),
    enabled: bool = Query(None),
    db: Session = Depends(get_db),
):
    rules = crud.get_rules(db, skip=skip, limit=limit, rule_type=rule_type, enabled=enabled)
    total = crud.count_rules(db, rule_type=rule_type, enabled=enabled)
    return {
        "total": total,
        "items": [PolishLearningRuleOut.model_validate(r).model_dump() for r in rules],
    }


@router.get("/{rule_id}", response_model=PolishLearningRuleOut)
def get_rule(rule_id: int, db: Session = Depends(get_db)):
    rule = crud.get_rule(db, rule_id)
    if not rule:
        raise HTTPException(status_code=404, detail="规则不存在")
    return PolishLearningRuleOut.model_validate(rule)


@router.post("/", response_model=PolishLearningRuleOut)
def create_rule(rule: PolishLearningRuleCreate, db: Session = Depends(get_db)):
    existing = crud.get_rule_by_key(db, rule.rule_key)
    if existing:
        raise HTTPException(status_code=400, detail=f"规则键 {rule.rule_key} 已存在")
    return PolishLearningRuleOut.from_orm(crud.create_rule(db, rule))


@router.put("/{rule_id}", response_model=PolishLearningRuleOut)
def update_rule(rule_id: int, rule_update: PolishLearningRuleUpdate, db: Session = Depends(get_db)):
    updated = crud.update_rule(db, rule_id, rule_update)
    if not updated:
        raise HTTPException(status_code=404, detail="规则不存在")
    return PolishLearningRuleOut.from_orm(updated)


@router.delete("/{rule_id}")
def delete_rule(rule_id: int, db: Session = Depends(get_db)):
    success = crud.delete_rule(db, rule_id)
    if not success:
        raise HTTPException(status_code=404, detail="规则不存在")
    return {"ok": True}


@router.post("/batch-delete")
def batch_delete_rules(body: PolishLearningRuleBatchDelete, db: Session = Depends(get_db)):
    deleted = crud.batch_delete_rules(db, body.ids)
    return {"ok": True, "deleted": deleted}


@router.get("/export/json")
def export_rules_json(
    rule_type: str = Query(None),
    db: Session = Depends(get_db),
):
    rules_data = crud.export_rules(db, rule_type=rule_type)
    content = json.dumps(rules_data, ensure_ascii=False, indent=2)
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=polish_rules.json"},
    )


@router.get("/export/csv")
def export_rules_csv(
    rule_type: str = Query(None),
    db: Session = Depends(get_db),
):
    rules_data = crud.export_rules(db, rule_type=rule_type)
    output = io.StringIO()
    if rules_data:
        writer = csv.DictWriter(output, fieldnames=rules_data[0].keys())
        writer.writeheader()
        writer.writerows(rules_data)
    content = output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=polish_rules.csv"},
    )


def build_rules_workbook(rows: list[dict]):
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        raise HTTPException(status_code=500, detail="Excel 生成依赖不可用")

    headers = ["规则名称", "规则分类", "匹配模式", "替换文本", "说明", "优先级", "是否启用"]
    wb = Workbook()
    ws = wb.active
    ws.title = "润色规则模板"
    ws.append(headers)

    for row in rows:
        ws.append([
            row.get("规则名称") or "",
            row.get("规则分类") or "术语替换",
            row.get("匹配模式") or "",
            row.get("替换文本") or "",
            row.get("说明") or "",
            row.get("优先级") or 0,
            "启用" if row.get("是否启用", True) else "禁用",
        ])

    header_fill = PatternFill("solid", fgColor="E8F1FF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    widths = [18, 16, 30, 30, 40, 12, 12]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    rule_type_validation = DataValidation(type="list", formula1='"系统规则,术语替换,禁止规则,句式适用,祈使句规则,格式规则"', allow_blank=False)
    enabled_validation = DataValidation(type="list", formula1='"启用,禁用"', allow_blank=False)
    ws.add_data_validation(rule_type_validation)
    ws.add_data_validation(enabled_validation)
    rule_type_validation.add("B2:B1000")
    enabled_validation.add("G2:G1000")

    note = wb.create_sheet("填写说明")
    note.append(["字段", "说明"])
    note.append(["规则分类", "请使用下拉选项：系统规则、术语替换、禁止规则、句式适用、祈使句规则、格式规则"])
    note.append(["是否启用", "请使用下拉选项：启用、禁用"])
    note.append(["优先级", "数字越大优先级越高，默认 0"])
    for cell in note[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    note.column_dimensions["A"].width = 18
    note.column_dimensions["B"].width = 70
    return wb


def workbook_response(workbook, filename: str):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}.xlsx"},
    )


@router.get("/export/xlsx")
def export_rules_xlsx(
    rule_type: str = Query(None),
    db: Session = Depends(get_db),
):
    rules_data = crud.export_rules(db, rule_type=rule_type)
    workbook = build_rules_workbook(rules_data)
    return workbook_response(workbook, "润色规则导出")


@router.get("/template/xlsx")
def download_rules_xlsx_template():
    sample = [{
        "规则名称": "示例规则名称",
        "规则分类": "术语替换",
        "匹配模式": "待匹配的原文或正则",
        "替换文本": "替换后的文本",
        "说明": "规则说明（可选）",
        "优先级": 0,
        "是否启用": True,
    }]
    workbook = build_rules_workbook(sample)
    return workbook_response(workbook, "润色规则模板")


@router.post("/import/json")
def import_rules_json(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="仅支持 .json 文件")
    try:
        content = file.file.read().decode("utf-8")
        rules_data = json.loads(content)
    except Exception:
        raise HTTPException(status_code=400, detail="JSON 解析失败")
    if not isinstance(rules_data, list):
        raise HTTPException(status_code=400, detail="JSON 文件应包含规则数组")

    for item in rules_data:
        normalize_rule_row(item)

    result = crud.import_rules(db, rules_data)
    return {"ok": True, **result}


@router.post("/import/csv")
def import_rules_csv(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="仅支持 .csv 文件")
    try:
        content = file.file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        rules_data = list(reader)
    except Exception:
        raise HTTPException(status_code=400, detail="CSV 解析失败")
    if not rules_data:
        raise HTTPException(status_code=400, detail="CSV 文件为空")

    # Convert types and translate headers
    for row in rules_data:
        normalize_rule_row(row)
    result = crud.import_rules(db, rules_data)
    return {"ok": True, **result}


@router.post("/import/xlsx")
def import_rules_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")
    try:
        from openpyxl import load_workbook
        content = file.file.read()
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook["润色规则模板"] if "润色规则模板" in workbook.sheetnames else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 解析失败")

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件为空")

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    rules_data = []
    for values in rows[1:]:
        if not values or not any(v not in (None, "") for v in values):
            continue
        row = {headers[index]: values[index] for index in range(min(len(headers), len(values))) if headers[index]}
        normalize_rule_row(row)
        rules_data.append(row)

    if not rules_data:
        raise HTTPException(status_code=400, detail="Excel 文件没有可导入的数据")

    result = crud.import_rules(db, rules_data)
    return {"ok": True, **result}
