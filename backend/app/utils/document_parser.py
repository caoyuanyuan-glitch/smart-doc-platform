import os
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from docx import Document as DocxDocument
import markdown
from app.utils.file_utils import read_file_safe as _read_file_safe


@dataclass
class TextBlock:
    text: str
    page_num: int
    block_type: str
    bbox: tuple


def _block_sort_key(block: TextBlock):
    return (block.page_num, round(block.bbox[1], 1), round(block.bbox[0], 1))


def _merge_pdf_paragraph_lines(text: str) -> str:
    paragraphs = re.split(r'\n{2,}', text)
    merged = []

    for paragraph in paragraphs:
        lines = [re.sub(r'\s+', ' ', line).strip() for line in paragraph.split('\n') if line.strip()]
        if not lines:
            continue

        current = lines[0]
        for line in lines[1:]:
            if re.match(r'^(?:\d+[\.)]|[-*•])\s*', line):
                current += '\n' + line
                continue
            if len(line.split()) <= 6 and re.match(r'^[A-Z][A-Za-z0-9 /&()_\-]{0,48}$', line):
                current += '\n' + line
                continue
            current += ' ' + line
        merged.append(current)

    return '\n\n'.join(merged)


def clean_pdf_text(text: str) -> str:
    text = str(text or '')
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = text.replace('\f', '\n\f\n')
    text = re.sub(r'(?<=\n)\s*y(?=[A-Za-z])', ' ', text)
    text = re.sub(r'(?<=[A-Za-z])\s*-\s*\n\s*(?=[A-Za-z])', '', text)
    text = re.sub(r'[\x00-\x08\x0b\x0e-\x1f\x7f-\x9f]', '', text)
    text = _merge_pdf_paragraph_lines(text)
    text = re.sub(r'\n*\f\n*', '\f', text)
    text = re.sub(r'[ \t]{2,}', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _extract_pdf_block_text(block: dict) -> str:
    lines = []
    for line in block.get("lines", []):
        line_text = "".join(span.get("text", "") for span in line.get("spans", []))
        line_text = re.sub(r'\s+', ' ', line_text).strip()
        if line_text:
            lines.append(line_text)
    return "\n".join(lines).strip()


def _classify_pdf_block(text: str, block: dict, page_height: float) -> str:
    y0 = float(block.get("bbox", [0, 0, 0, 0])[1] or 0)
    if y0 < max(50.0, page_height * 0.06):
        return 'header'
    if y0 > max(0.0, page_height - 50.0):
        return 'footer'

    lines = block.get("lines", [])
    first_span = lines[0].get("spans", [])[0] if lines and lines[0].get("spans") else {}
    font_size = float(first_span.get("size", 12) or 12)
    font_name = str(first_span.get("font", "") or "")
    is_bold = 'Bold' in font_name
    line_count = len(lines)

    if (is_bold or font_size > 14) and line_count <= 2:
        if any(keyword in text for keyword in ['Table', 'Figure', 'FIGURE']):
            return 'caption'
        return 'title'

    digit_ratio = sum(1 for ch in text if ch.isdigit()) / max(len(text), 1)
    if line_count >= 3 and digit_ratio > 0.15:
        return 'table'

    if digit_ratio > 0.2 and len(text.split()) <= 12:
        return 'table'

    return 'body'


def _looks_like_table_row(blocks):
    if len(blocks) < 2:
        return False

    texts = [block.text.strip() for block in blocks if block.text.strip()]
    if len(texts) < 2:
        return False

    joined = ' '.join(texts)
    short_blocks = sum(1 for text in texts if len(text) <= 40 and '\n' not in text)
    table_like_keywords = re.search(r'\b(?:revision|history|version|date|table|figure|well|tube|sample|reagent)\b', joined, re.IGNORECASE)
    numeric_hits = len(re.findall(r'\d+(?:\.\d+)?', joined))
    unit_hits = len(re.findall(r'\b(?:mL|μL|uL|mg|ng|mm|cm|kg|rpm|min|sec|h)\b', joined, re.IGNORECASE))
    explicit_table_blocks = sum(1 for block in blocks if block.block_type == 'table')

    if explicit_table_blocks >= 1 and short_blocks >= 2:
        return True
    if short_blocks >= 3 and (numeric_hits >= 1 or unit_hits >= 1):
        return True
    if short_blocks >= 2 and table_like_keywords:
        return True
    return False


def _merge_table_like_blocks(blocks, y_threshold=5.0):
    merged = []
    current_group = []
    current_key = None

    def flush_group():
        nonlocal current_group, current_key
        if not current_group:
            return
        row_blocks = sorted(current_group, key=lambda block: block.bbox[0])
        if _looks_like_table_row(row_blocks):
            row_text = '  |  '.join(block.text.strip().replace('\n', ' ') for block in row_blocks if block.text.strip())
            merged.append(TextBlock(
                text=row_text,
                page_num=row_blocks[0].page_num,
                block_type='table_row',
                bbox=row_blocks[0].bbox,
            ))
        else:
            merged.extend(row_blocks)
        current_group = []
        current_key = None

    for block in sorted(blocks, key=_block_sort_key):
        block_key = (block.page_num, round(block.bbox[1] / y_threshold))
        if current_key is None or block_key == current_key:
            current_group.append(block)
            current_key = block_key
            continue
        flush_group()
        current_group = [block]
        current_key = block_key

    flush_group()
    return merged


def extract_pdf(file_path):
    try:
        import fitz
    except ImportError as exc:
        raise ValueError("PDF解析依赖缺失: 请安装 PyMuPDF") from exc

    try:
        doc = fitz.open(file_path)
        blocks = []
        page_texts = []

        for page_num, page in enumerate(doc):
            page_dict = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)
            page_blocks = []
            for block in page_dict.get("blocks", []):
                if "lines" not in block:
                    continue

                block_text = _extract_pdf_block_text(block)
                if not block_text:
                    continue

                block_type = _classify_pdf_block(block_text, block, page.rect.height)
                text_block = TextBlock(
                    text=block_text,
                    page_num=page_num,
                    block_type=block_type,
                    bbox=tuple(block.get("bbox", (0, 0, 0, 0))),
                )
                page_blocks.append(text_block)

            page_blocks = _merge_table_like_blocks(page_blocks)
            page_text = "\n\n".join(block.text for block in sorted(page_blocks, key=_block_sort_key))
            blocks.extend(page_blocks)
            page_texts.append(page_text)

        return {
            'blocks': blocks,
            'page_texts': page_texts,
            'full_text': "\f\n\n".join(page_texts),
        }
    except Exception as e:
        raise ValueError(f"PDF解析失败: {str(e)}")
    finally:
        try:
            doc.close()
        except Exception:
            pass


def parse_pdf(file_path):
    extracted = extract_pdf(file_path)
    page_texts = extracted.get('page_texts', []) or []
    if page_texts:
        cleaned_pages = [clean_pdf_text(page_text) for page_text in page_texts if str(page_text or '').strip()]
        return '\f'.join(cleaned_pages)
    return clean_pdf_text(extracted.get('full_text', ''))

def parse_docx(file_path):
    try:
        doc = DocxDocument(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        raise ValueError(f"DOCX解析失败: {str(e)}")

def parse_markdown(file_path):
    try:
        content = _read_file_safe(file_path)
        return content.strip()
    except Exception as e:
        raise ValueError(f"MD解析失败: {str(e)}")

def parse_dita_zip(file_path):
    try:
        text = ""
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            for name in zip_ref.namelist():
                if name.endswith('.dita') or name.endswith('.xml'):
                    with zip_ref.open(name) as f:
                        content = f.read().decode('utf-8')
                        root = ET.fromstring(content)
                        text += ' '.join(root.itertext()) + "\n"
        return text.strip()
    except Exception as e:
        raise ValueError(f"DITA ZIP解析失败: {str(e)}")

def parse_text(file_path):
    try:
        return _read_file_safe(file_path).strip()
    except Exception as e:
        raise ValueError(f"文本文件解析失败: {str(e)}")

def parse_idml(file_path):
    try:
        text = ""
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            for name in zip_ref.namelist():
                if name.endswith('.xml'):
                    with zip_ref.open(name) as f:
                        content = f.read().decode('utf-8')
                        root = ET.fromstring(content)
                        text += ' '.join(root.itertext()) + "\n"
        return text.strip()
    except Exception as e:
        raise ValueError(f"IDML解析失败: {str(e)}")

def parse_dita(file_path):
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        ns = {"dita": "http://docs.oasis-open.org/dita/v1.1/ns.dita"}
        text = " ".join(root.itertext())
        return text.strip()
    except Exception as e:
        try:
            content = _read_file_safe(file_path)
            root = ET.fromstring(content)
            return " ".join(root.itertext()).strip()
        except Exception:
            raise ValueError(f"DITA解析失败: {str(e)}")

def parse_xlsx(file_path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        texts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            texts.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                row_texts = []
                for cell in row:
                    if cell is not None:
                        row_texts.append(str(cell))
                if row_texts:
                    texts.append("\t".join(row_texts))
        wb.close()
        return "\n".join(texts)
    except Exception as e:
        raise ValueError(f"XLSX解析失败: {str(e)}")

def parse_xls(file_path):
    try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        texts = []
        for sheet in wb.sheets():
            texts.append(f"[Sheet: {sheet.name}]")
            for row_index in range(sheet.nrows):
                row_texts = []
                for col_index in range(sheet.ncols):
                    value = sheet.cell_value(row_index, col_index)
                    if value not in (None, ''):
                        row_texts.append(str(value))
                if row_texts:
                    texts.append("\t".join(row_texts))
        return "\n".join(texts)
    except Exception as e:
        raise ValueError(f"XLS解析失败: {str(e)}")

def parse_pptx(file_path):
    try:
        text = ""
        import re
        with zipfile.ZipFile(file_path, 'r') as zf:
            slide_files = [n for n in zf.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")]
            slide_files.sort(key=lambda x: int(re.search(r'slide(\d+)', x).group(1)))

            for slide_name in slide_files:
                with zf.open(slide_name) as f:
                    content = f.read().decode('utf-8')
                    root = ET.fromstring(content)
                    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
                    slide_texts = []
                    for t_elem in root.iter(f"{{{ns['a']}}}t"):
                        if t_elem.text:
                            slide_texts.append(t_elem.text)
                    if slide_texts:
                        text += "\n".join(slide_texts) + "\n"
        return text.strip()
    except Exception as e:
        raise ValueError(f"PPTX解析失败: {str(e)}")

def parse_xlf(file_path):
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        ns = {"xliff": "urn:oasis:names:tc:xliff:document:1.2"}

        texts = []
        for tu in root.iter("{urn:oasis:names:tc:xliff:document:1.2}trans-unit"):
            source = tu.find("{urn:oasis:names:tc:xliff:document:1.2}source")
            if source is not None and source.text:
                texts.append(source.text.strip())

        if not texts:
            for source in root.iter("{urn:oasis:names:tc:xliff:document:1.2}source"):
                if source.text and source.text.strip():
                    texts.append(source.text.strip())
        
        return "\n".join(texts) if texts else ""
    except Exception as e:
        try:
            content = _read_file_safe(file_path)
            root = ET.fromstring(content)
            texts = []
            xliff_ns = "urn:oasis:names:tc:xliff:document:1.2"
            for tu in root.iter(f"{{{xliff_ns}}}trans-unit"):
                source = tu.find(f"{{{xliff_ns}}}source")
                if source is not None and source.text:
                    texts.append(source.text.strip())
            if not texts:
                for source in root.iter(f"{{{xliff_ns}}}source"):
                    if source.text and source.text.strip():
                        texts.append(source.text.strip())
            return "\n".join(texts) if texts else ""
        except Exception:
            raise ValueError(f"XLF解析失败: {str(e)}")

def parse_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.pdf':
        return parse_pdf(file_path)
    elif ext == '.docx':
        return parse_docx(file_path)
    elif ext == '.md':
        return parse_markdown(file_path)
    elif ext == '.zip':
        return parse_dita_zip(file_path)
    elif ext == '.dita':
        return parse_dita(file_path)
    elif ext == '.xls':
        return parse_xls(file_path)
    elif ext == '.xlsx':
        return parse_xlsx(file_path)
    elif ext == '.pptx':
        return parse_pptx(file_path)
    elif ext == '.xlf':
        return parse_xlf(file_path)
    elif ext == '.txt':
        return parse_text(file_path)
    elif ext == '.idml':
        return parse_idml(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")

def get_file_type(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.pdf':
        return 'pdf'
    elif ext == '.docx':
        return 'docx'
    elif ext == '.md':
        return 'md'
    elif ext == '.zip':
        return 'dita'
    elif ext == '.dita':
        return 'dita'
    elif ext == '.xls':
        return 'xlsx'
    elif ext == '.xlsx':
        return 'xlsx'
    elif ext == '.pptx':
        return 'pptx'
    elif ext == '.xlf':
        return 'xlf'
    elif ext == '.txt':
        return 'txt'
    elif ext == '.idml':
        return 'idml'
    return 'unknown'
