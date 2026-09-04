import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from PIL import Image

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.api.translation import (  # noqa: E402
    _apply_memory_glossary,
    _append_memory_entry_to_delimited_file,
    _append_memory_entry_to_excel,
    _append_memory_match_trace,
    _build_batch_separator,
    _build_memory_candidate_bundle,
    _build_memory_seed_file_path,
    _build_diff_spans,
    _collect_fuzzy_memory_matches,
    _count_translatable_text_units,
    _collect_memory_candidates,
    _do_translate,
    _ensure_memory_bank_entry,
    _find_memory_glossary,
    _filename_looks_non_translatable,
    _get_memory_file_candidates,
    _normalize_memory_file_ids,
    _get_memory_match_trace,
    _get_translate_task_status,
    _get_translation_usage_stats,
    _group_image_ocr_data,
    _detect_confirmation_button_blocks,
    _merge_image_ocr_blocks,
    _blocks_to_ocr_text,
    _ocr_image_file_to_text,
    _clean_ocr_document_text,
    _restore_ocr_alnum_tokens,
    _restore_ocr_latin_letters,
    _load_memory_file_entries,
    _looks_like_hallucination,
    _looks_like_invalid_translation,
    _match_memory_candidates,
    _mark_translation_canceled,
    _memory_match_qualifies_for_stats,
    _normalize_usage_counts,
    _split_batched_translation_output,
    _reset_memory_match_trace,
    _reset_translation_usage_stats,
    _record_passthrough_usage,
    _record_qualified_memory_usage,
    _thread_locals,
    _translate_tasks,
    _translate_tasks_lock,
    _sync_memory_file_to_seed,
    persist_memory_library_seed_if_needed,
    _translate_image,
    _sanitize_translated_filename,
    _translate_filename,
    _is_completed_translation_filename,
    _build_translation_stats_payload,
    translate_with_memory,
)
from app.utils.runtime_paths import runtime_memory_seed_dir  # noqa: E402


class HallucinationTest(unittest.TestCase):
    def test_prompt_leak_detected(self):
        result = "译文如下：这是翻译好的内容"
        self.assertTrue(_looks_like_hallucination(result, "一些原文内容"))

    def test_clean_translation_not_detected(self):
        result = "The system loads the configuration on startup."
        self.assertFalse(_looks_like_hallucination(result, "系统启动时加载配置。"))


class InvalidTranslationTest(unittest.TestCase):
    def test_empty_result_invalid(self):
        self.assertTrue(_looks_like_invalid_translation("", "some source text here", "zh", "en"))

    def test_single_word_ack_invalid(self):
        self.assertTrue(_looks_like_invalid_translation("ok", "这是一段很长的需要翻译的内容，包含了多个句子。", "zh", "en"))

    def test_english_source_to_zh_keeps_no_cjk_invalid(self):
        self.assertTrue(_looks_like_invalid_translation("nothing chinese here at all", "Some english source text here.", "en", "zh"))

    def test_identifier_like_source_to_zh_without_cjk_allowed(self):
        self.assertFalse(_looks_like_invalid_translation("STUM-TT004", "STUM-TT004", "en", "zh"))

    def test_markdown_placeholder_to_zh_without_cjk_allowed(self):
        self.assertFalse(_looks_like_invalid_translation("%%LINK0%%", "%%LINK0%%", "en", "zh"))


class CancelFlowTest(unittest.TestCase):
    def test_mark_canceled_updates_status(self):
        doc_id = 999999
        with _translate_tasks_lock:
            _translate_tasks[doc_id] = {"status": "processing"}
        _mark_translation_canceled(doc_id, message="用户取消")
        self.assertEqual(_get_translate_task_status(doc_id), "canceled")
        with _translate_tasks_lock:
            _translate_tasks.pop(doc_id, None)


class ImageTranslationTest(unittest.TestCase):
    def test_groups_ocr_words_into_ordered_lines(self):
        blocks = _group_image_ocr_data({
            "text": ["Open", "Settings", "Save"],
            "conf": [96, 93, 91],
            "left": [10, 50, 10],
            "top": [20, 20, 60],
            "width": [35, 55, 35],
            "height": [16, 16, 16],
            "block_num": [1, 1, 1],
            "par_num": [1, 1, 1],
            "line_num": [1, 1, 2],
        })

        self.assertEqual([block["text"] for block in blocks], ["Open Settings", "Save"])
        self.assertEqual(blocks[0]["left"], 10)
        self.assertEqual(blocks[0]["width"], 95)

    def test_joins_cjk_words_without_inserting_spaces(self):
        blocks = _group_image_ocr_data({
            "text": ["设置", "页面"],
            "conf": [96, 93],
            "left": [10, 50],
            "top": [20, 20],
            "width": [35, 35],
            "height": [16, 16],
            "block_num": [1, 1],
            "par_num": [1, 1],
            "line_num": [1, 1],
        })

        self.assertEqual(blocks[0]["text"], "设置页面")

    def test_discards_low_confidence_ocr_words(self):
        blocks = _group_image_ocr_data({
            "text": ["garbage", "Valid"],
            "conf": [12, 92],
            "left": [10, 10],
            "top": [20, 60],
            "width": [50, 40],
            "height": [16, 16],
            "block_num": [1, 1],
            "par_num": [1, 1],
            "line_num": [1, 2],
        })

        self.assertEqual([block["text"] for block in blocks], ["Valid"])

    def test_discards_single_character_ocr_noise(self):
        blocks = _group_image_ocr_data({
            "text": ["©", "设置"],
            "conf": [96, 93],
            "left": [10, 10],
            "top": [20, 60],
            "width": [16, 35],
            "height": [16, 16],
            "block_num": [1, 1],
            "par_num": [1, 1],
            "line_num": [1, 2],
        })

        self.assertEqual([block["text"] for block in blocks], ["设置"])

    def test_merges_complementary_ocr_layout_results(self):
        merged = _merge_image_ocr_blocks([
            [{"text": "确认测序对话框", "left": 140, "top": 220, "width": 130, "height": 18, "confidence": 95}],
            [{"text": "是否要测序？", "left": 145, "top": 100, "width": 110, "height": 20, "confidence": 86}],
        ])

        self.assertEqual([block["text"] for block in merged], ["是否要测序？", "确认测序对话框"])

    def test_prefers_higher_confidence_duplicate_ocr_block(self):
        merged = _merge_image_ocr_blocks([
            [{"text": "确认测序", "left": 140, "top": 220, "width": 100, "height": 18, "confidence": 60}],
            [{"text": "确认测序", "left": 142, "top": 220, "width": 100, "height": 18, "confidence": 93}],
        ])

        self.assertEqual(merged, [{"text": "确认测序", "left": 142, "top": 220, "width": 100, "height": 18}])

    def test_recovers_confirmation_button_labels_from_paired_borders(self):
        image = Image.new("L", (400, 240), "white")
        for x in range(80, 180):
            image.putpixel((x, 170), 0)
        for x in range(200, 300):
            image.putpixel((x, 170), 0)
        for x in range(80, 180):
            image.putpixel((x, 190), 0)
        for x in range(200, 300):
            image.putpixel((x, 190), 0)

        blocks = _detect_confirmation_button_blocks(image, [{"text": "是否继续操作？", "left": 140, "top": 90, "width": 100, "height": 20}])

        self.assertEqual([block["text"] for block in blocks], ["否", "是"])
        self.assertEqual([block["left"] for block in blocks], [82, 202])
        self.assertEqual([block["draw_background"] for block in blocks], [True, True])
        self.assertEqual([block["height"] for block in blocks], [16, 16])
        self.assertTrue(all(block["font_size"] <= block["height"] for block in blocks))
        self.assertEqual(blocks[0]["font_size"], blocks[1]["font_size"])

    def test_recovers_english_confirmation_button_labels_from_paired_borders(self):
        image = Image.new("L", (400, 240), "white")
        for x in range(80, 180):
            image.putpixel((x, 170), 0)
            image.putpixel((x, 190), 0)
        for x in range(200, 300):
            image.putpixel((x, 170), 0)
            image.putpixel((x, 190), 0)

        blocks = _detect_confirmation_button_blocks(image, [{"text": "Proceed with sequencing ?", "left": 120, "top": 90, "width": 140, "height": 20}])

        self.assertEqual([block["text"] for block in blocks], ["No", "Yes"])
        self.assertEqual(blocks[0]["font_size"], blocks[1]["font_size"])

    def test_button_blocks_override_tiny_duplicate_ocr_labels(self):
        merged = _merge_image_ocr_blocks([
            [{"text": "Yes", "left": 214, "top": 141, "width": 13, "height": 5, "confidence": 88}],
            [{"text": "Yes", "left": 203, "top": 156, "width": 112, "height": 22, "confidence": 999, "font_size": 14, "draw_background": True, "center_text": True}],
        ])

        self.assertEqual(merged, [{"text": "Yes", "left": 203, "top": 156, "width": 112, "height": 22, "font_size": 14, "draw_background": True, "center_text": True}])

    def test_translates_image_to_same_image_format(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "source.png"
            Image.new("RGB", (180, 80), "white").save(source_path)
            blocks = [{"text": "Open settings", "left": 10, "top": 10, "width": 120, "height": 24}]

            with patch("app.api.translation._extract_image_ocr_blocks", return_value=(Image.open(source_path), blocks)), \
                 patch("app.api.translation._translate_text_items", return_value=["Open configuration"]):
                translated, original, output = _translate_image(str(source_path), "ai", "kimi", "en", "en", db=None)

            self.assertEqual(original, ["Open settings"])
            self.assertEqual(output, ["Open configuration"])
            with Image.open(__import__("io").BytesIO(translated)) as image:
                self.assertEqual(image.format, "PNG")

    def test_blocks_to_ocr_text_keeps_reading_order(self):
        text = _blocks_to_ocr_text([
            {"text": "second", "top": 40, "left": 10},
            {"text": "first", "top": 10, "left": 10},
            {"text": "  ", "top": 80, "left": 10},
        ])
        self.assertEqual(text, "first\nsecond")

    def test_ocr_image_file_to_text_uses_document_ocr(self):
        with patch("app.api.translation._ocr_document_image_to_text", return_value="Hello\nWorld"):
            self.assertEqual(_ocr_image_file_to_text("/tmp/x.png"), "Hello\nWorld")

    def test_ocr_image_file_to_text_falls_back_to_blocks(self):
        with patch("app.api.translation._ocr_document_image_to_text", return_value=""), \
             patch("app.api.translation._extract_image_ocr_blocks", return_value=(None, [
                 {"text": "Hello", "top": 0, "left": 0},
                 {"text": "World", "top": 20, "left": 0},
             ])):
            self.assertEqual(_ocr_image_file_to_text("/tmp/x.png"), "Hello\nWorld")

    def test_ocr_image_file_to_text_raises_when_empty(self):
        with patch("app.api.translation._ocr_document_image_to_text", return_value=""), \
             patch("app.api.translation._extract_image_ocr_blocks", return_value=(None, [])):
            with self.assertRaises(ValueError):
                _ocr_image_file_to_text("/tmp/x.png")

    def test_clean_ocr_document_text_joins_wrapped_lines(self):
        text = _clean_ocr_document_text(
            "1.G99.测序仪支持多种规格芯片上机，定位芯\n: 。 片的位置。。\n\n2.此位置信息会受到影响。"
        )
        self.assertIn("定位芯片的位置。", text)
        self.assertIn("2.此位置信息会受到影响。", text)

    def test_restore_ocr_alnum_prefers_g99_over_699(self):
        text = _restore_ocr_alnum_tokens(
            "1.699.测序仪支持多种规格芯片上机",
            ["1.G99-测序仪支持多种规格芯片上机", "B12 and 812"],
        )
        self.assertIn("G99", text)
        self.assertNotIn("699", text)

    def test_restore_ocr_alnum_keeps_unrelated_numbers(self):
        text = _restore_ocr_alnum_tokens("订单号 1699 已发货", ["型号 G99"])
        self.assertEqual(text, "订单号 1699 已发货")

    def test_restore_ocr_latin_prefers_fov_over_fev(self):
        text = _restore_ocr_latin_letters(
            "投射在晶圆所有 Fev 排布",
            [
                ((237, 159, 186), "投射在晶圆所有 Fev 排布"),
                ((259, 159, 186), "投射在晶圆所有 Fov 排布"),
                ((195, 159, 188), "投射在品圆所有 Fov 排布"),
            ],
        )
        self.assertIn("Fov", text)
        self.assertNotIn("Fev", text)

    def test_document_ocr_keeps_numbered_paragraphs(self):
        fixture = Path(__file__).parent / "fixtures" / "ocr_numbered_paragraphs.png"
        if not fixture.exists():
            self.skipTest("OCR fixture missing")
        text = _ocr_image_file_to_text(str(fixture))
        self.assertIn("测序仪支持多种规格芯片上机", text)
        self.assertIn("芯片晶圆切割", text)
        self.assertIn("多芯片融合的拍照起点定位", text)
        self.assertRegex(text, r"Fov")
        self.assertIn("G99", text)
        self.assertNotIn("699", text)
        self.assertNotIn("Fev", text)
        self.assertNotIn("TALIA", text)
        self.assertNotIn("AML", text)

class HybridMemoryFallbackTest(unittest.TestCase):
    def tearDown(self):
        for attr in ["memory_bank", "memory_file_id", "memory_file_ids", "translation_usage_stats", "memory_candidate_cache"]:
            if hasattr(_thread_locals, attr):
                delattr(_thread_locals, attr)

    def test_hybrid_falls_back_to_partial_memory_when_ai_unavailable(self):
        bundle = _build_memory_candidate_bundle([
            {"source_text": "settings", "translated_text": "设置"}
        ])

        with patch("app.api.translation._get_memory_candidate_bundle", return_value=bundle), \
             patch("app.api.translation.translate_with_ai", side_effect=HTTPException(status_code=500, detail="AI翻译引擎不可用")):
            translated = _do_translate("Open settings page", "hybrid", "kimi", "en", "zh", db=None)

        self.assertEqual(translated, "Open 设置 page")

    def test_hybrid_multi_segment_falls_back_to_memory_when_ai_unavailable(self):
        bundle = _build_memory_candidate_bundle([
            {"source_text": "settings", "translated_text": "设置"}
        ])

        with patch("app.api.translation._get_memory_candidate_bundle", return_value=bundle), \
             patch("app.api.translation.translate_with_ai", side_effect=HTTPException(status_code=500, detail="AI翻译引擎不可用")):
            translated = _do_translate("Open settings page.\nClose settings page.", "hybrid", "kimi", "en", "zh", db=None)

        self.assertEqual(translated, "Open 设置 page.\nClose 设置 page.")


class EnsureMemoryBankEntryTest(unittest.TestCase):
    def test_creates_new_entry_when_missing(self):
        db = unittest.mock.MagicMock()
        db.query.return_value.filter.return_value.first.return_value = None

        entry, created = _ensure_memory_bank_entry(db, "hello", "你好", "en", "zh")

        self.assertTrue(created)
        self.assertEqual(entry.source_text, "hello")
        self.assertEqual(entry.translated_text, "你好")
        db.add.assert_called_once_with(entry)

    def test_skips_duplicate_entry(self):
        db = unittest.mock.MagicMock()
        existing = unittest.mock.MagicMock()
        db.query.return_value.filter.return_value.first.return_value = existing

        entry, created = _ensure_memory_bank_entry(db, "hello", "你好", "en", "zh")

        self.assertFalse(created)
        self.assertIs(entry, existing)
        db.add.assert_not_called()


class MemorySeedSyncTest(unittest.TestCase):
    def test_build_seed_path_preserves_folder_structure(self):
        root = SimpleNamespace(name="资源库", parent=None)
        child = SimpleNamespace(name="记忆库", parent=root)
        memory_file = SimpleNamespace(file_path="/tmp/source.xlsx", name="AI翻译语料写入Excel.xlsx", folder=child)

        seed_path = _build_memory_seed_file_path(memory_file, seed_root=Path("/tmp/seed-root"))

        self.assertEqual(seed_path, Path("/tmp/seed-root/资源库/记忆库/AI翻译语料写入Excel.xlsx"))

    def test_sync_memory_file_to_seed_copies_content(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            source_path = tmp_path / "source.xlsx"
            source_path.write_bytes(b"seed-content")

            root = SimpleNamespace(name="资源库", parent=None)
            child = SimpleNamespace(name="记忆库", parent=root)
            memory_file = SimpleNamespace(file_path=str(source_path), name="AI翻译语料写入Excel.xlsx", folder=child)

            seed_file_path = _sync_memory_file_to_seed(memory_file, seed_root=tmp_path / "seed")

            self.assertTrue(seed_file_path.exists())
            self.assertEqual(seed_file_path.read_bytes(), b"seed-content")
            self.assertEqual(seed_file_path, tmp_path / "seed" / "资源库" / "记忆库" / "AI翻译语料写入Excel.xlsx")

    def test_runtime_memory_seed_dir_is_shared_location(self):
        shared_dir = runtime_memory_seed_dir()
        self.assertIn(".smart-doc-platform", str(shared_dir))

    def test_persist_uploaded_csv_under_memory_library_to_seed(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            source_path = tmp_path / "source.csv"
            source_path.write_text("zh-CN,en-US\n基因,gene\n", encoding="utf-8")
            root = SimpleNamespace(name="资源库", parent=None)
            child = SimpleNamespace(name="记忆库", parent=root)
            memory_file = SimpleNamespace(
                file_path=str(source_path),
                name="admin-upload.csv",
                file_type="csv",
                folder=child,
            )
            seed_file_path = persist_memory_library_seed_if_needed(memory_file, seed_root=tmp_path / "seed")
            self.assertEqual(seed_file_path, tmp_path / "seed" / "资源库" / "记忆库" / "admin-upload.csv")
            self.assertTrue(seed_file_path.exists())
            self.assertIn("基因", seed_file_path.read_text(encoding="utf-8"))

    def test_persist_skips_files_outside_memory_library(self):
        root = SimpleNamespace(name="资源库", parent=None)
        other = SimpleNamespace(name="文件资料", parent=root)
        memory_file = SimpleNamespace(
            file_path="/tmp/manual.pdf",
            name="manual.pdf",
            file_type="pdf",
            folder=other,
        )
        self.assertIsNone(persist_memory_library_seed_if_needed(memory_file, seed_root=Path("/tmp/seed")))


class MemoryNormalizationTest(unittest.TestCase):
    def tearDown(self):
        if hasattr(_thread_locals, "memory_match_trace"):
            delattr(_thread_locals, "memory_match_trace")

    def test_match_ignores_case_spaces_and_special_characters(self):
        bundle = _build_memory_candidate_bundle([
            {"source_text": "Open Settings Page", "translated_text": "打开设置页面"}
        ])

        matched = _match_memory_candidates("open_settings-(page)", bundle)

        self.assertEqual(matched, "打开设置页面")
        self.assertEqual(_get_memory_match_trace()[-1]["reason"], "compact_exact")

    def test_match_preserves_symbols_when_candidate_is_embedded(self):
        bundle = _build_memory_candidate_bundle([
            {"source_text": "DNBSEQ T7 Gene Sequencer", "translated_text": "DNBSEQ T7 基因测序仪"}
        ])

        matched = _match_memory_candidates("[DNBSEQ-T7] Gene_Sequencer", bundle)

        self.assertEqual(matched, "DNBSEQ T7 基因测序仪")

    def test_match_hits_embedded_phrase_inside_longer_sentence(self):
        bundle = _build_memory_candidate_bundle([
            {"source_text": "open settings page", "translated_text": "打开设置页面"}
        ])

        matched = _match_memory_candidates("Please open - settings_page now.", bundle)

        self.assertEqual(matched, "Please 打开设置页面 now.")

    def test_match_prefers_longer_embedded_phrase(self):
        bundle = _build_memory_candidate_bundle([
            {"source_text": "settings", "translated_text": "设置"},
            {"source_text": "open settings page", "translated_text": "打开设置页面"}
        ])

        matched = _match_memory_candidates("Please open settings page now", bundle)

        self.assertEqual(matched, "Please 打开设置页面 now")

    def test_glossary_prefers_longer_phrases_and_replaces_multiple_terms(self):
        glossary = _find_memory_glossary(
            "Please open settings page and close settings page.",
            [
                {"source_text": "settings", "translated_text": "设置"},
                {"source_text": "open settings page", "translated_text": "打开设置页面"},
                {"source_text": "close settings page", "translated_text": "关闭设置页面"},
            ],
        )

        translated, replaced = _apply_memory_glossary(
            "Please open_settings-page and close settings page.",
            glossary,
        )

        self.assertTrue(replaced)
        self.assertEqual(translated, "Please 打开设置页面 and 关闭设置页面.")

    def test_match_prefers_more_specific_candidate_among_similar_terms(self):
        bundle = _build_memory_candidate_bundle([
            {"source_text": "gene sequencer", "translated_text": "测序仪"},
            {"source_text": "DNBSEQ T7 gene sequencer", "translated_text": "DNBSEQ T7 基因测序仪"},
        ])

        matched = _match_memory_candidates("The DNBSEQ-T7 Gene_Sequencer is ready.", bundle)

        self.assertEqual(matched, "The DNBSEQ T7 基因测序仪 is ready.")
        trace = _get_memory_match_trace()[-1]
        self.assertEqual(trace["reason"], "token_subsequence")
        self.assertEqual(trace["candidate_text"], "DNBSEQ T7 gene sequencer")

    def test_match_ignores_locale_suffix_metadata(self):
        bundle = _build_memory_candidate_bundle([
            {
                "source_text": "DNBSEQ-T7+RS Genetic Sequencer System Guide_Chinese_RUO_WH",
                "translated_text": "DNBSEQ-T7+RS 基因测序仪系统指南_中文_RUO_WH",
            }
        ])

        matched = _match_memory_candidates(
            "DNBSEQ-T7+RS Genetic Sequencer System Guide_English_RUO_WH",
            bundle,
        )

        self.assertEqual(matched, "DNBSEQ-T7+RS 基因测序仪系统指南_中文_RUO_WH")
        self.assertEqual(_get_memory_match_trace()[-1]["reason"], "metadata_exact")

    def test_match_ignores_version_metadata_variants(self):
        bundle = _build_memory_candidate_bundle([
            {
                "source_text": "Mammoth COOLING Control Board V3.0.0",
                "translated_text": "Mammoth 冷却控制板 V3.0.0",
            }
        ])

        matched = _match_memory_candidates("Mammoth COOLING Control Board V3.0.0.0", bundle)

        self.assertEqual(matched, "Mammoth 冷却控制板 V3.0.0")
        self.assertEqual(_get_memory_match_trace()[-1]["reason"], "metadata_exact")


    def test_fuzzy_sentence_match_accepts_eighty_percent_similarity(self):
        bundle = _build_memory_candidate_bundle([
            {
                "source_text": "参数确认无误后，点击运行。",
                "translated_text": "After confirming the parameters, click Run.",
            }
        ])

        matched = _match_memory_candidates("确认参数无误后点击运行。", bundle, threshold=0.8)

        self.assertEqual(matched, "After confirming the parameters, click Run.")
        self.assertGreaterEqual(_get_memory_match_trace()[-1]["score"], 0.8)

    def test_fuzzy_sentence_match_works_when_partial_preserve_is_disabled(self):
        bundle = _build_memory_candidate_bundle([
            {
                "source_text": "请在样本制备卡准备界面点击下一步。",
                "translated_text": "Click Next on the sample prep card screen.",
            }
        ])

        matched = _match_memory_candidates(
            "请在样品制备卡准备界面点击下一步。",
            bundle,
            threshold=0.8,
            preserve_sentence_unmatched=False,
        )

        self.assertEqual(matched, "Click Next on the sample prep card screen.")

    def test_fuzzy_sentence_match_ignores_terminal_punctuation_kind_mismatch(self):
        bundle = _build_memory_candidate_bundle([
            {
                "source_text": "点击运行按钮开始实验",
                "translated_text": "Click Run to start the experiment.",
            }
        ])

        matched = _match_memory_candidates("点击运行按钮开始实验。", bundle, threshold=0.8)

        self.assertEqual(matched, "Click Run to start the experiment.")

    def test_collect_fuzzy_memory_matches_includes_score_and_diff_spans(self):
        _reset_memory_match_trace()
        bundle = _build_memory_candidate_bundle([
            {
                "source_text": "参数确认无误后，点击运行。",
                "translated_text": "After confirming the parameters, click Run.",
            }
        ])

        matched = _match_memory_candidates("确认参数无误后点击运行。", bundle, threshold=0.8)
        details = _collect_fuzzy_memory_matches()

        self.assertEqual(matched, "After confirming the parameters, click Run.")
        self.assertEqual(len(details), 1)
        self.assertGreater(details[0].match_rate, 0)
        self.assertLess(details[0].match_rate, 100)
        self.assertTrue(any(span.tag != "equal" for span in details[0].source_spans))

    def test_collect_fuzzy_memory_matches_shows_contained_longer_memory_sentence(self):
        _reset_memory_match_trace()
        source = "结合不同的芯片拍照预设设置的起点位置（参考起始点）"
        candidate = "多flow cell融合场景下成像起点位置按如下方式确定：结合不同的芯片拍照预设设置的起点位置（参考起始点）"
        translated = (
            "The imaging starting point in multi-flow cell fusion scenario is positioned "
            "using the following manner: According to the pre-configured imaging starting positions."
        )
        _append_memory_match_trace(source, candidate, "similarity_ranked", score=1.0, translated_text=translated)

        details = _collect_fuzzy_memory_matches()

        self.assertEqual(len(details), 1)
        self.assertLess(details[0].match_rate, 100)
        self.assertGreater(details[0].match_rate, 0)
        self.assertTrue(any(span.tag != "equal" for span in details[0].candidate_spans))

    def test_collect_fuzzy_memory_matches_skips_true_100_percent_compact_match(self):
        _reset_memory_match_trace()
        _append_memory_match_trace(
            "点击运行。",
            "点击运行。",
            "normalized_exact",
            score=1.0,
            translated_text="Click Run.",
        )

        self.assertEqual(_collect_fuzzy_memory_matches(), [])

    def test_build_diff_spans_marks_replaced_tokens(self):
        left_spans, right_spans = _build_diff_spans("请在样本制备卡准备界面点击下一步。", "请在样品制备卡准备界面点击下一步。")

        self.assertTrue(any(span["tag"] in {"replace", "delete"} for span in left_spans))
        self.assertTrue(any(span["tag"] in {"replace", "insert"} for span in right_spans))


class TranslationStatisticsTest(unittest.TestCase):
    def test_count_translatable_text_units_skips_codes_versions_and_locale_tags(self):
        count = _count_translatable_text_units(
            "DNBSEQ-T7+RS Genetic Sequencer System Guide_English_RUO_WH V3.0.0"
        )

        self.assertEqual(count, 4)


class MemoryFileSelectionTest(unittest.TestCase):
    def test_normalize_memory_file_ids_merges_and_deduplicates(self):
        self.assertEqual(_normalize_memory_file_ids(memory_file_ids=[3, "2", 3, 0, None], memory_file_id=2), [3, 2])

    def test_collect_memory_candidates_merges_multiple_files_in_order(self):
        db = unittest.mock.MagicMock()
        db.query.return_value.filter.return_value.all.return_value = []

        with patch("app.api.translation._get_memory_file_candidates") as get_candidates:
            get_candidates.side_effect = [
                [
                    {"source_text": "Alpha", "translated_text": "阿尔法"},
                    {"source_text": "Beta", "translated_text": "贝塔"},
                ],
                [
                    {"source_text": "Beta", "translated_text": "贝塔"},
                    {"source_text": "Gamma", "translated_text": "伽马"},
                ],
            ]

            candidates = _collect_memory_candidates(db, "en", "zh", memory_file_ids=[11, 12])

        self.assertEqual(
            candidates,
            [
                {"source_text": "Alpha", "translated_text": "阿尔法"},
                {"source_text": "Beta", "translated_text": "贝塔"},
                {"source_text": "Gamma", "translated_text": "伽马"},
            ],
        )


class BatchedTranslationHelpersTest(unittest.TestCase):
    def test_build_batch_separator_avoids_existing_marker(self):
        separator = _build_batch_separator(["alpha [[MC_DOCSEG]] beta", "gamma"])

        self.assertNotEqual(separator.strip(), "[[MC_DOCSEG]]")
        self.assertNotIn(separator.strip(), "alpha [[MC_DOCSEG]] beta")

    def test_split_batched_translation_output_uses_exact_separator(self):
        separator = _build_batch_separator(["第一段", "第二段"])

        parts = _split_batched_translation_output(
            f"译文一{separator}译文二",
            separator,
            2,
            "batch_split_error",
        )

        self.assertEqual(parts, ["译文一", "译文二"])

    def test_split_batched_translation_output_rejects_missing_separator(self):
        with self.assertRaises(ValueError):
            _split_batched_translation_output("only one part", "\n[[MC_DOCSEG]]\n", 2, "batch_split_error")


class MemoryFileWriteTest(unittest.TestCase):
    def test_append_csv_memory_entry_respects_language_headers(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "memory.csv"
            csv_path.write_bytes("zh-CN,en-US\n已有译文,Existing Source\n".encode("gb18030"))
            memory_file = SimpleNamespace(file_path=str(csv_path), file_type="csv")

            _append_memory_entry_to_delimited_file(
                memory_file,
                source_text="New Source",
                translated_text="新增译文",
                source_lang="en",
                target_lang="zh",
            )

            entries = _load_memory_file_entries(str(csv_path), "csv")
            self.assertTrue(any(item.get("bilingual_values", {}).get("en") == "New Source" and item.get("bilingual_values", {}).get("zh") == "新增译文" for item in entries))

            db = unittest.mock.MagicMock()
            db.query.return_value.filter.return_value.first.return_value = SimpleNamespace(
                id=1,
                file_path=str(csv_path),
                file_type="csv",
                updated_at=None,
            )
            candidates = _get_memory_file_candidates(db, 1, "en", "zh")
            self.assertTrue(any(item["source_text"] == "New Source" and item["translated_text"] == "新增译文" for item in candidates))

            content = csv_path.read_bytes().decode("gb18030")
            self.assertIn("新增译文,New Source", content)

    def test_append_excel_memory_entry_respects_language_headers(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "memory.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "zh-CN"
            ws.cell(row=1, column=2).value = "en-US"
            wb.save(xlsx_path)
            wb.close()

            memory_file = SimpleNamespace(file_path=str(xlsx_path), file_type="xlsx")
            _append_memory_entry_to_excel(
                memory_file,
                source_text="New Source",
                translated_text="新增译文",
                source_lang="en",
                target_lang="zh",
            )

            workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            worksheet = workbook.active
            self.assertEqual(worksheet.cell(row=2, column=1).value, "新增译文")
            self.assertEqual(worksheet.cell(row=2, column=2).value, "New Source")
            workbook.close()


class FilenameTranslationTest(unittest.TestCase):
    def test_sanitize_accepts_long_but_reasonable_titles(self):
        original = "DNBSEQ-T7RS High-throughput Sequencing Set User Manual"
        translated = "DNBSEQ-T7RS 高通量测序试剂套装用户手册"
        self.assertEqual(_sanitize_translated_filename(translated, original), translated)

    def test_sanitize_rejects_oversized_hallucination(self):
        original = "Manual"
        translated = "x" * 400
        self.assertIsNone(_sanitize_translated_filename(translated, original))

    def test_chinese_title_is_translatable(self):
        self.assertFalse(
            _filename_looks_non_translatable("DNBSEQ-T7基因测序仪系统使用说明书", "zh")
        )

    def test_memory_hit_is_used_before_ai(self):
        db = SimpleNamespace()
        with patch("app.api.translation.translate_with_memory", return_value=("基因测序仪说明书", True)), \
             patch("app.api.translation.ai_client.chat") as chat:
            result = _translate_filename(
                "Gene Sequencer Manual",
                "en",
                "zh",
                model="qwen",
                engine="hybrid",
                db=db,
            )
        self.assertEqual(result, "基因测序仪说明书")
        chat.assert_not_called()

    def test_memory_engine_keeps_original_when_no_hit(self):
        db = SimpleNamespace()
        with patch("app.api.translation.translate_with_memory", return_value=("Gene Sequencer Manual", False)):
            result = _translate_filename(
                "Gene Sequencer Manual",
                "en",
                "zh",
                engine="memory",
                db=db,
            )
        self.assertEqual(result, "Gene Sequencer Manual")


class CompletedTranslationListTest(unittest.TestCase):
    def test_completed_filename_detection(self):
        self.assertTrue(_is_completed_translation_filename("说明书.docx"))
        self.assertFalse(_is_completed_translation_filename(""))
        self.assertFalse(_is_completed_translation_filename("ERROR:timeout"))
        self.assertFalse(_is_completed_translation_filename("CANCELED:stopped"))


class TranslationStatsAggregationTest(unittest.TestCase):
    def test_normalize_usage_counts_keeps_uncategorized_remainder(self):
        self.assertEqual(
            _normalize_usage_counts(100, 40, 20),
            {"source_count": 100, "ai_count": 40, "memory_count": 20},
        )

    def test_passthrough_usage_is_not_counted_as_memory(self):
        _reset_translation_usage_stats()
        _record_passthrough_usage("这是一段未匹配的原文内容")
        stats = _get_translation_usage_stats()
        self.assertEqual(stats["memory_word_count"], 0)
        self.assertEqual(stats["ai_word_count"], 0)

    def test_low_score_trace_does_not_qualify_for_memory_stats(self):
        _reset_memory_match_trace()
        _append_memory_match_trace("源句", "候选句", "token_subsequence", score=0.4, translated_text="candidate")
        self.assertFalse(_memory_match_qualifies_for_stats("源句"))

    def test_high_score_trace_qualifies_for_memory_stats(self):
        _reset_memory_match_trace()
        _append_memory_match_trace("源句", "候选句", "similarity_ranked", score=0.86, translated_text="candidate")
        self.assertTrue(_memory_match_qualifies_for_stats("源句"))

    def test_glossary_partial_hit_does_not_record_memory_stats(self):
        bundle = _build_memory_candidate_bundle([
            {
                "source_text": "基因",
                "translated_text": "gene",
                "source_lang": "zh",
                "target_lang": "en",
                "priority": 1,
            }
        ])
        _reset_translation_usage_stats()
        _reset_memory_match_trace()
        with patch("app.api.translation._get_memory_candidate_bundle", return_value=bundle):
            result, hit = translate_with_memory("检测基因序列", "zh", "en", db=SimpleNamespace())
        self.assertTrue(hit)
        self.assertIn("gene", result)
        _record_qualified_memory_usage("检测基因序列")
        self.assertEqual(_get_translation_usage_stats()["memory_word_count"], 0)

    def test_stats_payload_uses_all_records_for_total_word_count(self):
        class FakeQuery:
            def filter(self, *args, **kwargs):
                return self

            def order_by(self, *args, **kwargs):
                return self

            def first(self):
                return None

        class FakeDB:
            def query(self, *args, **kwargs):
                return FakeQuery()

        def fake_summary(db, file_type=None, batch_id=None):
            if file_type == "text":
                return {"doc_count": 2, "doc_word_count": 30, "ai_word_count": 10, "memory_word_count": 8}
            if file_type == "file":
                return {"doc_count": 4, "doc_word_count": 70, "ai_word_count": 40, "memory_word_count": 20}
            return {"doc_count": 6, "doc_word_count": 100, "ai_word_count": 50, "memory_word_count": 28}

        with patch("app.api.translation._refresh_missing_translation_doc_word_counts"), \
             patch("app.api.translation._query_translation_doc_summary", side_effect=fake_summary):
            payload = _build_translation_stats_payload(FakeDB(), None)

        self.assertEqual(payload["total_word_count"], 100)
        self.assertEqual(payload["doc_count"], 4)
        self.assertEqual(payload["ai_word_count"], 50)
        self.assertEqual(payload["memory_word_count"], 28)
        self.assertEqual(payload["text_word_count"], 30)
        self.assertEqual(payload["text_ai_word_count"], 10)
        self.assertEqual(payload["text_memory_word_count"], 8)


if __name__ == "__main__":
    unittest.main()
