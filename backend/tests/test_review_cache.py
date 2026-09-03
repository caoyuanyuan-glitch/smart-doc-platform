import json
import asyncio
import re
from types import SimpleNamespace

from app.api import review as review_api
from app.api import review_rules
from app.crud import rule as crud_rule
from app.crud import review as crud_review
from app.review_engine import pipeline as review_pipeline
from app.review_engine import validation as review_validation


def test_normalize_providers_keeps_single_provider_only():
    assert review_api._normalize_providers(provider="qwen") == ["qwen"]
    assert review_api._normalize_providers(providers="qwen,kimi,deepseek") == ["qwen"]
    assert review_api._normalize_providers(provider="qwen", providers="kimi,deepseek") == ["kimi"]


def test_build_review_cache_key_changes_with_mode(monkeypatch):
    monkeypatch.setattr(review_api, "_review_cache_version", lambda: "v1")
    document = SimpleNamespace(
        id=1,
        filename="demo.docx",
        file_type="docx",
        file_size=128,
        content="same content",
    )

    rule_key = review_api._build_review_cache_key(document, "rule")
    hybrid_key = review_api._build_review_cache_key(document, "hybrid")

    assert rule_key != hybrid_key


def test_build_review_cache_key_isolates_visual_input(monkeypatch):
    monkeypatch.setattr(review_api, "_review_cache_version", lambda: "v1")
    document = SimpleNamespace(
        id=1,
        filename="demo.docx",
        file_type="docx",
        file_size=128,
        content="same content",
    )
    visual_a = SimpleNamespace(id=2, filename="a.pdf", file_type="pdf", file_size=10, content="pdf-a")
    visual_b = SimpleNamespace(id=3, filename="b.pdf", file_type="pdf", file_size=11, content="pdf-b")
    single = review_api._build_review_cache_key(document, "hybrid")
    paired_a = review_api._build_review_cache_key(document, "hybrid", visual_document=visual_a, pairing_confirmed=True)
    paired_b = review_api._build_review_cache_key(document, "hybrid", visual_document=visual_b, pairing_confirmed=True)
    assert single != paired_a
    assert paired_a != paired_b


def test_build_ai_chunk_cache_key_changes_with_document_name(monkeypatch):
    monkeypatch.setattr(review_api, "_review_cache_version", lambda: "v1")
    monkeypatch.setattr(review_api, "_ai_provider_cache_fingerprint", lambda: "provider-v1")

    key_a = review_api._build_ai_chunk_cache_key("same content", "cn", "same basis", document_name="A.pdf")
    key_b = review_api._build_ai_chunk_cache_key("same content", "cn", "same basis", document_name="B.pdf")

    assert key_a != key_b


def test_provider_status_uses_health_check_available_providers(monkeypatch):
    def fake_provider_status(include_health=False):
        assert include_health is True
        return {
            "default_provider": "qwen",
            "available": ["kimi"],
            "health": {
                "healthy": True,
                "ok_providers": 1,
                "total_providers": 2,
                "primary": "kimi",
                "primary_status": "ok",
                "providers": {
                    "qwen": {"status": "error", "model": "qwen-max", "error": "401"},
                    "kimi": {"status": "ok", "model": "moonshot-v1-8k", "latency_ms": 42},
                },
            },
        }

    monkeypatch.setattr(review_api.ai_client, "provider_status", fake_provider_status)

    payload = asyncio.run(review_api.get_provider_status(None))

    assert payload["available_providers"] == ["kimi"]
    assert payload["models"] == [{
        "name": "kimi",
        "label": "Kimi (Moonshot)",
        "available": True,
        "status": "ok",
        "latency_ms": 42,
        "model": "moonshot-v1-8k",
    }]
    assert payload["health"]["ok_providers"] == 1


def test_pdf_page_metadata_collects_suspicious_table_pages(monkeypatch):
    document = SimpleNamespace(file_type="pdf", filename="demo.pdf")
    monkeypatch.setattr(review_api, "_get_document_upload_path", lambda document: review_api.PROJECT_ROOT / "backend")
    monkeypatch.setattr(
        review_api,
        "extract_pdf",
        lambda path: {
            "page_texts": [
                "版本  日期\n2.0  2026年7月15日\n1.0  2025年11月17日\n3",
                "正文内容\n图 1 系统组成\nDoc. No. H-020-001246-00",
            ],
            "blocks": [
                SimpleNamespace(page_num=0, block_type="table_row"),
                SimpleNamespace(page_num=0, block_type="table_row"),
                SimpleNamespace(page_num=1, block_type="caption"),
            ],
        },
    )

    metadata = review_api._extract_pdf_page_metadata(document, "fallback")
    suspicious = review_api._collect_pdf_suspicious_pages(metadata)

    assert metadata["enabled"] is True
    assert metadata["source"] == "pdf_blocks"
    assert metadata["page_count"] == 2
    assert suspicious["candidate_count"] == 2
    first_page = suspicious["candidates"][0]
    assert first_page["page_number"] == 1
    assert "table_layout" in first_page["reasons"]


def test_run_filename_audit_skips_known_product_output_filename():
    issues = review_api._run_filename_audit(
        "ROINAMEquantification.csv",
        "Export file generated by software",
        "en",
    )

    assert issues == []


def test_run_filename_audit_skips_revision_and_absent_cover_product():
    filename = "H-020-001368-00 MGIDL-G99RS 全自动样本加载仪快速操作指南_中文_RUO_WH-R02.pdf"
    content = "触摸屏显示用户交互界面。长按【长按启动】按钮旁的指纹区域。"

    issues = review_api._run_filename_audit(filename, content, "cn")

    assert not any(issue["rule"] == "FILENAME-002" for issue in issues)


def test_run_filename_audit_flags_conflicting_product_model():
    issues = review_api._run_filename_audit(
        "DNBSEQ-T7 操作指南.pdf",
        "本指南适用于 DNBSEQ-T10 测序仪。",
        "cn",
    )

    issue = next(item for item in issues if item["rule"] == "FILENAME-002")
    assert issue["original_text"] == "DNBSEQ-T7"
    assert "DNBSEQ-T10" in issue["description"]


def test_resolve_compare_row_level_accepts_email_only_support_contact_for_english_manuals():
    field_def = {
        "name": "技术支持联系方式（电话/邮箱）",
        "priority": "P1",
    }
    reference_summary = {
        "status": "stable",
        "values": ["邮箱: support@example.com"],
    }

    level, conclusion = review_api._resolve_compare_row_level(field_def, "", reference_summary)

    assert level == "一致"
    assert "仅保留邮箱联系方式" in conclusion


def test_review_cache_version_tracks_review_basis_files():
    for path in review_api.REVIEW_BASIS_VERSION_FILES:
        assert path in review_api.REVIEW_CACHE_VERSION_FILES

    assert review_api.PROJECT_ROOT / "backend" / "app" / "crud" / "rule.py" in review_api.REVIEW_CACHE_VERSION_FILES
    assert review_api.PROJECT_ROOT / "backend" / "seed" / "review_rule_library_seed.json" in review_api.REVIEW_CACHE_VERSION_FILES


def test_select_relevant_ai_review_basis_prefers_es_sections(monkeypatch):
    basis_sections = [
        {
            "label": "中文技术文档写作风格指南",
            "text": "【中文技术文档写作风格指南】\n统一术语表达。",
            "priority": 4,
            "language": "cn",
            "basis_type": "style_guide",
        },
        {
            "label": "说明书发布前自检 Checklist",
            "text": "【说明书发布前自检 Checklist】\n检查版本记录与引用完整性。",
            "priority": 5,
            "language": "both",
            "basis_type": "checklist",
            "is_checklist": True,
        },
    ]

    monkeypatch.setattr(
        review_api,
        "_fetch_es_relevant_basis_sections",
        lambda content, sections, document_language, limit=4: [
            {
                "label": "说明书发布前自检 Checklist",
                "text": "【说明书发布前自检 Checklist】\n重点检查版本记录日期与图表引用。",
                "priority": 5,
                "language": "both",
                "basis_type": "checklist",
                "is_checklist": True,
            }
        ],
    )

    selected = review_api._select_relevant_ai_review_basis("请检查版本记录日期与图表引用", basis_sections, document_language="cn")

    assert "重点检查版本记录日期与图表引用" in selected
    assert "检查版本记录与引用完整性" not in selected


def test_select_relevant_ai_review_basis_uses_es_as_primary_layer(monkeypatch):
    basis_sections = [
        {
            "label": "技术文档常见错误清单",
            "text": "【技术文档常见错误清单】\n避免术语不一致和引用缺失。",
            "priority": 3,
            "language": "both",
            "basis_type": "common_errors",
        },
        {
            "label": "中文技术文档写作风格指南",
            "text": "【中文技术文档写作风格指南】\n统一术语表达。",
            "priority": 4,
            "language": "cn",
            "basis_type": "style_guide",
        },
    ]

    monkeypatch.setattr(
        review_api,
        "_fetch_es_relevant_basis_sections",
        lambda content, sections, document_language, limit=4: [
            {
                "label": "CYY人工审核经验基线-术语",
                "text": "【CYY人工审核经验基线-术语】\n术语章节重点关注缩写首次定义是否完整。",
                "priority": 3,
                "language": "both",
                "basis_type": "human_baseline",
                "is_cyy_example": True,
            }
        ],
    )

    selected = review_api._select_relevant_ai_review_basis("缩写首次定义缺失", basis_sections, document_language="cn")

    assert "缩写首次定义是否完整" in selected
    assert "避免术语不一致和引用缺失" not in selected
    assert "统一术语表达" not in selected


def test_select_relevant_ai_review_basis_falls_back_without_es(monkeypatch):
    monkeypatch.setattr(review_api, "_fetch_es_relevant_basis_sections", lambda *args, **kwargs: [])
    basis_sections = [
        {
            "label": "技术文档常见错误清单",
            "text": "【技术文档常见错误清单】\n避免术语不一致和引用缺失。",
            "priority": 3,
            "language": "both",
            "basis_type": "common_errors",
        }
    ]

    selected = review_api._select_relevant_ai_review_basis("术语不一致，引用缺失", basis_sections, document_language="cn")

    assert "技术文档常见错误清单" in selected
    assert "术语不一致和引用缺失" in selected


def test_find_cached_completed_review_matches_cache_key(monkeypatch):
    expected_key = "cache-key-1"
    monkeypatch.setattr(review_api, "_build_review_cache_key", lambda document, mode, **kwargs: expected_key)

    matched_review = SimpleNamespace(
        id=12,
        status="completed",
        total_issues=3,
        summary=json.dumps({"cache_key": expected_key, "total": 3}),
    )
    unmatched_review = SimpleNamespace(
        id=11,
        status="completed",
        total_issues=5,
        summary=json.dumps({"cache_key": "other-key", "total": 5}),
    )
    monkeypatch.setattr(review_api, "get_reviews", lambda db, document_id, limit=20: [unmatched_review, matched_review])

    review, summary = review_api._find_cached_completed_review(None, SimpleNamespace(id=7), "rule")

    assert review.id == 12
    assert summary["total"] == 3


def test_seed_external_review_rules_updates_existing_rule(monkeypatch):
    payload = {
        "source": "外部评审规则库",
        "export_date": "2026-08-17",
        "rules": [{
            "rule_id": "R013",
            "rule_content": "标点符号使用必须符合规范",
            "category": "格式",
            "severity": "一般",
            "applicable_scenarios": ["PDF"],
        }],
    }
    existing = SimpleNamespace(
        rule_no="EXT-R013",
        category="旧分类",
        description="旧规则",
        regex="old-regex",
        example="old-example",
        suggestion="old-suggestion",
        audit_basis="old-basis",
        severity="serious",
        language="cn",
    )
    commits = []
    db = SimpleNamespace(add=lambda item: None, commit=lambda: commits.append(True))

    monkeypatch.setattr(
        crud_rule,
        "REVIEW_RULE_LIBRARY_SEED_PATH",
        SimpleNamespace(
            exists=lambda: True,
            read_text=lambda encoding="utf-8": json.dumps(payload, ensure_ascii=False),
        ),
    )
    monkeypatch.setattr(crud_rule, "get_rule_by_no", lambda db_obj, rule_no: existing if rule_no == "EXT-R013" else None)

    changed = crud_rule.seed_external_review_rules(db)

    assert changed == 1
    assert existing.description == "标点符号使用必须符合规范"
    assert existing.regex == "(?!)"
    assert existing.severity == "general"
    assert existing.language == "both"
    assert commits == [True]


def test_convert_rule_content_to_regex_disables_ui_bracket_semantic_rule():
    regex = crud_rule._convert_rule_content_to_regex(
        "仅可交互UI元素（按钮、菜单、输入框、选项等）使用【】标注，界面名称/标题（如主界面、测序界面等）不使用【】"
    )

    assert regex == "(?!)"


def test_convert_rule_content_to_regex_disables_quote_semantic_rule():
    regex = crud_rule._convert_rule_content_to_regex(
        "文档中所有需要使用引号的场景统一使用双引号，不得混用单引号"
    )

    assert regex == "(?!)"


def test_review_rules_do_not_hardcode_single_quote_or_screen_name_as_errors():
    punctuation_patterns = {rule["pattern"] for rule in review_rules.CHINESE_PUNCTUATION_RULES}
    terminology_patterns = {rule["pattern"] for rule in review_rules.CHINESE_TERMINOLOGY_RULES}

    assert r"[\u2018\u2019]" not in punctuation_patterns
    assert r"【测序界面】" not in terminology_patterns


def test_system_prompt_keeps_ui_bracket_guidance_without_forcing_screen_names():
    prompt = review_rules.SYSTEM_PROMPT_TEMPLATE

    assert "仅按钮、菜单、输入框、选项等可交互UI元素使用【】标注；界面名称和标题保持原文写法" in prompt
    assert "单引号仅在明确影响语义或格式规范时报告" in prompt


def test_find_cached_completed_review_skips_non_completed(monkeypatch):
    monkeypatch.setattr(review_api, "_build_review_cache_key", lambda document, mode, **kwargs: "cache-key-1")
    reviews = [
        SimpleNamespace(id=21, status="running", summary=json.dumps({"cache_key": "cache-key-1"})),
        SimpleNamespace(id=22, status="failed", summary=json.dumps({"cache_key": "cache-key-1"})),
    ]
    monkeypatch.setattr(review_api, "get_reviews", lambda db, document_id, limit=20: reviews)

    review, summary = review_api._find_cached_completed_review(None, SimpleNamespace(id=7), "rule")

    assert review is None
    assert summary is None


def test_list_reviews_batches_document_lookup(monkeypatch):
    reviews = [
        SimpleNamespace(id=3, document_id=11, status="running", summary="", total_issues=0),
        SimpleNamespace(id=2, document_id=12, status="completed", summary="{}", total_issues=4),
    ]
    documents = [
        SimpleNamespace(id=11, filename="a.docx", file_type="docx"),
        SimpleNamespace(id=12, filename="b.pdf", file_type="pdf"),
    ]
    calls = []

    monkeypatch.setattr(review_api, "_query_review_rows", lambda *args, **kwargs: reviews)
    monkeypatch.setattr(review_api, "_count_review_rows", lambda *args, **kwargs: len(reviews))
    monkeypatch.setattr(review_api, "_reconcile_review_runtime_state", lambda db, review: review)
    monkeypatch.setattr(review_api, "get_progress", lambda review_id: {"status": "running", "progress": 35, "message": "处理中"})

    def fake_get_documents_by_ids(db, document_ids):
        calls.append(list(document_ids))
        return documents

    monkeypatch.setattr(review_api, "get_documents_by_ids", fake_get_documents_by_ids)

    result = asyncio.run(review_api.list_reviews(db=None, skip=0, limit=20))
    items = result["items"]

    assert calls == [[11, 12]]
    assert result["total"] == 2
    assert items[0]["document_name"] == "a.docx"
    assert items[0]["progress"]["progress"] == 35
    assert items[1]["document_file_type"] == "pdf"


def test_normalize_review_status_accepts_supported_values():
    assert review_api._normalize_review_status(None) is None
    assert review_api._normalize_review_status("all") is None
    assert review_api._normalize_review_status("RUNNING") == "running"


def test_list_reviews_supports_filters(monkeypatch):
    reviews = [SimpleNamespace(id=8, document_id=21, status="running", summary="", total_issues=0)]
    documents = [SimpleNamespace(id=21, filename="demo.docx", file_type="docx")]
    captured = {}

    def fake_query_review_rows(db, document_id=None, status=None, latest_only=False, limit=100, skip=0):
        captured.update({
            "document_id": document_id,
            "status": status,
            "latest_only": latest_only,
            "limit": limit,
            "skip": skip,
        })
        return reviews

    monkeypatch.setattr(review_api, "_query_review_rows", fake_query_review_rows)
    monkeypatch.setattr(review_api, "_count_review_rows", lambda *args, **kwargs: 1)
    monkeypatch.setattr(review_api, "get_documents_by_ids", lambda db, document_ids: documents)
    monkeypatch.setattr(review_api, "_reconcile_review_runtime_state", lambda db, review: review)
    monkeypatch.setattr(review_api, "get_progress", lambda review_id: {"status": "running", "progress": 42, "message": "处理中"})

    result = asyncio.run(
        review_api.list_reviews(document_id=21, status="running", latest_only=True, limit=25, db=None)
    )

    assert captured == {"document_id": 21, "status": "running", "latest_only": True, "limit": 25, "skip": 0}
    assert result["items"][0]["document_name"] == "demo.docx"
    assert result["items"][0]["progress"]["progress"] == 42


def test_list_reviews_filters_reconciled_runtime_status(monkeypatch):
    stale_running_review = SimpleNamespace(id=9, document_id=22, status="running", summary="", total_issues=0)
    document_calls = []

    monkeypatch.setattr(review_api, "_query_review_rows", lambda *args, **kwargs: [stale_running_review])
    monkeypatch.setattr(review_api, "_count_review_rows", lambda *args, **kwargs: 1)
    monkeypatch.setattr(
        review_api,
        "_reconcile_review_runtime_state",
        lambda db, review: SimpleNamespace(**{**review.__dict__, "status": "failed"}),
    )
    monkeypatch.setattr(review_api, "get_documents_by_ids", lambda db, document_ids: document_calls.append(document_ids) or [])

    result = asyncio.run(review_api.list_reviews(status="running", db=None))

    assert result["items"] == []
    assert document_calls == [[]]


def test_export_review_excel_appends_red_opinion_columns(tmp_path, monkeypatch):
    from openpyxl import Workbook, load_workbook

    upload_dir = tmp_path / "uploads"
    export_dir = tmp_path / "exports"
    upload_dir.mkdir()
    export_dir.mkdir()

    workbook_path = upload_dir / "demo.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["中文", "英文"])
    ws.append(["保存", "Save"])
    wb.save(workbook_path)

    monkeypatch.setattr(review_api, "UPLOAD_DIR", upload_dir)
    monkeypatch.setattr(review_api, "REVIEW_EXPORT_DIR", export_dir)
    monkeypatch.setattr(review_api, "_select_export_issues", lambda payload: payload)

    document = SimpleNamespace(filename="demo.xlsx")
    review = SimpleNamespace(id=18)
    issues = [{
        "severity": "general",
        "category": "格式问题",
        "rule": "XLS-FMT-001",
        "description": "英文译文存在连续空格。",
        "suggestion": "建议改为 Save",
        "status": "pending",
        "source": "excel",
        "confidence": 95,
        "position": review_api._encode_issue_position_with_meta(0, 0, sheet="Sheet1", row=2, source_column=1, target_column=2),
    }]

    export_path, _, _ = review_api._export_review_excel(review, document, issues)

    exported = load_workbook(export_path)
    ws = exported["Sheet1"]
    assert ws.cell(row=1, column=3).value == "审核意见"
    assert ws.cell(row=2, column=3).value == "[格式问题] 英文译文存在连续空格。 → Save"
    assert ws.cell(row=2, column=3).font.color.type == "rgb"
    assert ws.cell(row=2, column=3).font.color.rgb.endswith("FF0000")
    assert ws.cell(row=2, column=4).value == "待修改"
    assert ws.cell(row=2, column=5).value == 1


def test_generate_excel_review_html_content_groups_rows(monkeypatch):
    monkeypatch.setattr(review_api, "_format_report_datetime", lambda value=None: "2026-08-18 12:00:00 (UTC+8)")
    monkeypatch.setattr(review_api, "_select_export_issues", lambda payload: payload)
    review = SimpleNamespace(id=16, document_id=8, created_at=None, completed_at=None)
    document = SimpleNamespace(filename="demo.xlsx", file_type="xlsx")
    issues = [
        {
            "severity": "serious",
            "category": "完整性",
            "rule": "XLS-COMP-002",
            "description": "英文译文为空，需补充。",
            "suggestion": "补充英文译文",
            "context": "中文: 保存 | 英文: ",
            "status": "pending",
            "source": "excel",
            "confidence": 95,
            "position": review_api._encode_issue_position_with_meta(0, 0, sheet="Sheet1", row=2, source_column=1, target_column=2),
        }
    ]

    html = review_api._generate_excel_review_html_content(review, document, issues)

    assert "Excel 审核报告" in html
    assert "Sheet1 / 第 2 行" in html
    assert "英文译文为空，需补充。 → 补充英文译文" in html
    assert "命中行数" in html


def test_run_excel_review_audit_covers_manual_excel_findings(tmp_path, monkeypatch):
    from openpyxl import Workbook

    upload_dir = tmp_path / "uploads"
    upload_dir.mkdir()
    workbook_path = upload_dir / "manual-gap.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["中文", "英文"])
    ws.append(["系统设置", "System Settings"])
    ws.append(["长按3s启动", "Press and hold for 3s to start"])
    ws.append(["锁屏时间", "Lock Screen Timeout Period"])
    ws.append(["时间日期   ", "Date&Time"])
    ws.append(["目标速度-", "Target Speed"])
    ws.append(["目标角度-", "Traget Angle"])
    ws.append(["错误信息", "Error message"])
    ws.append(["确认恢复出厂设置", "Confirm factory reset?"])
    ws.append(["确认固件升级", "Confirm to start firmware upgrade?"])
    ws.append(["确认结束老化测试", "Confirm to end aging test?"])
    ws.append(["确认清除老化次数", "Confirm to clear aging count?"])
    ws.append(["语言设置", "Language"])
    wb.save(workbook_path)

    monkeypatch.setattr(review_api, "UPLOAD_DIR", upload_dir)
    monkeypatch.setattr(review_api, "get_terms", lambda db, limit=10000: [])

    issues = review_api._run_excel_review_audit(None, SimpleNamespace(filename="manual-gap.xlsx"))

    original_texts = {issue["original_text"] for issue in issues}
    rules = {issue["rule"] for issue in issues}
    pair_002_issue = next(issue for issue in issues if issue["rule"] == "XLS-PAIR-002")

    assert "System Settings" not in original_texts
    assert "Press and hold for 3s to start" not in original_texts
    assert "Traget Angle" in original_texts
    assert "Confirm to start firmware upgrade?" in original_texts
    assert "Confirm to end aging test?" in original_texts
    assert "Confirm to clear aging count?" in original_texts
    assert pair_002_issue["suggestion"] == "建议改为 Screen Timeout"
    assert "XLS-CN-FMT-001" in rules
    assert "XLS-CN-FMT-002" in rules
    assert "XLS-PUNCT-001" in rules
    assert "XLS-PAIR-001" in rules
    assert "XLS-EN-STYLE-002" in rules
    assert "XLS-LANG-003" in rules


def test_select_relevant_ai_review_basis_prefers_matching_sections():
    sections = [
        {"label": "通用风格", "text": "【通用风格】\n统一标点和语法。", "priority": 2},
        {"label": "版本记录", "text": "【版本记录】\nrevision history copyright year version history", "priority": 5},
        {"label": "术语", "text": "【术语】\n术语一致性与缩略语。", "priority": 3},
    ]

    selected = review_api._select_relevant_ai_review_basis(
        "Please check revision history and copyright year consistency.",
        sections,
        max_sections=2,
        char_budget=500,
    )

    assert "版本记录" in selected
    assert "revision history" in selected


def test_extract_basis_tokens_filters_stopwords_and_noise():
    tokens = review_api._extract_basis_tokens(
        "请检查这个页面的相关问题，并确认 revision history 与 copyright year 是否一致。",
        max_tokens=20,
    )

    assert "revision" in tokens
    assert "history" in tokens
    assert "copyright" in tokens
    assert "请" not in tokens
    assert "相关" not in tokens
    assert "问题" not in tokens


def test_select_relevant_ai_review_basis_keeps_summary_first_and_only_when_irrelevant():
    sections = [
        {"label": "CYY人工审核经验基线摘要", "text": "【CYY人工审核经验基线摘要】\n关注高频问题。", "priority": 5},
        {"label": "版本记录", "text": "【版本记录】\nrevision history copyright year version history", "priority": 5},
        {"label": "术语", "text": "【术语】\n术语一致性与缩略语。", "priority": 3},
    ]

    selected = review_api._select_relevant_ai_review_basis(
        "完全不相关的机械噪声字段 abc xyz",
        sections,
        max_sections=3,
        char_budget=500,
    )

    assert selected.startswith("## CYY人工审核经验基线摘要")
    assert "## 版本记录" not in selected
    assert "## 术语" not in selected


def test_select_relevant_ai_review_basis_skips_summary_when_specific_sections_match():
    sections = [
        {"label": "CYY人工审核经验基线摘要", "text": "【CYY人工审核经验基线摘要】\n关注高频问题。", "priority": 5},
        {"label": "说明书发布前自检 Checklist", "text": "【说明书发布前自检 Checklist】\nrevision history date version", "priority": 5},
        {"label": "版本记录", "text": "【版本记录】\nrevision history copyright year version history", "priority": 4},
    ]

    selected = review_api._select_relevant_ai_review_basis(
        "Please check revision history and copyright year consistency.",
        sections,
        max_sections=2,
        char_budget=500,
    )

    assert "## 版本记录" in selected
    assert "## 说明书发布前自检 Checklist" in selected
    assert "## CYY人工审核经验基线摘要" not in selected


def test_build_ai_review_basis_sections_include_english_and_common_error_specs():
    sections = review_api._build_ai_review_basis_sections(
        {
            "en_style": "Use present tense. Use active voice. Use sentence style headings.",
            "common_errors": "中文文档统一使用全角标点。单位统一使用 μL、mL、kg。",
        },
        "both",
    )

    labels = {section["label"] for section in sections}

    assert "英文技术文档写作风格指南" in labels
    assert "技术文档常见错误清单" in labels


def test_build_ai_review_basis_sections_include_structured_cyy_sections(monkeypatch):
    monkeypatch.setattr(
        review_api,
        "_load_cyy_human_review_basis_sections",
        lambda: [
            {"label": "CYY人工审核经验基线摘要", "text": "【CYY人工审核经验基线摘要】\n总览", "priority": 5},
            {"label": "CYY人工审核经验基线-步骤结构", "text": "【CYY人工审核经验基线 - 步骤结构】\n关注步骤缺失", "priority": 5},
        ],
    )

    sections = review_api._build_ai_review_basis_sections({}, "cn")
    labels = [section["label"] for section in sections]

    assert "CYY人工审核经验基线摘要" in labels
    assert "CYY人工审核经验基线-步骤结构" in labels


def test_load_cyy_human_review_basis_sections_omits_comment_details(monkeypatch):
    review_api.CYY_HUMAN_REVIEW_BASIS_SECTIONS_CACHE = None
    monkeypatch.setattr(
        review_api,
        "CYY_HUMAN_REVIEW_BASELINE_PATH",
        SimpleNamespace(
            exists=lambda: True,
            read_text=lambda encoding="utf-8": json.dumps(
                {
                    "summary": {"total": 2, "by_category": {"步骤结构": 2}},
                    "annotations": [{"category": "步骤结构", "comment": "步骤3之后直接跳到步骤5"}],
                },
                ensure_ascii=False,
            ),
        ),
    )

    sections = review_api._load_cyy_human_review_basis_sections()

    assert any(section["label"] == "CYY人工审核经验基线-步骤结构" for section in sections)
    assert all("步骤3之后直接跳到步骤5" not in section["text"] for section in sections)


def test_select_relevant_ai_review_basis_prefers_matching_cyy_category_sections():
    sections = [
        {"label": "CYY人工审核经验基线摘要", "text": "【CYY人工审核经验基线摘要】\n关注高频问题。", "priority": 5},
        {"label": "CYY人工审核经验基线-步骤结构", "text": "【CYY人工审核经验基线 - 步骤结构】\n步骤编号跳号、步骤缺失、操作不可执行。", "priority": 5},
        {"label": "CYY人工审核经验基线-版本记录", "text": "【CYY人工审核经验基线 - 版本记录】\nrevision history 版本记录 日期。", "priority": 4},
    ]

    selected = review_api._select_relevant_ai_review_basis(
        "步骤3之后直接跳到步骤5，导致操作步骤缺失。",
        sections,
        max_sections=2,
        char_budget=500,
    )

    assert "步骤结构" in selected
    assert "操作不可执行" in selected


def test_select_relevant_ai_review_basis_limits_cyy_sections():
    sections = [
        {"label": "CYY人工审核经验基线摘要", "text": "【CYY人工审核经验基线摘要】\n总览。", "priority": 5},
        {"label": "CYY人工审核经验基线-步骤结构", "text": "【CYY人工审核经验基线 - 步骤结构】\n步骤编号跳号、步骤缺失。", "priority": 5},
        {"label": "CYY人工审核经验基线-术语一致性", "text": "【CYY人工审核经验基线 - 术语一致性】\n术语混用、字段命名不一致。", "priority": 5},
        {"label": "版本记录", "text": "【版本记录】\nrevision history date version", "priority": 4},
    ]

    selected = review_api._select_relevant_ai_review_basis(
        "步骤5之后跳到步骤7，术语名称也前后不一致。",
        sections,
        max_sections=4,
        char_budget=800,
    )

    assert selected.count("## CYY人工审核经验基线-") <= 2


def test_build_ai_review_basis_respects_char_budget(monkeypatch):
    monkeypatch.setenv("REVIEW_AI_BASIS_CHAR_BUDGET", "80")

    basis = review_api._build_ai_review_basis(
        {
            "en_style": "A" * 120,
            "common_errors": "B" * 120,
        },
        "both",
    )

    assert len(basis) <= 80


def test_review_recall_floor_env_defaults_disabled(monkeypatch):
    monkeypatch.delenv("REVIEW_RECALL_FLOOR", raising=False)
    assert review_api._review_env_bool("REVIEW_RECALL_FLOOR", False) is False

    monkeypatch.setenv("REVIEW_RECALL_FLOOR", "1")
    assert review_api._review_env_bool("REVIEW_RECALL_FLOOR", False) is True


def test_finalize_review_issues_defaults_to_pipeline_only(monkeypatch):
    monkeypatch.delenv("REVIEW_USE_LEGACY_POST_FILTERS", raising=False)
    monkeypatch.delenv("REVIEW_RECALL_FLOOR", raising=False)

    calls = []

    monkeypatch.setattr(review_api, "dedupe_issues_by_original", lambda issues: list(issues))
    monkeypatch.setattr(review_api, "_sanitize_issue_suggestions", lambda issues: list(issues))
    monkeypatch.setattr(
        review_api,
        "_filter_ai_issues_without_document_evidence_with_reasons",
        lambda issues, content: (list(issues), {"missing_evidence": 1}),
    )
    monkeypatch.setattr(review_api, "_apply_false_positive_signature_penalty", lambda issues, signatures: list(issues))

    def fake_false_positive(issues):
        calls.append("false_positive")
        return issues

    def fake_low_value(issues):
        calls.append("low_value")
        return issues

    def fake_pipeline(issues, *args, **kwargs):
        calls.append("pipeline")
        return list(issues)

    monkeypatch.setattr(review_api, "_filter_review_false_positives", fake_false_positive)
    monkeypatch.setattr(review_api, "_filter_low_value_review_issues", fake_low_value)
    monkeypatch.setattr(review_api, "pipeline_select_review_issues", fake_pipeline)

    issues, diagnostics = review_api._finalize_review_issues(
        [{"source": "ai", "rule": "AI", "original_text": "demo", "suggestion": "fix", "severity": "general"}],
        "demo content",
        set(),
    )

    assert len(issues) == 1
    assert calls == ["pipeline"]
    assert diagnostics["legacy_post_filters_enabled"] is False
    assert diagnostics["recall_floor_enabled"] is False
    assert diagnostics["document_evidence_drop_reasons"] == {"missing_evidence": 1}


def test_sample_ai_evidence_filter_drops_includes_compact_reason():
    samples = review_api._sample_ai_evidence_filter_drops(
        [
            {
                "source": "ai",
                "rule": "AI-GRAMMAR",
                "category": "语法表达",
                "severity": "general",
                "confidence": 82,
                "original_text": "missing phrase",
                "suggestion": "replacement phrase",
                "description": "AI suggested a correction",
            },
            {
                "source": "rule",
                "rule": "R001",
                "original_text": "ignored",
            },
        ],
        "document content without that phrase",
    )

    assert samples == [
        {
            "reason": "original_text_not_found",
            "rule": "AI-GRAMMAR",
            "category": "语法表达",
            "severity": "general",
            "confidence": 82,
            "original_text": "missing phrase",
            "suggestion": "replacement phrase",
            "description": "AI suggested a correction",
        }
    ]


def test_finalize_review_issues_records_ai_evidence_drop_samples(monkeypatch):
    monkeypatch.delenv("REVIEW_USE_LEGACY_POST_FILTERS", raising=False)
    monkeypatch.delenv("REVIEW_RECALL_FLOOR", raising=False)

    monkeypatch.setattr(review_api, "dedupe_issues_by_original", lambda issues: list(issues))
    monkeypatch.setattr(review_api, "_sanitize_issue_suggestions", lambda issues: list(issues))
    monkeypatch.setattr(review_api, "_apply_false_positive_signature_penalty", lambda issues, signatures: list(issues))
    monkeypatch.setattr(review_api, "pipeline_select_review_issues", lambda issues: list(issues))

    issues, diagnostics = review_api._finalize_review_issues(
        [
            {
                "source": "ai",
                "rule": "AI-GRAMMAR",
                "category": "语法表达",
                "severity": "general",
                "confidence": 75,
                "original_text": "not in document",
                "suggestion": "in document",
                "description": "AI correction",
            }
        ],
        "document body",
        set(),
    )

    assert issues == []
    assert diagnostics["document_evidence_drop_reasons"] == {"original_text_not_found": 1}
    assert diagnostics["document_evidence_drop_samples"][0]["reason"] == "original_text_not_found"
    assert diagnostics["document_evidence_drop_samples"][0]["original_text"] == "not in document"


def test_false_positive_signature_filter_drops_matched_issues():
    issues = [
        {"rule": "SPELL", "category": "拼写", "original_text": "consumbles", "source": "spellcheck"},
        {"rule": "SPELL", "category": "拼写", "original_text": "digestive", "source": "spellcheck"},
    ]

    filtered = review_api._apply_false_positive_signature_penalty(
        issues,
        {"SPELL|拼写|consumbles"},
    )

    assert len(filtered) == 2
    marked = next(item for item in filtered if item["original_text"] == "consumbles")
    kept = next(item for item in filtered if item["original_text"] == "digestive")
    assert marked.get("possible_false_positive") is True
    assert kept.get("possible_false_positive") is not True


def test_sync_false_positive_memory_adds_and_removes_entries():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    from app.database import Base
    from app.models.issue import Issue as IssueModel

    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)

    db = SessionLocal()
    try:
        issue = IssueModel(
            review_id=1,
            severity="general",
            category="拼写",
            rule="SPELL",
            chapter="Intro",
            original_text="consumbles",
            context="consumbles in intro",
            suggestion="consumables",
            description="拼写错误",
            audit_basis="basis",
            confidence=90,
            source="spellcheck",
            status="false_positive",
            position="{}",
        )
        db.add(issue)
        db.commit()
        db.refresh(issue)

        review_api._sync_false_positive_memory(db, issue)
        signatures = review_api.list_false_positive_memory_signatures(db)
        assert "spell|拼写|consumbles" in signatures
        assert "spell|consumbles" in signatures
        assert "拼写|consumbles" in signatures
        assert "consumbles" not in signatures

        issue.status = "confirmed"
        db.commit()
        db.refresh(issue)

        review_api._sync_false_positive_memory(db, issue)
        signatures = review_api.list_false_positive_memory_signatures(db)
        assert "consumbles" not in signatures
    finally:
        db.close()


def test_seed_preset_false_positive_memory_is_idempotent():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    from app.database import Base

    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)

    db = SessionLocal()
    try:
        assert crud_review.seed_preset_false_positive_memory(db) == 12
        assert crud_review.seed_preset_false_positive_memory(db) == 0

        items, total = crud_review.list_false_positive_memory(db, limit=20)

        assert total == 12
        assert len(items) == 12
        assert {item["document_name"] for item in items} == {"预置"}
        assert all(item["source_issue_id"] == 0 for item in items)
    finally:
        db.close()


def test_false_positive_penalty_keeps_same_text_on_different_rule():
    issues = [
        {"rule": "SPELL", "category": "拼写", "original_text": "consumbles", "source": "spellcheck"},
        {"rule": "R013", "category": "术语", "original_text": "consumbles", "source": "ai"},
    ]

    filtered = review_api._apply_false_positive_signature_penalty(
        issues,
        {"SPELL|拼写|consumbles"},
    )

    assert len(filtered) == 2
    spell_issue = next(item for item in filtered if item["rule"] == "SPELL")
    term_issue = next(item for item in filtered if item["rule"] == "R013")
    assert spell_issue.get("possible_false_positive") is True
    assert term_issue.get("possible_false_positive") is not True


def test_apply_pdf_visual_verification_filters_rejected_ai_issue(monkeypatch):
    document = SimpleNamespace(file_type="pdf", filename="demo.pdf")
    issues = [
        {
            "source": "ai",
            "severity": "general",
            "category": "术语",
            "original_text": "注册手册号",
            "context": "输入注册手册号后继续。",
            "description": "疑似术语错误",
            "suggestion": "改为注册手机号",
            "position": json.dumps({"page_number": 2}, ensure_ascii=False),
        },
        {
            "source": "rule",
            "severity": "general",
            "category": "结构",
            "original_text": "图4",
            "context": "图4",
            "description": "图号跳号",
            "suggestion": "补齐图3",
            "position": json.dumps({"page_number": 3}, ensure_ascii=False),
        },
    ]

    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_enabled", lambda: True)
    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_limit", lambda: 6)
    monkeypatch.setattr(review_api, "_get_document_upload_path", lambda document: review_api.PROJECT_ROOT / "backend")
    monkeypatch.setattr(review_api, "_render_pdf_page_png_bytes", lambda file_path, page_number, scale=None: b"png-bytes")
    monkeypatch.setattr(review_api.ai_client, "kimi_client", object())
    monkeypatch.setattr(
        review_api.ai_client,
        "verify_review_issue_from_image",
        lambda image_bytes, issue_payload, page_number=None, review_id=None, request_label="review.visual_verify": {
            "decision": "reject",
            "reason": "重复文本层伪影",
            "confidence": 92,
            "is_extraction_artifact": True,
        },
    )
    monkeypatch.setattr(review_api, "set_progress", lambda *args, **kwargs: None)

    filtered, diagnostics = review_api._apply_pdf_visual_verification(123, document, "dummy", issues)

    assert len(filtered) == 1
    assert filtered[0]["source"] == "rule"
    assert diagnostics["enabled"] is True
    assert diagnostics["candidate_count"] == 1
    assert diagnostics["rejected_count"] == 1
    issue_meta = json.loads(issues[0]["position"])
    assert issue_meta["visual_verification"]["decision"] == "reject"
    assert issue_meta["visual_verification"]["is_extraction_artifact"] is True


def test_apply_pdf_visual_verification_adds_confirmed_page_candidate_issue(monkeypatch):
    document = SimpleNamespace(file_type="pdf", filename="demo.pdf")
    issues = []
    suspicious_pages = {
        "enabled": True,
        "candidate_count": 1,
        "candidates": [{
            "page_number": 3,
            "reasons": ["table_layout", "dense_table_or_layout"],
            "char_start": 20,
            "char_end": 80,
            "text_preview": "版本 日期 2.0 2026年7月15日 1.0 2025年11月17日",
        }],
    }

    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_enabled", lambda: True)
    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_limit", lambda: 6)
    monkeypatch.setattr(review_api, "_get_document_upload_path", lambda document: review_api.PROJECT_ROOT / "backend")
    monkeypatch.setattr(review_api, "_render_pdf_page_png_bytes", lambda file_path, page_number, scale=None: b"png-bytes")
    monkeypatch.setattr(review_api.ai_client, "kimi_client", object())
    monkeypatch.setattr(
        review_api.ai_client,
        "verify_review_issue_from_image",
        lambda image_bytes, issue_payload, page_number=None, review_id=None, request_label="review.visual_verify": {
            "decision": "confirm",
            "reason": "表格列宽分布不均，页面显示拥挤",
            "confidence": 88,
            "is_extraction_artifact": False,
        },
    )
    monkeypatch.setattr(review_api, "set_progress", lambda *args, **kwargs: None)

    filtered, diagnostics = review_api._apply_pdf_visual_verification(123, document, "dummy", issues, suspicious_pages)

    assert len(filtered) == 1
    assert filtered[0]["rule"] == "PDF-VISUAL-001"
    assert filtered[0]["category"] == "表格/版式"
    assert diagnostics["page_candidate_count"] == 1
    assert diagnostics["page_issue_count"] == 1
    issue_meta = json.loads(filtered[0]["position"])
    assert issue_meta["page_number"] == 3
    assert issue_meta["visual_verification"]["page_candidate"] is True


def test_apply_pdf_visual_verification_adds_manual_page_issue_on_visual_error(monkeypatch):
    document = SimpleNamespace(file_type="pdf", filename="demo.pdf")
    issues = []
    suspicious_pages = {
        "enabled": True,
        "candidate_count": 1,
        "candidates": [{
            "page_number": 3,
            "reasons": ["table_layout"],
            "char_start": 20,
            "char_end": 80,
            "text_preview": "版本记录 日期 版本 修订 2026年7月15日 2.0 编制 2025年11月17日 1.0",
        }],
    }

    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_enabled", lambda: True)
    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_limit", lambda: 6)
    monkeypatch.setattr(review_api, "_get_document_upload_path", lambda document: review_api.PROJECT_ROOT / "backend")
    monkeypatch.setattr(review_api, "_render_pdf_page_png_bytes", lambda file_path, page_number, scale=None: b"png-bytes")
    monkeypatch.setattr(review_api.ai_client, "kimi_client", object())
    monkeypatch.setattr(
        review_api.ai_client,
        "verify_review_issue_from_image",
        lambda image_bytes, issue_payload, page_number=None, review_id=None, request_label="review.visual_verify": {
            "decision": "error",
            "reason": "visual model unavailable",
            "confidence": 0,
            "is_extraction_artifact": False,
        },
    )
    monkeypatch.setattr(review_api, "set_progress", lambda *args, **kwargs: None)

    filtered, diagnostics = review_api._apply_pdf_visual_verification(123, document, "dummy", issues, suspicious_pages)

    assert len(filtered) == 1
    assert filtered[0]["rule"] == "PDF-VISUAL-001"
    assert filtered[0]["category"] == "表格/版式"
    assert filtered[0]["source"] == "rule"
    assert "平均分布列" in filtered[0]["description"]
    assert "版本 2.0 1.0" in filtered[0]["context"]
    assert "版本列值: 2.0 1.0" in filtered[0]["context"]
    assert diagnostics["page_candidate_count"] == 1
    assert diagnostics["page_issue_count"] == 1
    issue_meta = json.loads(filtered[0]["position"])
    assert issue_meta["page_number"] == 3
    assert issue_meta["visual_verification"]["decision"] == "needs_manual_review"


def test_apply_pdf_visual_verification_filters_known_quote_mapping_artifact_after_uncertain(monkeypatch):
    document = SimpleNamespace(file_type="pdf", filename="demo.pdf")
    issues = [
        {
            "source": "ai",
            "severity": "general",
            "category": "格式规范",
            "original_text": '“Run settings".',
            "context": 'The text layer shows “Run settings". in the UI message.',
            "description": '引号和句号位置疑似异常',
            "suggestion": '建议统一引号和句号位置',
            "position": json.dumps({"page_number": 2}, ensure_ascii=False),
        },
    ]

    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_enabled", lambda: True)
    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_limit", lambda: 6)
    monkeypatch.setattr(review_api, "_get_document_upload_path", lambda document: review_api.PROJECT_ROOT / "backend")
    monkeypatch.setattr(review_api, "_render_pdf_page_png_bytes", lambda file_path, page_number, scale=None: b"png-bytes")
    monkeypatch.setattr(review_api.ai_client, "kimi_client", object())
    monkeypatch.setattr(
        review_api.ai_client,
        "verify_review_issue_from_image",
        lambda image_bytes, issue_payload, page_number=None, review_id=None, request_label="review.visual_verify": {
            "decision": "uncertain",
            "reason": "截图证据不足",
            "confidence": 61,
            "is_extraction_artifact": False,
        },
    )
    monkeypatch.setattr(review_api, "set_progress", lambda *args, **kwargs: None)

    filtered, diagnostics = review_api._apply_pdf_visual_verification(123, document, "dummy", issues)

    assert filtered == []
    assert diagnostics["rejected_count"] == 1
    issue_meta = json.loads(issues[0]["position"])
    assert issue_meta["visual_verification"]["decision"] == "reject"
    assert issue_meta["visual_verification"]["reason"] == "quote_mapping_artifact"


def test_apply_pdf_visual_verification_filters_known_following_status_artifact_after_uncertain(monkeypatch):
    document = SimpleNamespace(file_type="pdf", filename="demo.pdf")
    issues = [
        {
            "source": "ai",
            "severity": "general",
            "category": "语言质量",
            "original_text": 'following status',
            "context": 'Check the following status of the module.',
            "description": '短语表达疑似异常',
            "suggestion": '建议改写为 following statuses',
            "position": json.dumps({"page_number": 4}, ensure_ascii=False),
        },
    ]

    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_enabled", lambda: True)
    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_limit", lambda: 6)
    monkeypatch.setattr(review_api, "_get_document_upload_path", lambda document: review_api.PROJECT_ROOT / "backend")
    monkeypatch.setattr(review_api, "_render_pdf_page_png_bytes", lambda file_path, page_number, scale=None: b"png-bytes")
    monkeypatch.setattr(review_api.ai_client, "kimi_client", object())
    monkeypatch.setattr(
        review_api.ai_client,
        "verify_review_issue_from_image",
        lambda image_bytes, issue_payload, page_number=None, review_id=None, request_label="review.visual_verify": {
            "decision": "uncertain",
            "reason": "截图证据不足",
            "confidence": 58,
            "is_extraction_artifact": False,
        },
    )
    monkeypatch.setattr(review_api, "set_progress", lambda *args, **kwargs: None)

    filtered, diagnostics = review_api._apply_pdf_visual_verification(123, document, "dummy", issues)

    assert filtered == []
    assert diagnostics["rejected_count"] == 1
    issue_meta = json.loads(issues[0]["position"])
    assert issue_meta["visual_verification"]["decision"] == "reject"
    assert issue_meta["visual_verification"]["reason"] == "accepted_pdf_phrase_following_status"


def test_apply_pdf_visual_verification_filters_known_turn_on_it_artifact_after_uncertain(monkeypatch):
    document = SimpleNamespace(file_type="pdf", filename="demo.pdf")
    issues = [
        {
            "source": "ai",
            "severity": "general",
            "category": "语言质量",
            "original_text": 'turn on it',
            "context": 'Press the switch and turn on it before use.',
            "description": '短语表达疑似异常',
            "suggestion": '建议改写为 turn it on',
            "position": json.dumps({"page_number": 5}, ensure_ascii=False),
        },
    ]

    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_enabled", lambda: True)
    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_limit", lambda: 6)
    monkeypatch.setattr(review_api, "_get_document_upload_path", lambda document: review_api.PROJECT_ROOT / "backend")
    monkeypatch.setattr(review_api, "_render_pdf_page_png_bytes", lambda file_path, page_number, scale=None: b"png-bytes")
    monkeypatch.setattr(review_api.ai_client, "kimi_client", object())
    monkeypatch.setattr(
        review_api.ai_client,
        "verify_review_issue_from_image",
        lambda image_bytes, issue_payload, page_number=None, review_id=None, request_label="review.visual_verify": {
            "decision": "uncertain",
            "reason": "截图证据不足",
            "confidence": 57,
            "is_extraction_artifact": False,
        },
    )
    monkeypatch.setattr(review_api, "set_progress", lambda *args, **kwargs: None)

    filtered, diagnostics = review_api._apply_pdf_visual_verification(123, document, "dummy", issues)

    assert filtered == []
    assert diagnostics["rejected_count"] == 1
    issue_meta = json.loads(issues[0]["position"])
    assert issue_meta["visual_verification"]["decision"] == "reject"
    assert issue_meta["visual_verification"]["reason"] == "accepted_pdf_phrase_turn_on_it"


def test_apply_pdf_visual_verification_filters_missing_icon_text_layer_artifact_after_uncertain(monkeypatch):
    document = SimpleNamespace(file_type="pdf", filename="demo.pdf")
    issues = [
        {
            "source": "ai",
            "severity": "general",
            "category": "完整性",
            "original_text": 'icon',
            "context": 'The review thinks a toolbar icon is missing in the text layer.',
            "description": 'missing icon in screenshot area',
            "suggestion": '补充缺少的图标说明',
            "position": json.dumps({"page_number": 6}, ensure_ascii=False),
        },
    ]

    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_enabled", lambda: True)
    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_limit", lambda: 6)
    monkeypatch.setattr(review_api, "_get_document_upload_path", lambda document: review_api.PROJECT_ROOT / "backend")
    monkeypatch.setattr(review_api, "_render_pdf_page_png_bytes", lambda file_path, page_number, scale=None: b"png-bytes")
    monkeypatch.setattr(review_api.ai_client, "kimi_client", object())
    monkeypatch.setattr(
        review_api.ai_client,
        "verify_review_issue_from_image",
        lambda image_bytes, issue_payload, page_number=None, review_id=None, request_label="review.visual_verify": {
            "decision": "uncertain",
            "reason": "截图中控件可见，但文本层无法确认",
            "confidence": 55,
            "is_extraction_artifact": False,
        },
    )
    monkeypatch.setattr(review_api, "set_progress", lambda *args, **kwargs: None)

    filtered, diagnostics = review_api._apply_pdf_visual_verification(123, document, "dummy", issues)

    assert filtered == []
    assert diagnostics["rejected_count"] == 1
    issue_meta = json.loads(issues[0]["position"])
    assert issue_meta["visual_verification"]["reason"] == "missing_icon_text_layer_artifact"


def test_apply_pdf_visual_verification_filters_duplicated_text_layer_artifact_after_uncertain(monkeypatch):
    document = SimpleNamespace(file_type="pdf", filename="demo.pdf")
    issues = [
        {
            "source": "ai",
            "severity": "general",
            "category": "重复内容",
            "original_text": '一旦您开始使',
            "context": '一旦您开始使 一旦您开始使 用本软件',
            "description": '重复文本层伪影',
            "suggestion": '删除重复文本',
            "position": json.dumps({"page_number": 7}, ensure_ascii=False),
        },
    ]

    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_enabled", lambda: True)
    monkeypatch.setattr(review_api, "_review_pdf_visual_verify_limit", lambda: 6)
    monkeypatch.setattr(review_api, "_get_document_upload_path", lambda document: review_api.PROJECT_ROOT / "backend")
    monkeypatch.setattr(review_api, "_render_pdf_page_png_bytes", lambda file_path, page_number, scale=None: b"png-bytes")
    monkeypatch.setattr(review_api.ai_client, "kimi_client", object())
    monkeypatch.setattr(
        review_api.ai_client,
        "verify_review_issue_from_image",
        lambda image_bytes, issue_payload, page_number=None, review_id=None, request_label="review.visual_verify": {
            "decision": "uncertain",
            "reason": "截图证据不足",
            "confidence": 54,
            "is_extraction_artifact": False,
        },
    )
    monkeypatch.setattr(review_api, "set_progress", lambda *args, **kwargs: None)

    filtered, diagnostics = review_api._apply_pdf_visual_verification(123, document, "dummy", issues)

    assert filtered == []
    assert diagnostics["rejected_count"] == 1
    issue_meta = json.loads(issues[0]["position"])
    assert issue_meta["visual_verification"]["reason"] == "duplicated_text_layer_artifact"


def test_finalize_review_issues_can_enable_legacy_post_filters(monkeypatch):
    monkeypatch.setenv("REVIEW_USE_LEGACY_POST_FILTERS", "1")
    monkeypatch.delenv("REVIEW_RECALL_FLOOR", raising=False)

    calls = []

    monkeypatch.setattr(review_api, "dedupe_issues_by_original", lambda issues: list(issues))
    monkeypatch.setattr(review_api, "_sanitize_issue_suggestions", lambda issues: list(issues))
    monkeypatch.setattr(
        review_api,
        "_filter_ai_issues_without_document_evidence_with_reasons",
        lambda issues, content: (list(issues), {}),
    )
    monkeypatch.setattr(review_api, "_apply_false_positive_signature_penalty", lambda issues, signatures: list(issues))

    def fake_false_positive(issues):
        calls.append("false_positive")
        return list(issues)

    def fake_low_value(issues):
        calls.append("low_value")
        return list(issues)

    def fake_pipeline(issues, *args, **kwargs):
        calls.append("pipeline")
        return list(issues)

    monkeypatch.setattr(review_api, "_filter_review_false_positives", fake_false_positive)
    monkeypatch.setattr(review_api, "_filter_low_value_review_issues", fake_low_value)
    monkeypatch.setattr(review_api, "pipeline_select_review_issues", fake_pipeline)

    issues, diagnostics = review_api._finalize_review_issues(
        [{"source": "ai", "rule": "AI", "original_text": "demo", "suggestion": "fix", "severity": "general"}],
        "demo content",
        set(),
    )

    assert len(issues) == 1
    assert calls == ["false_positive", "low_value", "pipeline"]
    assert diagnostics["legacy_post_filters_enabled"] is True


def test_run_cached_ai_chunk_review_reuses_cached_result(monkeypatch):
    review_api._ai_review_chunk_cache.clear()
    calls = []

    monkeypatch.setattr(review_api, "_review_cache_version", lambda: "v-cache")
    monkeypatch.setattr(review_api, "_ai_provider_cache_fingerprint", lambda: "provider-a")

    def fake_call_with_timeout(func, timeout_seconds, *args, **kwargs):
        calls.append((timeout_seconds, args[0], args[2], args[4], args[5]))
        return {"issues": [{"rule": "AI", "original_text": "demo"}]}

    monkeypatch.setattr(review_api, "_call_with_timeout", fake_call_with_timeout)

    first, first_hit = review_api._run_cached_ai_chunk_review(88, "chunk-a", "en", "basis-a", 18)
    second, second_hit = review_api._run_cached_ai_chunk_review(88, "chunk-a", "en", "basis-a", 18)

    assert first_hit is False
    assert second_hit is True
    assert first == second
    assert len(calls) == 1
    assert calls[0] == (18, "chunk-a", "basis-a", 88, "review.audit_chunk")


def test_pipeline_filters_single_char_punctuation_rule_issue():
    issues = [
        {
            "source": "rule",
            "rule": "EXT-R013",
            "category": "格式错误",
            "severity": "general",
            "original_text": ".",
            "suggestion": "文档中标点符号使用必须符合规范",
            "description": "单字符标点噪声",
            "audit_basis": "标点规范",
        }
    ]

    selected = review_pipeline.select_review_issues(issues)

    assert selected == []


def test_pipeline_filters_quote_rule_without_quote_evidence():
    issues = [
        {
            "source": "rule",
            "rule": "EXT-R011",
            "category": "格式错误",
            "severity": "general",
            "original_text": "不能",
            "context": "A 不能。盘点状态下的物料将被系统锁定。",
            "suggestion": "文档中所有需要使用引号的场景统一使用双引号，不得混用单引号",
            "description": "文档中所有需要使用引号的场景统一使用双引号，不得混用单引号",
            "audit_basis": "引号规范",
        }
    ]

    selected = review_pipeline.select_review_issues(issues)

    assert selected == []


def test_pipeline_keeps_serious_real_spelling_issue():
    issues = [
        {
            "source": "spellcheck",
            "rule": "SPELL",
            "category": "拼写/用词错误",
            "severity": "serious",
            "original_text": "consumbles",
            "suggestion": "consumables",
            "description": "拼写错误",
            "audit_basis": "英文拼写规范",
            "confidence": 96,
        }
    ]

    selected = review_pipeline.select_review_issues(issues)

    assert len(selected) == 1
    assert selected[0]["original_text"] == "consumbles"


def test_pipeline_filters_low_value_generic_spelling_noise():
    issues = [
        {
            "source": "spellcheck",
            "rule": "SPELL",
            "category": "拼写/用词错误",
            "severity": "general",
            "original_text": "StandardMPS",
            "suggestion": "standard's",
            "description": "拼写错误",
            "audit_basis": "英文拼写规范",
            "confidence": 80,
        }
    ]

    selected = review_pipeline.select_review_issues(issues)

    assert selected == []


def test_pipeline_filters_low_precision_cyy_reference_check():
    issues = [
        {
            "source": "rule",
            "rule": "CYY-CN-REF-002",
            "category": "交叉引用",
            "severity": "general",
            "original_text": "Barcode 管理界面",
            "suggestion": "请检查该引用是否需要同步到全文一致的页码或章节",
            "description": "未更新交叉引用，通查。",
            "audit_basis": "CYY人工审核经验基线 - 交叉引用通查",
            "confidence": 90,
        }
    ]

    selected = review_pipeline.select_review_issues(issues)

    assert selected == []


def test_pipeline_filters_short_time_placeholder_noise():
    issues = [
        {
            "source": "rule",
            "rule": "CYY-CN-PLACEHOLDER-001",
            "category": "占位符残留",
            "severity": "general",
            "original_text": "XX:XX:XX",
            "suggestion": "请替换为正式内容或删除该占位符",
            "description": "疑似模板占位符残留。",
            "audit_basis": "CYY人工审核经验基线 - 占位符检查",
            "confidence": 90,
        }
    ]

    selected = review_pipeline.select_review_issues(issues)

    assert selected == []


def test_should_drop_unicode_equivalent_issue_accepts_mu_variants():
    issue = {
        "original_text": "20 µL",
        "suggestion": "20 μL",
    }

    assert review_api._should_drop_unicode_equivalent_issue(issue) is True


def test_run_logic_integrity_audit_skips_mixed_separator_false_positive():
    content = "4) Load sample\n12. Secondary heading\n"

    issues = review_api._run_logic_integrity_audit(content)

    assert not any(issue.get("rule") == "LOGIC-001" for issue in issues)


def test_dedupe_issues_by_original_prefers_spellcheck_over_ai():
    ai_issue = {
        "severity": "serious",
        "category": "语法",
        "rule": "AI-GRAMMAR",
        "chapter": "Section 1",
        "original_text": "Reviewing parameters",
        "source": "ai",
        "confidence": 95,
        "position": "10-30",
    }
    spell_issue = {
        "severity": "general",
        "category": "拼写/用词错误",
        "rule": "SPELL",
        "chapter": "Section 1",
        "original_text": "Reviewing parameters",
        "source": "spellcheck",
        "confidence": 80,
        "position": "10-30",
    }

    deduped = review_api.dedupe_issues_by_original([ai_issue, spell_issue])

    assert len(deduped) == 1
    assert deduped[0]["source"] == "spellcheck"


def test_dedupe_issues_by_original_prefers_spellcheck_when_chapter_differs_but_suggestion_matches():
    ai_issue = {
        "severity": "general",
        "category": "Terminology",
        "rule": "AI-TERM",
        "chapter": "Section 2.2 User-supplied equipment, reagent, and consumbles",
        "original_text": "consumbles",
        "suggestion": "consumables",
        "source": "ai",
        "confidence": 100,
        "position": "2898-2908",
    }
    spell_issue = {
        "severity": "general",
        "category": "拼写/用词错误",
        "rule": "SPELL",
        "chapter": "2.2 User-supplied equipment, reagent, and",
        "original_text": "consumbles",
        "suggestion": "consumables",
        "source": "spellcheck",
        "confidence": 95,
        "position": "2898-2908",
    }

    deduped = review_api.dedupe_issues_by_original([ai_issue, spell_issue])

    assert len(deduped) == 1
    assert deduped[0]["source"] == "spellcheck"


def test_dedupe_similar_but_not_identical_issues_kept():
    first_issue = {
        "severity": "general",
        "category": "操作步骤",
        "rule": "AI-STEP",
        "chapter": "Section 1",
        "original_text": "点击【新增】按钮，填写物料编码，点击【确定】保存。",
        "source": "ai",
        "confidence": 92,
        "position": "10-30",
    }
    second_issue = {
        "severity": "general",
        "category": "操作步骤",
        "rule": "AI-STEP",
        "chapter": "Section 1",
        "original_text": "点击【新增】按钮，填写物料名称，点击【确定】保存。",
        "source": "ai",
        "confidence": 91,
        "position": "31-51",
    }

    deduped = review_api.dedupe_issues_by_original([first_issue, second_issue])

    assert len(deduped) == 2


def test_pipeline_keeps_substantive_high_confidence_ai_grammar_issue():
    issue = {
        "severity": "general",
        "category": "Grammar",
        "rule": "AI-GRAMMAR",
        "chapter": "Section 1",
        "original_text": "The sample are ready for loading.",
        "suggestion": "The samples are ready for loading.",
        "description": "Subject-verb agreement issue.",
        "source": "ai",
        "confidence": 91,
        "position": "10-45",
    }

    selected = review_pipeline.select_review_issues([issue])

    assert len(selected) == 1
    assert selected[0]["source"] == "ai"
    assert selected[0]["review_value_score"] >= 45


def test_pipeline_still_filters_low_value_ai_template_issue():
    issue = {
        "severity": "general",
        "category": "Grammar",
        "rule": "AI-GRAMMAR",
        "chapter": "Section 1",
        "original_text": "Browse",
        "suggestion": "Click Browse.",
        "description": "Generic UI wording suggestion.",
        "source": "ai",
        "confidence": 96,
        "position": "10-16",
    }

    selected = review_pipeline.select_review_issues([issue])

    assert selected == []


def test_parse_ai_issue_with_fatal_severity():
    payload = json.dumps({
        "issues": [{
            "severity": "fatal",
            "type": "Compliance",
            "category": "法规合规",
            "location": "Warnings",
            "original": "WARNING label is missing.",
            "expected": "Add the WARNING label.",
            "context": "Before operation, the warning label is missing from the section.",
            "rule": "SAFE-001",
            "basis": "Safety checklist clause 2.1",
            "source": "ai",
            "status": "open",
            "confidence": 96,
        }],
        "summary": {
            "total": 1,
            "fatal": 1,
            "serious": 0,
            "general": 0,
            "suggestion": 0,
            "categories": {"法规合规": 1},
        },
    })

    data = review_api.ai_client._extract_json(payload, {"issues": []})
    issues = review_api.ai_client.normalize_audit_issues(data.get("issues", []), "WARNING label is missing.")

    assert len(issues) == 1
    assert issues[0]["severity"] == "fatal"
    assert issues[0]["category"] == "法规合规"
    assert issues[0]["context"]
    assert issues[0]["audit_basis"] == "Safety checklist clause 2.1"
    assert issues[0]["source"] == "ai"
    assert issues[0]["status"] == "open"


def test_validate_ai_issue_candidate_rejects_ruo_template_rewrite():
    issue = {
        "source": "ai",
        "original_text": "WARNING This product is intended only for scientific research and should not be used for clinical diagnosis.",
        "suggestion": "WARNING: This product is for research use only (RUO) and is not intended for clinical diagnostic use.",
        "description": "",
        "rule": "Compliance",
        "chapter": "Warnings",
        "audit_basis": "basis",
    }
    content = issue["original_text"]

    result = review_validation.validate_ai_issue_candidate(issue, content)

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_dnb_expansion_rewrite():
    issue = {
        "source": "ai",
        "original_text": "The device adopts the advanced DNB and the core technology of combinatorial probe-anchor synthesis (cPAS)",
        "suggestion": "The system uses DNA Nanoballs (DNBs) and the core technology of combinatorial probe-anchor synthesis (cPAS)",
        "description": "",
        "rule": "Terminology",
        "chapter": "Introduction",
        "audit_basis": "basis",
    }
    content = issue["original_text"]

    result = review_validation.validate_ai_issue_candidate(issue, content)

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_article_only_rewrite():
    issue = {
        "source": "ai",
        "original_text": "ensure that proper protections for personnel and device are implemented to avoid contamination by sequencing reagents.",
        "suggestion": "Ensure that proper protections for personnel and the device are implemented to avoid contamination by sequencing reagents.",
        "description": "",
        "rule": "Operation",
        "chapter": "CAUTION",
        "audit_basis": "basis",
    }
    content = issue["original_text"]

    result = review_validation.validate_ai_issue_candidate(issue, content)

    assert result.accepted is False
    assert result.reason == "low_value_english_rewrite"


def test_validate_ai_issue_candidate_rejects_dnb_solution_expansion_rewrite():
    issue = {
        "source": "ai",
        "original_text": "DNB 102",
        "suggestion": "DNB solution 102",
        "description": "",
        "rule": "Table",
        "chapter": "Table 3",
        "audit_basis": "basis",
    }
    content = "Table 3 includes DNB 102 in the component list."

    result = review_validation.validate_ai_issue_candidate(issue, content)

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_speculative_completion():
    issue = {
        "source": "ai",
        "original_text": "Add 34 µL of",
        "suggestion": "Add 34 µL of DNB loading mixture.",
        "description": "",
        "rule": "Operation",
        "chapter": "Step 5",
        "audit_basis": "basis",
    }
    content = "Then add 34 µL of to the tube for the next step."

    result = review_validation.validate_ai_issue_candidate(issue, content)

    assert result.accepted is False
    assert result.reason == "speculative_completion"


def test_validate_ai_issue_candidate_rejects_speculative_completion_with_bullet_prefix():
    issue = {
        "source": "ai",
        "original_text": "y If a reagent cartridge has been",
        "suggestion": "If a reagent cartridge has been thawed and the signal protein mixture has been added into the MSP well according to Preparing the reagent cartridge on Page 18, but it cannot be used in time, store it at room temperature and use it within 2 hours.",
        "description": "",
        "rule": "Completeness",
        "chapter": "Storage",
        "audit_basis": "basis",
    }
    content = "y If a reagent cartridge has been"

    result = review_validation.validate_ai_issue_candidate(issue, content)

    assert result.accepted is False
    assert result.reason == "speculative_completion"


def test_validate_ai_issue_candidate_rejects_unsupported_cn_source_expansion():
    issue = {
        "source": "ai",
        "original_text": "系统可同步系统内部已有的实验耗材、试剂等物料，同时支持手动新增普通物料。",
        "suggestion": "系统可同步外部系统或本平台内已有的实验耗材、试剂等物料，同时支持手动新增普通物料。",
        "description": "",
        "rule": "Terminology",
        "chapter": "Q&A章节（页码70）",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_aggressive_rewrite_blocks_template_rewrites():
    assert review_validation.ai_suggestion_is_aggressive_rewrite(
        "This reagent set is for research use only, and cannot be used for clinical diagnosis.",
        "This reagent set is for research use only (RUO) and is not intended for clinical diagnosis, patient management, or any diagnostic purposes.",
    ) is True
    assert review_validation.ai_suggestion_is_aggressive_rewrite(
        "This product is for research use only. Please read the instructions for use of the product carefully before use.",
        "This product is for research use only (RUO).",
    ) is True
    assert review_validation.ai_suggestion_is_aggressive_rewrite(
        "The reagent cartridge is completely thawed when there is no sound of cracked ice during shaking.",
        "The reagent cartridge is completely thawed when no sound of cracking ice is heard during gentle shaking.",
    ) is True


def test_aggressive_rewrite_passes_local_fixes():
    assert review_validation.ai_suggestion_is_aggressive_rewrite(
        "This instructions for use describes how to perform sequencing",
        "These instructions for use describe how to perform sequencing",
    ) is False
    assert review_validation.ai_suggestion_is_aggressive_rewrite(
        "Disgestive Buffer 250 uL per tube",
        "Digestive Buffer 250 uL per tube",
    ) is False
    assert review_validation.ai_suggestion_is_aggressive_rewrite(
        "click OK, The device begins unloading",
        "Tap OK, The device begins unloading",
    ) is False


def test_validate_ai_issue_candidate_accepts_local_grammar_fix():
    issue = {
        "source": "ai",
        "original_text": "This instructions for use describes how to perform sequencing",
        "suggestion": "These instructions for use describe how to perform sequencing",
        "description": "",
        "rule": "Grammar",
        "chapter": "Introduction",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is True
    assert result.reason == "accepted"


def test_validate_ai_issue_candidate_accepts_digestive_spelling_fix():
    issue = {
        "source": "ai",
        "original_text": "Disgestive Buffer 250 uL per tube",
        "suggestion": "Digestive Buffer 250 uL per tube",
        "description": "",
        "rule": "Spelling",
        "chapter": "Reagents",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is True
    assert result.reason == "accepted"


def test_validate_ai_issue_candidate_rejects_high_overlap_style_rewrite():
    issue = {
        "source": "ai",
        "original_text": "Data output (GB/flow cell)",
        "suggestion": "Data output (GB per flow cell)",
        "description": "",
        "rule": "Terminology",
        "chapter": "Specifications",
        "audit_basis": "basis",
    }
    content = "The specification table contains Data output (GB/flow cell)."

    result = review_validation.validate_ai_issue_candidate(issue, content)

    assert result.accepted is False
    assert result.reason == "low_value_english_rewrite"


def test_validate_ai_issue_candidate_rejects_low_value_cn_term_swap():
    issue = {
        "source": "ai",
        "original_text": "点击【导出物料】，弹出【选择文件保存位置】窗口。",
        "suggestion": "点击【导出物料】，弹出【选择文件保存位置】对话框。",
        "description": "",
        "rule": "Terminology",
        "chapter": "批量导出物料编码",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "low_value_cn_term_swap"


def test_validate_ai_issue_candidate_rejects_support_policy_rewrite():
    issue = {
        "source": "ai",
        "original_text": "For the username and password, contact the technical support.",
        "suggestion": "Default credentials are not provided. Contact technical support to obtain authorized login credentials.",
        "description": "",
        "rule": "Operation",
        "chapter": "Login",
        "audit_basis": "basis",
    }
    content = issue["original_text"]

    result = review_validation.validate_ai_issue_candidate(issue, content)

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_instructions_for_use_semantic_rewrite():
    issue = {
        "source": "ai",
        "original_text": "The instructions for use are provided in the package.",
        "suggestion": "The IFU is provided in the package.",
        "description": "",
        "rule": "Terminology",
        "chapter": "Introduction",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_mixedly_use_rewrite():
    issue = {
        "source": "ai",
        "original_text": "Do not mixedly use the MDA T-Regent from different models of reagent kits.",
        "suggestion": "Do not mix the MDA T-Reagent from different reagent kit models.",
        "description": "",
        "rule": "Terminology",
        "chapter": "Tips",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_app_library_pluralization_rewrite():
    issue = {
        "source": "ai",
        "original_text": "For App library",
        "suggestion": "For App libraries",
        "description": "",
        "rule": "Terminology",
        "chapter": "4.4 Quantifying DNBs",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_choose_scheme_semantic_rewrite():
    issue = {
        "source": "ai",
        "original_text": "y The recipe you selected should be consistent with the selected type in the Choose scheme interface.",
        "suggestion": "The recipe you selected must match the assay type selected in the Choose scheme interface.",
        "description": "",
        "rule": "Terminology",
        "chapter": "Tips",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_local_regulations_compliance_expansion():
    issue = {
        "source": "ai",
        "original_text": "y Dispose of the flow cell, reagent cartridge, waste, waste container, and PCR tube in accordance with local regulations and safety standards.",
        "suggestion": "Dispose of the flow cell, reagent cartridge, waste, waste container, and PCR tube in accordance with applicable local, national, and institutional biosafety and hazardous waste regulations.",
        "description": "",
        "rule": "Compliance",
        "chapter": "7.7 Removing the flow cell, reagent cartridge, and waste container",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_split_barcode_default_value_expansion():
    issue = {
        "source": "ai",
        "original_text": "If the selected recipe includes the barcode length, you need to select whether to split barcode.",
        "suggestion": "If the selected recipe includes a barcode, select whether to split the barcode. 'Yes' is selected by default.",
        "description": "",
        "rule": "Clarity",
        "chapter": "step 5",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_flow_cell_validation_logic_expansion():
    issue = {
        "source": "ai",
        "original_text": "Scan the QR code on the plastic package of the sequencing flow cell. Flow cell ID, Throughput, and Expiration date are automatically filled in.",
        "suggestion": "Scan the QR code on the plastic package of the sequencing flow cell. Flow cell ID, Throughput, and Expiration date are automatically filled in. If the flow cell ID is invalid or expired, the system displays an error and prevents proceeding.",
        "description": "",
        "rule": "Completeness",
        "chapter": "step 7.3, item 1",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_reagent_pluralization_inside_faulty_title():
    issue = {
        "source": "ai",
        "original_text": "reagent",
        "suggestion": "reagents",
        "description": "",
        "rule": "Terminology",
        "chapter": "2.2 User-supplied equipment, reagent, and consumbles",
        "audit_basis": "basis",
    }
    content = "2.2 User-supplied equipment, reagent, and consumbles"

    result = review_validation.validate_ai_issue_candidate(issue, content)

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_dnb_rationale_completion():
    issue = {
        "source": "ai",
        "original_text": "Do not centrifuge, vortex, or shake the tube.",
        "suggestion": "Do not centrifuge, vortex, or shake the tube — these actions may shear or aggregate DNBs.",
        "description": "",
        "rule": "Completeness",
        "chapter": "Tips under Step 3",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "protected_meaning_changed"


def test_validate_ai_issue_candidate_rejects_visual_icon_gap_from_pdf_text_loss():
    issue = {
        "source": "ai",
        "original_text": "双击计算机桌面软件图标 启动操作软件。",
        "suggestion": "请补充需要双击的具体软件图标名称，避免仅写图标。",
        "description": "操作步骤缺少具体图标对象，用户无法确定点击哪个图标。",
        "rule": "Completeness",
        "category": "操作步骤",
        "chapter": "开机与软件准备",
        "audit_basis": "basis",
        "context": "4. 先按计算机电源按钮开机，再按仪器电源开关，启动仪器。5. 双击计算机桌面软件图标 启动操作软件。6. 等待软件完成自检流程。",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["context"])

    assert result.accepted is False
    assert result.reason == "visual_control_ambiguity"


def test_validate_ai_issue_candidate_rejects_visual_button_gap_from_pdf_text_loss():
    issue = {
        "source": "ai",
        "original_text": "点击采集模式后方的 按钮，可设置模式参数。",
        "suggestion": "请补充该按钮或图标的具体名称，避免用户无法定位。",
        "description": "操作步骤缺少具体按钮对象。",
        "rule": "Completeness",
        "category": "操作步骤",
        "chapter": "采集模式",
        "audit_basis": "basis",
        "context": "实时显示眼底图像。点击采集模式后方的 按钮，可设置模式参数。显示相应眼别的采集任务。",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["context"])

    assert result.accepted is False
    assert result.reason == "visual_control_ambiguity"


def test_validate_ai_issue_candidate_rejects_visual_blank_click_from_pdf_text_loss():
    issue = {
        "source": "ai",
        "original_text": "在登录界面，点击 ，弹出服务器和机构设置界面窗口。",
        "suggestion": "在登录界面，点击【设置】或对应齿轮图标，弹出服务器和机构设置界面窗口。",
        "description": "UI交互元素完全缺失，疑似图标丢失。",
        "rule": "Operation",
        "category": "操作步骤",
        "chapter": "设置服务器地址和机构代码",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "visual_control_ambiguity"


def test_validate_ai_issue_candidate_rejects_visual_blank_click_with_blank_button_reasoning():
    issue = {
        "source": "ai",
        "original_text": "点击 ，弹出结项确认对话框。",
        "suggestion": "点击【查询】，弹出结项确认对话框。",
        "description": "",
        "rule": "空白按钮无语义，未指明具体按钮，关键操作不可执行。",
        "category": "图文引用",
        "chapter": "项目结项",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "visual_control_ambiguity"


def test_validate_ai_issue_candidate_rejects_blank_click_with_guessed_button_name():
    issue = {
        "source": "ai",
        "original_text": "在界面右上角查询区域，输入流程编号或名称，点击 ，查询目标流程。",
        "suggestion": "在界面右上角查询区域，输入流程编号或名称，点击【查询】，查询目标流程。",
        "description": "",
        "rule": "UI交互元素必须明确标注按钮名称。",
        "category": "图文引用",
        "chapter": "审核项目",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "visual_control_ambiguity"


def test_validate_ai_issue_candidate_rejects_visual_ocr_artifact_text_loss():
    issue = {
        "source": "ai",
        "original_text": "（可选）点击目标物料大类/物料小类【操作】栏目的文 ，对字典值进行多语言翻译。",
        "suggestion": "（可选）点击目标物料大类/物料小类【操作】栏的【翻译】，对字典值进行多语言翻译。",
        "description": "UI交互元素缺失，原文疑似 OCR 识别残片。",
        "rule": "Operation",
        "category": "操作步骤",
        "chapter": "管理物料分类",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "visual_control_ambiguity"


def test_validate_ai_issue_candidate_rejects_low_value_ui_bracket_labeling():
    issue = {
        "source": "ai",
        "original_text": "点击【设计耗材】，进入实验室器具定制界面。",
        "suggestion": "点击【设计耗材】按钮，进入实验室器具定制界面。",
        "description": "UI元素未按规范标注：'【设计耗材】' 作为可点击按钮，但未说明其为按钮。",
        "rule": "Operation",
        "category": "操作步骤",
        "chapter": "实验室器具定制",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "low_value_ui_bracket_labeling"


def test_validate_ai_issue_candidate_keeps_true_missing_ui_object_issue():
    issue = {
        "source": "ai",
        "original_text": "点击【仓库标签】行的,弹出仓库标签编辑窗口。",
        "suggestion": "点击【仓库标签】行末【操作】栏的【编辑】图标，弹出仓库标签编辑窗口。",
        "description": "操作步骤缺少具体操作对象，无法确定点击位置。",
        "rule": "Completeness",
        "category": "操作步骤",
        "chapter": "仓库标签",
        "audit_basis": "basis",
        "context": "字典管理 > 仓库标签。点击【仓库标签】行的,弹出仓库标签编辑窗口。",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["context"])

    assert result.accepted is True
    assert result.reason == "accepted"


def test_validate_ai_issue_candidate_rejects_operation_column_blank_icon_issue():
    issue = {
        "source": "ai",
        "original_text": "点击【操作】列的 ，禁用该物料编码。",
        "suggestion": "点击【操作】列的【禁用】，禁用该物料编码。",
        "description": "UI交互元素缺失，疑似操作列图标在 PDF 文本中丢失。",
        "rule": "Operation",
        "category": "术语",
        "chapter": "启用/禁用物料状态",
        "audit_basis": "basis",
    }

    result = review_validation.validate_ai_issue_candidate(issue, issue["original_text"])

    assert result.accepted is False
    assert result.reason == "visual_control_ambiguity"


def test_pipeline_filters_visual_layout_external_rule_issue():
    issues = [
        {
            "source": "rule",
            "rule": "EXT-R018",
            "category": "格式错误",
            "severity": "suggestion",
            "original_text": "说明书中",
            "suggestion": "说明书中图片和图注需左对齐",
            "description": "说明书中图片和图注需左对齐",
            "audit_basis": "飞书多维表格 - 技术文档评审规则库",
        }
    ]

    selected = review_pipeline.select_review_issues(issues)

    assert selected == []


def test_run_english_heuristic_audit_detects_this_instructions_grammar_issue():
    issues = review_api._run_english_heuristic_audit(
        "This instructions for use describes how to perform sequencing.",
        file_type="pdf",
    )

    assert any(issue.get("rule") == "GRAMMAR-007" for issue in issues)


def test_run_english_heuristic_audit_keeps_official_global_site():
    issues = review_api._run_english_heuristic_audit(
        "For more information, visit https://global-mgitech.com/ for the latest manuals.",
        file_type="pdf",
    )

    assert not any(issue.get("rule") in {"HR001", "HR012"} for issue in issues)


def test_run_manual_engineering_audit_detects_duplicate_sentence_and_step_leadin():
    content = (
        "Ubuntu is a registered trademark of Canonical Ltd.\n\n"
        "Some setup notes are listed here.\n\n"
        "Ubuntu is a registered trademark of Canonical Ltd.\n\n"
        "Perform the following steps:\n"
        "1. Open the software.\n"
        "Perform the following steps:\n"
        "2. Confirm the settings.\n"
    )

    issues = review_api._run_manual_engineering_audit(content, file_type="pdf")
    rules = {issue["rule"] for issue in issues}

    assert "DOC-DUP-001" in rules
    assert "DOC-PROC-002" in rules


def test_run_manual_engineering_audit_skips_duplicate_sentence_for_protocol_like_steps():
    content = (
        "Keep the PCR tube on the magnetic separation rack for 30 seconds, and then remove and discard the supernatant.\n\n"
        "Some intermediate notes are listed here.\n\n"
        "Keep the PCR tube on the magnetic separation rack for 30 seconds, and then remove and discard the supernatant.\n"
    )

    issues = review_api._run_manual_engineering_audit(content, file_type="pdf")

    assert not any(issue["rule"] == "DOC-DUP-001" for issue in issues)


def test_run_manual_engineering_audit_detects_missing_space_before_parentheses():
    issues = review_api._run_manual_engineering_audit(
        "The storage summary includes Distribution(TB) and Capacity(GB).",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-FMT-003" and issue["original_text"] == "Distribution(TB)" for issue in issues)


def test_run_manual_engineering_audit_detects_missing_space_before_units():
    issues = review_api._run_manual_engineering_audit(
        "Description 24VDC, 5A DNBSEQ-E25RS 20VDC, 11.5A DNBSEQ-E25ARS 100-240 V~, 50 /60 Hz, 300 VA ±10% II",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-UNIT-001" and issue["original_text"] == "24VDC" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_revision_table_column_layout():
    issues = review_api._run_chinese_human_baseline_rules(
        "修订记录 修订版本 发布日期 版本 2.0 1.0 2024-01-01 2023-01-01",
    )

    assert any(issue["rule"] == "CYY-CN-LAYOUT-008" and issue["original_text"] == "版本 2.0 1.0" for issue in issues)


def test_run_manual_engineering_audit_keeps_global_site_for_english_manual():
    issues = review_api._run_manual_engineering_audit(
        "Download the instructions for use from https://global-mgitech.com.",
        file_type="pdf",
    )

    assert not any(issue["rule"] == "DOC-URL-001" for issue in issues)


def test_should_skip_rule_match_skips_ext_r005_global_site_for_english_manual():
    rule = SimpleNamespace(rule_no="EXT-R005")
    content = "For more information, visit https://global-mgitech.com for the latest manuals."
    match = next(re.finditer(r"https://global-mgitech\.com", content))

    assert review_api._should_skip_rule_match(rule, match, content, "en", file_type="pdf") is True


def test_run_manual_engineering_audit_detects_neither_not():
    issues = review_api._run_manual_engineering_audit(
        "DNBSEQ-G400RS sequencing software version and read length in the case of neither pooling samples not sequencing the barcode.",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-GRAM-001" for issue in issues)


def test_run_manual_engineering_audit_detects_a_appropriate():
    issues = review_api._run_manual_engineering_audit(
        "Use a appropriate cell strainer for the sample.",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-GRAM-002" for issue in issues)


def test_run_manual_engineering_audit_detects_library_spelling():
    issues = review_api._run_manual_engineering_audit(
        "Table 41 DNBSEQ-G400RS making DNB requirements. Libary type cDNA library, TCR&BCR libraries.",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-SPELL-001" for issue in issues)


def test_run_manual_engineering_audit_detects_basecall_repeat():
    issues = review_api._run_manual_engineering_audit(
        "Basecall version Basecall Basecall_1.0.8.208 or later version.",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-DUP-006" for issue in issues)


def test_run_manual_engineering_audit_detects_by_use():
    issues = review_api._run_manual_engineering_audit(
        "It is recommended to mix the sample thoroughly by use the pipette.",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-GRAM-003" for issue in issues)


def test_run_manual_engineering_audit_detects_missing_data_placeholder():
    content = (
        "About the sequencing set\n"
        "Table 7 Recommended library insert size\n"
        "Model\n"
        "Recommended insert size (bp)\n"
        "Data output (GB/flow cell)\n\n"
        "E25 App-D FCU SE100 200 to 400\n"
        "About\n\n"
        "E25 App-D FCS SE100 200 to 400 About 2.5\n"
    )

    issues = review_api._run_manual_engineering_audit(content, file_type="pdf")

    assert any(issue["rule"] == "DOC-DATA-001" and issue["original_text"] == "About" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_temperature_range():
    issues = review_api._run_chinese_human_baseline_rules(
        "y Cytoactivity > 80% y Clumping rate < 5% y Impurity rate < 5% y Cytoactivity < 5% y Clumping rate < 5% y Impurity rate < 5% Recommended cell input Recommended cell concentration (cell/μL) 2 ℃ to 8 ℃ (36 ℉) and -25 ℃ to -15 ℃.",
    )

    assert any(issue["rule"] == "CYY-CN-UNIT-006" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_navigation_entry_mismatch():
    issues = review_api._run_chinese_human_baseline_rules(
        "系统管理模块 点击左侧导航栏的【物料管理】，默认进入用户管理界面。"
    )

    assert any(issue["rule"] == "CYY-CN-NAV-001" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_wrong_material_code_reference():
    issues = review_api._run_chinese_human_baseline_rules(
        "目录包含 管理物料编码规则 30。 物料编码规则设置参考第70 页“编辑字典”。"
    )

    assert any(issue["rule"] == "CYY-CN-REF-003" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_repeated_category_suffix():
    issues = review_api._run_chinese_human_baseline_rules(
        "可查看物料、耗材、测试物料大类、试剂等大类，并按类别筛选右侧物料列表。"
    )

    assert any(issue["rule"] == "CYY-CN-DUP-002" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_new_create_term_drift():
    issues = review_api._run_chinese_human_baseline_rules(
        "在出库登记界面，点击【新增】，进入新建出库登记界面。"
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-016" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_step_reference_style_drift():
    issues = review_api._run_chinese_human_baseline_rules(
        "4) （可选）如需为物料配置多个属性，可重复步骤3。"
    )

    assert any(issue["rule"] == "CYY-CN-STEP-003" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_cytoactivity_threshold():
    issues = review_api._run_chinese_human_baseline_rules(
        "y Cytoactivity > 80% y Clumping rate < 5% y Impurity rate < 5% y Cytoactivity < 5% y Clumping rate < 5% y Impurity rate < 5% Recommended cell input Recommended cell concentration (cell/μL)",
    )

    assert any(issue["rule"] == "CYY-CN-RANGE-002" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_strip_case():
    issues = review_api._run_chinese_human_baseline_rules(
        "8-Strip tube × 2 8-Strip tube × 2 8-Strip tube × 2 8-Strip tube × 2",
    )

    assert any(issue["rule"] == "CYY-CN-TERM-005" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_storage_term():
    issues = review_api._run_chinese_human_baseline_rules(
        "用于连接扫码枪或外部储存设备。",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-016" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_login_term():
    issues = review_api._run_chinese_human_baseline_rules(
        "用户账户和登陆密码信息",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-017" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_extra_particle():
    issues = review_api._run_chinese_human_baseline_rules(
        "在仓门开启时，检查的托盘表面是否有灰尘。",
    )

    assert any(issue["rule"] == "CYY-CN-GRAMMAR-007" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_pipette_term():
    issues = review_api._run_chinese_human_baseline_rules(
        "用移液枪将制备好的DNB样本加入到测序载片。",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-018" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_storage_environment_term():
    issues = review_api._run_chinese_human_baseline_rules(
        "运输/ 储存环 境温度和相对湿度。",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-019" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_long_storage_phrase():
    issues = review_api._run_chinese_human_baseline_rules(
        "用于连接扫码枪或外部储存设备。",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-020" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_login_context_phrase():
    issues = review_api._run_chinese_human_baseline_rules(
        "参数 y 内容：设备运行状况数据/日志 y 储存：主机硬盘/计算模块 y 内容：用户账户和登陆密码信息 y 储存：加密储存于主机硬盘。",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-021" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_extra_word_phrase():
    issues = review_api._run_chinese_human_baseline_rules(
        "本框，用弹出的软键盘中输入时间。",
    )

    assert any(issue["rule"] == "CYY-CN-GRAMMAR-008" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_pipette_context_phrase():
    issues = review_api._run_chinese_human_baseline_rules(
        "用移液枪将制备好的DNB样本加入到测序载片。",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-022" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_transport_storage_phrase():
    issues = review_api._run_chinese_human_baseline_rules(
        "污染等级 使用场地 运输/ 储存环 境 温度 相对湿度 随机附件 详见装箱清单",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-023" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_reference_page():
    issues = review_api._run_chinese_human_baseline_rules(
        "分析，此处DNB ID 需与第33 页“装载样本”一致。",
    )

    assert any(issue["rule"] == "CYY-CN-REF-001" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_external_storage_device():
    issues = review_api._run_chinese_human_baseline_rules(
        "用于连接扫码枪或外部存储设备。",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-024" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_login_password_phrase():
    issues = review_api._run_chinese_human_baseline_rules(
        "参数 y 内容：用户账户和登陆密码信息",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-025" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_read_length_term():
    issues = review_api._run_chinese_human_baseline_rules(
        "一链读长和二连读长暗反应的循环数（cycle）。",
    )

    assert any(issue["rule"] == "CYY-CN-SPELL-003" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_repeated_page_number():
    issues = review_api._run_chinese_human_baseline_rules(
        "正文第一页\n21\f正文第二页\n23\f正文第三页\n24\f正文第四页\n23",
    )

    page_issue = next(issue for issue in issues if issue["rule"] == "CYY-CN-PAGE-001")
    assert page_issue["original_text"] == "页码 23（第4个PDF页面）"


def test_run_chinese_human_baseline_rules_detects_suspicious_address():
    issues = review_api._run_chinese_human_baseline_rules(
        "生物医学科技（绍兴）有限公司 绍兴市越城区稽山街道越南大道2 号C2 栋2 层",
    )

    assert any(issue["rule"] == "CYY-CN-ADDR-001" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_cable_typo():
    issues = review_api._run_chinese_human_baseline_rules(
        "扫码枪 线揽 本表格依据GB 26572-2025 的规定编制。",
    )

    assert any(issue["rule"] == "CYY-CN-SPELL-005" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_jiaohu_typo():
    issues = review_api._run_chinese_human_baseline_rules(
        "触摸屏 显示用户交户界面。",
    )

    issue = next(item for item in issues if item["rule"] == "CYY-CN-SPELL-007")
    assert issue["original_text"] == "交户"
    assert "交互" in issue["suggestion"]


def test_run_chinese_human_baseline_rules_detects_similar_ui_labels():
    issues = review_api._run_chinese_human_baseline_rules(
        "主界面显示【长按开启】。\n操作时请长按【长按启动】按钮旁的指纹区域。",
    )

    issue = next(item for item in issues if item["rule"] == "CYY-CN-UI-002")
    assert "长按开启" in issue["original_text"]
    assert "长按启动" in issue["original_text"]


def test_run_chinese_human_baseline_rules_ignores_similar_english_ui_codes():
    issues = review_api._run_chinese_human_baseline_rules(
        "载片类型包括：【FTN】、【FTS】、【FTL】。",
    )

    assert not any(issue["rule"] == "CYY-CN-UI-002" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_wrong_step_page_reference():
    content = (
        "具体操作，参考第2页“加载DNB”步骤5。\n01\f"
        "加载DNB\n1. 打开加载仪的翻盖。\n02\f"
        "4. 轻轻按压翻盖。\n5. 上下滑动主界面，选择载片类型。\n03"
    )
    issues = review_api._run_chinese_human_baseline_rules(content)

    issue = next(item for item in issues if item["rule"] == "CYY-CN-REF-005")
    assert "步骤5" in issue["original_text"]
    assert "第 3 页" in issue["suggestion"]


def test_run_chinese_human_baseline_rules_keeps_matching_step_page_reference():
    content = (
        "具体操作，参考第2页“加载DNB”步骤1。\n01\f"
        "加载DNB\n1. 打开加载仪的翻盖。\n02\f"
        "4. 轻轻按压翻盖。\n5. 上下滑动主界面，选择载片类型。\n03"
    )
    issues = review_api._run_chinese_human_baseline_rules(content)

    assert not any(issue["rule"] == "CYY-CN-REF-005" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_figure_table_spacing():
    issues = review_api._run_chinese_human_baseline_rules(
        "入库登记\n在入库登记界面，可新建、查看待提交和已完成的入库登记单。\n图 1 入库登记界面\n项目 说明\n1 左侧导航栏",
    )

    assert any(issue["rule"] == "CYY-CN-LAYOUT-007" for issue in issues)


def test_restore_high_value_rule_issues_keeps_page_number_issue():
    selected = []
    candidates = [
        {
            "rule": "CYY-CN-PAGE-001",
            "category": "页码异常",
            "original_text": "23",
            "description": "页码重复",
            "source": "rule",
        }
    ]

    restored = review_api._restore_high_value_rule_issues(selected, candidates)

    assert restored == candidates


def test_run_chinese_human_baseline_rules_detects_hard_sentence():
    issues = review_api._run_chinese_human_baseline_rules(
        "禁止使用与设备零部件或设备内所含材料发生化学反应而引起危险的清洗剂或消毒剂。",
    )

    assert any(issue["rule"] == "CYY-CN-STYLE-002" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_ocr_storage_fragment():
    issues = review_api._run_chinese_human_baseline_rules(
        "合。 部储存设",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-026" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_storage_colon_phrase():
    issues = review_api._run_chinese_human_baseline_rules(
        "参数 y 内容：设备运行状况数据/ 日志 y 储存：主机硬盘/ 计算模块。",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-027" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_peripheral_device_term():
    issues = review_api._run_chinese_human_baseline_rules(
        "空间足够容纳相关配套或外围设备。",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-028" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_peripheral_device_ocr_fragment():
    issues = review_api._run_chinese_human_baseline_rules(
        "建议参 或外围设 热、线缆",
    )

    assert any(issue["rule"] == "CYY-CN-CONSIST-029" for issue in issues)


def test_run_chinese_human_baseline_rules_detects_reference_crosscheck():
    issues = review_api._run_chinese_human_baseline_rules(
        "该测序方案中如需自定义Barcode 文件，参考第19 页“Barcode 管理界面”。",
    )

    assert any(issue["rule"] == "CYY-CN-REF-002" for issue in issues)


def test_run_manual_engineering_audit_detects_hyphen_term_drift():
    issues = review_api._run_manual_engineering_audit(
        "The single-base sequencing workflow is described below. The sing-base analysis is used in the appendix.",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-TERM-003" and issue["original_text"] == "sing-base" for issue in issues)


def test_run_manual_engineering_audit_skips_case_only_hyphen_variants():
    issues = review_api._run_manual_engineering_audit(
        "Double-click the icon to continue. If needed, double-click the icon again.",
        file_type="pdf",
    )

    assert not any(issue["rule"] == "DOC-TERM-003" for issue in issues)


def test_run_manual_engineering_audit_detects_repeated_short_phrase():
    issues = review_api._run_manual_engineering_audit(
        "Click Back to return to the to the analysis result directory.",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-DUP-004" and issue["original_text"] == "to the to the" for issue in issues)


def test_run_manual_engineering_audit_skips_non_target_repeated_short_phrase():
    issues = review_api._run_manual_engineering_audit(
        "The samples are grouped by groups by category in the OCR output.",
        file_type="pdf",
    )

    assert not any(issue["rule"] == "DOC-DUP-004" for issue in issues)


def test_run_manual_engineering_audit_skips_title_case_repeated_phrase():
    issues = review_api._run_manual_engineering_audit(
        "Electrical safety Electrical safety is listed in the chapter title.",
        file_type="pdf",
    )

    assert not any(issue["rule"] == "DOC-DUP-004" for issue in issues)


def test_run_manual_engineering_audit_skips_cross_sentence_case_shift_phrase():
    issues = review_api._run_manual_engineering_audit(
        "Use the device The device can then be restarted.",
        file_type="pdf",
    )

    assert not any(issue["rule"] == "DOC-DUP-004" for issue in issues)


def test_run_manual_engineering_audit_detects_self_referential_imperative():
    issues = review_api._run_manual_engineering_audit(
        "Perform the following steps: 1. Power off the power.",
        file_type="pdf",
    )

    assert any(issue["rule"] == "DOC-DUP-005" and issue["original_text"] == "Power off the power" for issue in issues)


def test_review_ai_chunk_timeout_seconds_uses_higher_default_for_large_chunks(monkeypatch):
    monkeypatch.delenv("REVIEW_AI_CHUNK_TIMEOUT", raising=False)

    timeout = review_api._review_ai_chunk_timeout_seconds("A" * 3000)

    assert timeout == 50.0


def test_log_review_ai_usage_prints_summary(monkeypatch, capsys):
    monkeypatch.setattr(
        review_api.ai_client,
        "summarize_usage_events",
        lambda request_label=None, review_id=None, limit=200: {
            "calls": 2,
            "prompt_tokens": 120,
            "completion_tokens": 30,
            "total_tokens": 150,
            "providers": {"kimi": 2},
        },
    )

    review_api._log_review_ai_usage(66, "review.audit_chunk", "AI深度审核Token统计")

    output = capsys.readouterr().out
    assert "AI深度审核Token统计" in output
    assert "total_tokens=150" in output


def test_review_ai_chunk_limit_defaults_to_ten(monkeypatch):
    monkeypatch.delenv("REVIEW_AI_MAX_CHUNKS", raising=False)

    assert review_api._review_ai_chunk_limit() == 10


def test_review_ai_chunk_limit_uses_dynamic_value_under_budget(monkeypatch):
    monkeypatch.delenv("REVIEW_AI_MAX_CHUNKS", raising=False)

    assert review_api._review_ai_chunk_limit(6000) == 3
    assert review_api._review_ai_chunk_limit(12000) == 5
    assert review_api._review_ai_chunk_limit(30000) == 10


def test_review_ai_chunk_limit_caps_long_documents_by_env_budget(monkeypatch):
    monkeypatch.setenv("REVIEW_AI_MAX_CHUNKS", "6")

    assert review_api._review_ai_chunk_limit(60000) == 6


def test_review_ai_token_budget_defaults_to_zero(monkeypatch):
    monkeypatch.delenv("REVIEW_AI_TOKEN_BUDGET", raising=False)

    assert review_api._review_ai_token_budget() == 0


def test_run_ai_deep_review_stops_when_budget_reached(monkeypatch):
    chunks = [
        (1, 0, "chunk-1"),
        (2, 100, "chunk-2"),
        (3, 200, "chunk-3"),
    ]
    processed = []
    usage_state = {"tokens": 0}

    monkeypatch.setattr(review_api, "_iter_ai_audit_chunks", lambda content: chunks)
    monkeypatch.setattr(review_api, "_review_ai_chunk_limit", lambda content_length: 3)
    monkeypatch.setattr(review_api, "_review_ai_token_budget", lambda: 200)
    monkeypatch.setattr(review_api, "set_progress", lambda *args, **kwargs: None)
    monkeypatch.setattr(review_api, "_select_relevant_ai_review_basis", lambda chunk, sections: "basis")

    def fake_run_cached_ai_chunk_review(review_id, chunk, document_language, audit_basis, chunk_timeout, force_provider=None, **kwargs):
        processed.append(chunk)
        usage_state["tokens"] += 120
        return ([{"rule": "AI", "chapter": ""}], False)

    monkeypatch.setattr(review_api, "_run_cached_ai_chunk_review", fake_run_cached_ai_chunk_review)
    monkeypatch.setattr(
        review_api.ai_client,
        "summarize_usage_events",
        lambda request_label=None, review_id=None, limit=200: {
            "calls": len(processed),
            "prompt_tokens": max(usage_state["tokens"] - 20 * len(processed), 0),
            "completion_tokens": 20 * len(processed),
            "total_tokens": usage_state["tokens"],
            "providers": {"kimi": len(processed)} if processed else {},
        },
    )

    issues, chunk_meta = review_api._run_ai_deep_review(9, "demo content", "en", [{"text": "basis"}])

    assert processed == ["chunk-1", "chunk-2"]
    assert len(issues) == 2
    assert len(chunk_meta) == 2  # P0-3: 新增分块元数据返回


def test_chinese_human_baseline_rules_cover_common_spellcheck_sample_cases():
    content = (
        "试剂盒组份包括 Buffer。使用移液枪吸取 10ul 溶液，反应体系总体积为 20uL，在室温下静置 5 mins。"
        "将样本加入至离心管中，该试剂盒可用于用于 RNA 提取。"
        "否则的话，该步骤非常的重要。建议用户在使用前仔细阅读说明书。"
        "PCR 扩增条件：95℃ 预变性。设置转数为12000rpm，避免污然样本。"
        "该产品的有效期为 12 个月，请在有效期内使用，过期请勿使用。"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)
    originals = {issue["original_text"] for issue in issues}
    rules = {issue["rule"] for issue in issues}

    assert {
        "组份",
        "10ul",
        "20uL",
        "5 mins",
        "加入至",
        "可用于用于",
        "否则的话",
        "非常的重要",
        "建议用户",
        "95℃",
        "转数",
        "12000rpm",
        "污然",
        "过期请勿使用",
    }.issubset(originals)
    assert {
        "CYY-CN-SPELL-001",
        "CYY-CN-UNIT-001",
        "CYY-CN-UNIT-002",
        "CYY-CN-UNIT-003",
        "CYY-CN-TERM-004",
        "CYY-CN-UNIT-004",
        "CYY-CN-SPELL-002",
        "CYY-CN-GRAMMAR-001",
        "CYY-CN-GRAMMAR-002",
        "CYY-CN-GRAMMAR-003",
        "CYY-CN-GRAMMAR-004",
        "CYY-CN-GRAMMAR-005",
        "CYY-CN-GRAMMAR-006",
    }.issubset(rules)


def test_chinese_human_baseline_rules_cover_format_mixed_language_and_redundancy_cases():
    content = (
        "操作步骤如下:1. 加入试剂。注意事项：a)避免污染。请勿将试剂暴露于强光下!"
        "加入 10μL Buffer，5μL Enzyme 和 2μL Primer。该步骤非常重要……请严格按照规范执行。"
        "Q: 使用\"引物\"还是'引物'？Q: PCR(Polymerase Chain Reaction) 的优化条件？"
        "记录日期：2026-07-27、2026/07/27、2026.07.27。温度控制：37°C、37℃、37 度 C。"
        "反应时间：min、分钟、mins。章节编号：1.、1）、(1)、①。温度要求 37°C 或室温。"
        "使用 RNA extraction kit 进行 RNA 提取，将 sample 加入 collection tube 中。"
        "PCR 产物通过 Agarose Gel 电泳检测。设置 thermocycler 程序。使用 ddH2O 配制溶液。OD260/OD280 比值应在 1.8-2.0 之间。"
        "该试剂盒仅供科研使用，不得用于临床诊断。"
        "操作过程中应注意安全，注意安全事项，确保安全操作。"
        "结果判读：根据说明书中的结果判读方法进行判读。"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)
    originals = {issue["original_text"] for issue in issues}
    rules = {issue["rule"] for issue in issues}

    assert {
        "如下:1.",
        "a)避免污染",
        "10μL Buffer，5μL Enzyme 和",
        "强光下!",
        "非常重要……",
        "'引物'",
        "PCR(Polymerase",
        "2026-07-27、2026/07/27、2026.07.27",
        "37°C、37℃、37 度 C",
        "min、分钟、mins",
        "1.、1）、(1)、①",
        "37°C 或室温",
        "RNA extraction kit",
        "sample",
        "collection tube",
        "Agarose Gel",
        "thermocycler",
        "ddH2O",
        "OD",
        "仅供科研使用，不得用于临床诊断",
        "应注意安全，注意安全事项，确保安全操作",
        "结果判读：根据说明书中的结果判读方法进行判读",
    }.issubset(originals)
    assert {
        "CYY-CN-PUNCT-001",
        "CYY-CN-PUNCT-002",
        "CYY-CN-PUNCT-006",
        "CYY-CN-PUNCT-003",
        "CYY-CN-PUNCT-007",
        "CYY-CN-PUNCT-004",
        "CYY-CN-PUNCT-005",
        "CYY-CN-FORMAT-001",
        "CYY-CN-FORMAT-002",
        "CYY-CN-FORMAT-003",
        "CYY-CN-FORMAT-004",
        "CYY-CN-LOGIC-001",
        "CYY-CN-MIXED-001",
        "CYY-CN-MIXED-002",
        "CYY-CN-MIXED-003",
        "CYY-CN-MIXED-004",
        "CYY-CN-MIXED-005",
        "CYY-CN-MIXED-006",
        "CYY-CN-MIXED-007",
        "CYY-CN-REDUNDANCY-001",
        "CYY-CN-REDUNDANCY-002",
        "CYY-CN-REDUNDANCY-003",
    }.issubset(rules)


def test_chinese_human_baseline_rules_cover_consistency_logic_and_product_cases():
    content = (
        "本文档详细介绍了单细胞 RNA 文库制备试剂盒套装。"
        "图1标题为流程图，图2标题为示意图。"
        "前文用 Catalog Number，后文用 Cat. No.。表格1使用 Cat.No。"
        "使用 PBMC 和外周血单个核细胞进行分析。RNA 完整性对实验结果至关重要，RNA 质量直接影响数据可靠性。"
        "表3列出了试剂盒的5个组分。步骤3要求使用试剂A进行反应。声称无 RNase 污染。"
        "货号：H-020-000898-01。规格：16 RXN / 4 RXN。请将规格单位 RXN 改为反应。"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)
    originals = {issue["original_text"] for issue in issues}
    rules = {issue["rule"] for issue in issues}

    assert {
        "套装",
        "示意图",
        "Catalog Number",
        "Cat. No.",
        "外周血单个核细胞",
        "RNA 质量",
        "5个组分",
        "试剂A",
        "无 RNase 污染。",
        "H-020-000898-01",
        "RXN",
    }.issubset(originals)
    assert {
        "CYY-CN-CONSIST-003",
        "CYY-CN-CONSIST-004",
        "CYY-CN-CONSIST-005",
        "CYY-CN-CONSIST-006",
        "CYY-CN-CONSIST-007",
        "CYY-CN-LOGIC-002",
        "CYY-CN-LOGIC-003",
        "CYY-CN-LOGIC-004",
        "CYY-CN-PRODUCT-001",
        "CYY-CN-PRODUCT-002",
        "CYY-CN-PRODUCT-003",
    }.issubset(rules)


def test_chinese_human_baseline_rules_cover_storage_login_and_power_unit_format_cases():
    content = (
        "设备运行状况数据存储于主机硬盘，用户账户和登陆密码信息也会储存。"
        "首次登录前请确认设备配置。"
        "电源参数为 24VDC,5A 20VDC,11.5A 100-240V~,50/60 Hz,300VA。"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)
    originals = {issue["original_text"] for issue in issues}
    rules = {issue["rule"] for issue in issues}

    assert {
        "存储",
        "登录",
        "24VDC",
        "5A",
        "20VDC",
        "11.5A",
        "300VA",
    }.issubset(originals)
    assert {
        "CYY-CN-CONSIST-014",
        "CYY-CN-CONSIST-015",
        "CYY-CN-UNIT-005",
    }.issubset(rules)


def test_chinese_human_baseline_rules_cover_spacing_figure_and_info_completeness_cases():
    content = (
        "前面试剂用量为 10 μL，后续实验使用 10μL。"
        "图1引用为图 1，展示文库制备流程。"
        "储存条件：-20°C 避光保存。警告：避免反复冻融。"
        "步骤3要求使用试剂A进行反应。操作时间约 30 分钟。"
        "保质期标注可靠，生产日期格式规范。"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)
    originals = {issue["original_text"] for issue in issues}
    rules = {issue["rule"] for issue in issues}

    assert {
        "10μL",
        "图 1",
        "-20°C 避光保存。警告：避免反复冻融",
        "操作时间约 30 分钟",
        "生产日期格式规范",
        "试剂A",
    }.issubset(originals)
    assert {
        "CYY-CN-FORMAT-005",
        "CYY-CN-CONSIST-008",
        "CYY-CN-LOGIC-003",
        "CYY-CN-LOGIC-005",
        "CYY-CN-LOGIC-006",
        "CYY-CN-INFO-001",
    }.issubset(rules)


def test_chinese_human_baseline_rules_cover_alpha_lab_manual_gaps():
    content = (
        "需先选择标准计量单位，再选择包装单位及转化系数。例如：标准计量单位选‘个’，包装单位选‘箱’且转换系数设为10。"
        "首次登录前点击【图标】打开弹窗确认服务地址和机构。具体操作，参考第31页‘设置服务器地址和机构代码’。"
        "在左侧菜单栏点击【系统管理】。其余步骤写为在左侧导航栏点击【角色管理】。"
        "Q：如何区分同步物料与自增物料？A：系统支持手动新增普通物料。"
        "图59图注为修改密码界面，正文步骤均写为重置密码。"
        "处理器：X86 架构，8核及以上。默认安装在C盘，可点击【Browse】可选择目标安装目录。"
        "按照界面指引，拖拽αLab Studio 软件图标到右侧的Application 文件夹进行安装。安装完成后，点击【Install】开始安装系统，点击【Finish】完成安装。"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)
    originals = {issue["original_text"] for issue in issues}
    rules = {issue["rule"] for issue in issues}

    assert {
        "转化系数",
        "服务地址",
        "左侧菜单栏",
        "自增物料",
        "修改密码界面",
        "X86 架构",
        "可点击【Browse】可选择",
    }.issubset(originals)
    assert any("【Install】" in issue["original_text"] and "【Finish】" in issue["original_text"] for issue in issues)
    assert {
        "CYY-CN-CONSIST-009",
        "CYY-CN-CONSIST-010",
        "CYY-CN-CONSIST-011",
        "CYY-CN-CONSIST-012",
        "CYY-CN-CONSIST-013",
        "CYY-CN-TERM-005",
        "CYY-CN-GRAMMAR-007",
        "CYY-CN-LOGIC-007",
    }.issubset(rules)


def test_chinese_human_baseline_rules_detect_macos_install_conflict_with_intermediate_steps():
    content = (
        "在MacOS 端安装本系统\n"
        "1. 下载MacOS 安装包，双击该文件进入系统安装流程。\n"
        "2. 按照界面指引，拖拽αLab Studio 软件图标到右侧的Application 文件夹进行安装。\n"
        "3. 首次安装本平台，计算机将提示无法打开本系统。点击【好】，并进入【系统设置】> 【隐私与安全性】进入隐私与安全性设置界面。选择【仍要打开】进行软件安全确认。\n"
        "4. 如系统再次提醒，点击【打开】打开本系统。\n"
        "5. 点击【Install】，开始安装系统。安装完成后，点击【Finish】完成安装。\n"
        "6. 安装完成后，可通过启动台或应用中心，点击 打开本系统。\n"
        "在Android 平板端安装本系统"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)

    assert any(issue["rule"] == "CYY-CN-LOGIC-007" for issue in issues)


def test_chinese_human_baseline_rules_detect_adjacent_duplicate_steps():
    content = (
        "14. 核对入库信息。"
        "15. 在入库定位界面右侧选择目标库位，左侧填写待分配物料数量，点击【保存】，回到入库定位页签。"
        "16. 在入库定位界面左侧填写待分配物料数量，点击【保存】，回到入库定位页签。"
        "17. 继续下一步操作。"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)
    assert any(issue["rule"] == "CYY-CN-STRUCT-001" for issue in issues)
    assert any("步骤 15 和步骤 16" in issue["suggestion"] for issue in issues if issue["rule"] == "CYY-CN-STRUCT-001")


def test_chinese_human_baseline_rules_skip_adjacent_steps_with_different_actions_or_targets():
    content = (
        "1. 按需修改字典值code、字典值名称及颜色。"
        "2. 点击【添加】，按需填写字典值code 及字典值名称，并设置标签颜色。"
        "3. 点击【物料大类】右侧的【添加】，按需填写字典值code 及名称。"
        "4. 点击【物料小类】右侧的【添加】，按需填写字典值code 及名称，为对应大类添加小类。"
        "5. 点击【新增】按钮，填写物料编码，点击【确定】保存。"
        "6. 点击【新增】按钮，填写物料名称，点击【确定】保存。"
        "7. 取样本加入裂解液，充分混匀，室温静置。"
        "8. 取样本加入裂解液，充分混匀，室温静置，12000rpm 离心 10 分钟。"
        "9. Click the Add button to open the dialog."
        "10. Click the Edit button to open the dialog."
        "11. 打开系统设置，选择高级选项，点击保存。"
        "12. 打开系统设置，选择高级选项，点击保存。"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)

    struct_issues = [issue for issue in issues if issue["rule"] == "CYY-CN-STRUCT-001"]

    assert not any("填写物料名称" in issue["original_text"] for issue in struct_issues)
    assert not any("12000rpm 离心 10 分钟" in issue["original_text"] for issue in struct_issues)
    assert not any("Click the Edit button" in issue["original_text"] for issue in struct_issues)
    assert any("12. 打开系统设置，选择高级选项，点击保存。" == issue["original_text"] for issue in struct_issues)


def test_chinese_human_baseline_rules_skip_cross_section_step_capture():
    content = (
        "1. 点击弹窗中的【立即更新】，系统开始下载最新版本安装包。\n"
        "2. 下载完成后，在弹出的界面点击【立即安装】。\n"
        "3. 平台关闭并弹出安装窗口，在窗口中点击【Install】。安装完成后，点击【Finish】，完成更新。\n"
        "17\n更新系统\n非强制更新\n"
        "1) 在弹窗中点击【稍后再说】。\n"
        "2) 在导航栏，点击 ，进入设置界面。默认进入账号与安全界面。\n"
        "3) 点击【通用设置】>【检查更新】，系统将检查版本。\n"
    )

    issues = review_api._run_chinese_human_baseline_rules(content)

    assert not any(issue["rule"] == "CYY-CN-STRUCT-001" and "【Install】" in issue["original_text"] for issue in issues)


def test_known_false_positive_filter_drops_placeholder_and_duplicate_ai_items():
    ai_placeholder = {
        'original_text': '（版本号待补充）',
        'rule': '信息完整性：关键章节内容缺失',
        'source': 'ai',
        'description': '',
        'suggestion': '需补充版本号',
        'audit_basis': '',
    }
    ai_duplicate = {
        'original_text': '10μL',
        'rule': '单位规则：数字与单位之间需保留空格',
        'source': 'ai',
        'description': '',
        'suggestion': '10 μL',
        'audit_basis': '',
    }
    rule_duplicate = {
        'original_text': '用于用于',
        'rule': 'CYY-CN-DUP-002',
        'source': 'rule',
        'description': '连续重复中文词可能是复制或编辑残留。',
        'suggestion': '用于',
        'audit_basis': 'baseline',
    }

    assert review_api._should_drop_known_false_positive_issue(ai_placeholder) is True
    assert review_api._should_drop_known_false_positive_issue(ai_duplicate) is True
    assert review_api._should_drop_known_false_positive_issue(rule_duplicate) is True


def test_known_false_positive_filter_drops_official_global_site_issue():
    issue = {
        'original_text': 'https://global-mgitech.com/',
        'rule': 'EXT-R005',
        'source': 'ai',
        'description': '海外官网地址应统一。',
        'suggestion': '建议替换为 https://global-mgitech.com/',
        'audit_basis': '公司特定规范 - 海外官网地址',
        'context': 'For more information, visit https://global-mgitech.com/ for the latest manuals.',
    }

    assert review_api._should_drop_known_false_positive_issue(issue) is True


def test_run_chinese_human_baseline_rules_keeps_correct_ml_spacing():
    content = '表 4 推荐耗材清单\n低吸附 EP 管 0.2 mL，用于样本制备。'

    issues = review_api._run_chinese_human_baseline_rules(content)

    assert not any(issue['rule'] == 'CYY-CN-FMT-002' for issue in issues)


def test_run_chinese_human_baseline_rules_detects_missing_ml_spacing():
    content = '表 4 推荐耗材清单\n低吸附 EP 管 0.2mL，用于样本制备。'

    issues = review_api._run_chinese_human_baseline_rules(content)

    assert any(issue['rule'] == 'CYY-CN-FMT-002' and issue['original_text'] == '0.2mL' for issue in issues)


def test_clean_issue_suggestion_for_display_strips_chinese_quotes():
    assert review_api._clean_issue_suggestion_for_display('建议统一为“2 ℃ ~ 8 ℃”') == '2 ℃ ~ 8 ℃'


def test_normalize_ext_r021_uses_grammar_category_and_replacement_suggestion():
    issue = {
        'original_text': '如操作不当或不避免',
        'rule': 'EXT-R021',
        'category': '格式错误',
        'suggestion': '避免"不避免"类语法错误。"如操作不当或不避免"不通顺，应改为"如不按照说明操作"或"如操作不当或未加避免"',
        'description': '避免"不避免"类语法错误。',
    }

    review_api._sanitize_issue_suggestions([issue])

    assert issue['category'] == '语法错误'
    assert issue['suggestion'] == '如未按照说明进行操作'
    assert issue['description'] == '原文中的“不避免”搭配不通顺，应改为“如未按照说明进行操作”。'


def test_known_false_positive_filter_drops_noop_chinese_baseline_issue():
    issue = {
        'original_text': '外部存储设备',
        'rule': 'CYY-CN-CONSIST-024',
        'source': 'rule',
        'suggestion': '建议统一为“外部存储设备”',
        'description': '用于连接扫码枪或外部存储设备。',
        'audit_basis': 'CYY人工审核经验基线 - 外部存储设备直写',
        'context': '用于连接扫码枪或外部存储设备',
    }

    assert review_api._should_drop_known_false_positive_issue(issue) is True


def test_known_false_positive_filter_keeps_unit_spacing_suggestion():
    issue = {
        'original_text': '24VDC，5A',
        'rule': 'CYY-CN-FMT-003',
        'source': 'rule',
        'suggestion': '24 VDC，5 A',
        'description': '单位前面要加空格。',
        'audit_basis': 'CYY人工审核经验基线 - 电源规格单位空格',
        'context': '说明 24VDC，5A 100-240V~，50/60 Hz，300VA',
    }

    assert review_api._should_drop_known_false_positive_issue(issue) is False


def test_known_false_positive_filter_drops_space_term_usage_issue():
    issue = {
        'original_text': '空格',
        'rule': 'EXT-R009',
        'source': 'rule',
        'suggestion': '文档内容中不得出现多余的空格、空行等排版问题',
        'description': '检查多余空格。',
        'audit_basis': '通用格式规则',
        'context': 'Barcode 序列间禁止使用空格，长度需大于等于 6，且相等。',
    }

    assert review_api._should_drop_known_false_positive_issue(issue) is True


def test_known_false_positive_filter_drops_rohs_title_fragment_issue():
    issue = {
        'original_text': '品中有害',
        'rule': '需包含产品中有害物质的名称及含有物质表',
        'source': 'ai',
        'category': '合规问题',
        'severity': 'serious',
        'suggestion': '建议按下方修改。需包含产品中有害物质的名称及含有物质表',
        'description': '需包含产品中有害物质的名称及含有物质表',
        'audit_basis': '合规规则',
        'context': '产品中有害物质的名称及含有物质表',
    }

    assert review_api._should_drop_known_false_positive_issue(issue) is True


def test_finalize_review_issues_filters_known_false_positives_by_default():
    issues = [
        {
            'original_text': '空格',
            'rule': 'EXT-R009',
            'source': 'rule',
            'suggestion': '文档内容中不得出现多余的空格、空行等排版问题',
            'description': '检查多余空格。',
            'audit_basis': '通用格式规则',
            'context': 'Barcode 序列间禁止使用空格，长度需大于等于 6，且相等。',
        },
        {
            'original_text': '物质的名',
            'rule': 'EXT-R019',
            'source': 'ai',
            'category': '合规问题',
            'severity': 'serious',
            'suggestion': '建议按下方修改。需包含产品中有害物质的名称及含有物质表',
            'description': '需包含产品中有害物质的名称及含有物质表',
            'audit_basis': '合规规则',
            'context': '产品中有害物质的名称及含有物质表',
        },
        {
            'original_text': 'XXXXXXXXXXX',
            'rule': 'CYY-CN-PLACEHOLDER-001',
            'source': 'rule',
            'category': '占位符残留',
            'severity': 'serious',
            'suggestion': '替换为真实信息。',
            'description': '文档存在占位符残留。',
            'audit_basis': '中文人工审核规则',
            'context': '联系电话：XXXXXXXXXXX',
        },
    ]

    filtered, diagnostics = review_api._finalize_review_issues(issues, '联系电话：XXXXXXXXXXX', set())

    originals = {issue['original_text'] for issue in filtered}
    space_issue = next(issue for issue in filtered if issue['original_text'] == '空格')
    assert space_issue.get('possible_false_positive') is True
    assert '物质的名' not in originals
    assert 'XXXXXXXXXXX' in originals
    assert diagnostics['after_known_false_positive_ai'] == 1


def test_should_skip_rule_match_skips_ext_r013_for_pdf():
    rule = SimpleNamespace(rule_no="EXT-R013")
    match = next(re.finditer(r"\.", "Overview."))

    assert review_api._should_skip_rule_match(rule, match, "Overview.", "en", file_type="pdf") is True


def test_should_drop_punctuation_issue_drops_generic_single_punctuation_noise():
    issue = {
        'original_text': '。',
        'rule': 'PUNCT-001',
        'source': 'rule',
        'suggestion': '文档中标点符号使用必须符合规范，不得遗漏必要的标点（逗号、句号等）',
        'description': '英文文档中不应混入全角或中文标点。',
        'context': 'Overview。',
    }

    assert review_api._should_drop_punctuation_issue(issue) is True


def test_prepare_report_issues_filters_generic_punctuation_noise_before_aggregation():
    issues = [
        {
            'original_text': '。',
            'rule': 'PUNCT-001',
            'source': 'rule',
            'suggestion': '文档中标点符号使用必须符合规范，不得遗漏必要的标点（逗号、句号等）',
            'description': '英文文档中不应混入全角或中文标点。',
            'context': 'Overview。',
            'position': '10-11',
            'status': 'open',
        },
        {
            'original_text': '，',
            'rule': 'PUNCT-001',
            'source': 'rule',
            'suggestion': '文档中标点符号使用必须符合规范，不得遗漏必要的标点（逗号、句号等）',
            'description': '英文文档中不应混入全角或中文标点。',
            'context': 'Scope，',
            'position': '20-21',
            'status': 'open',
        },
    ]

    prepared = review_api._prepare_report_issues(issues, 'Overview。 Scope，')

    assert prepared == []


def test_normalize_review_issue_display_falls_back_to_description_when_suggestion_missing():
    issue = SimpleNamespace(
        suggestion='',
        description='问题说明：术语表缺少对应中文释义',
        rule='TERM-001',
        category='术语一致性',
        chapter='1.1 Scope',
        context='',
        original_text='',
        position=None,
    )

    result = review_api._normalize_review_issue_display([issue])

    assert result[0].suggestion == '术语表缺少对应中文释义'


def test_rulebook_false_positive_drops_nested_list_numbering():
    issue = {
        "source": "ai",
        "rule": "AI-STYLE-001",
        "category": "格式规范",
        "severity": "general",
        "original_text": "1)",
        "context": "1. Prepare the sample\n  1) Mix the buffer\n  2) Incubate",
        "suggestion": "外层与内层编号格式不统一，请统一为 1. 2. 3.",
        "description": "嵌套有序列表编号差异",
        "audit_basis": "格式规范",
        "confidence": 88,
    }

    assert review_api.is_rulebook_false_positive(issue) is True
    assert review_pipeline.select_review_issues([issue]) == []


def test_rulebook_false_positive_drops_official_global_site():
    issue = {
        "source": "ai",
        "rule": "DET-URL-001",
        "category": "官网地址",
        "severity": "serious",
        "original_text": "https://global-mgitech.com/",
        "context": "Visit https://global-mgitech.com/ for support.",
        "suggestion": "官网地址错误，应改为 en.mgi-tech.com",
        "description": "术语一致性：官网地址不正确",
        "audit_basis": "官网规范",
        "confidence": 90,
    }

    filtered, _diagnostics = review_api._finalize_review_issues([issue], "Visit https://global-mgitech.com/", set())
    assert filtered == []


def test_rulebook_false_positive_drops_english_email_only_contact():
    issue = {
        "source": "ai",
        "rule": "AI-CHECK-001",
        "category": "信息完整性",
        "severity": "general",
        "original_text": "MGI-service@mgi-tech.com",
        "context": "Technical support: MGI-service@mgi-tech.com",
        "suggestion": "The manufacturer contact is missing a telephone number.",
        "description": "English manual lacks a phone number",
        "audit_basis": "contact completeness",
        "confidence": 86,
    }

    assert review_api.is_rulebook_false_positive(issue) is True


def test_rulebook_false_positive_keeps_real_url_issue():
    issue = {
        "source": "rule",
        "rule": "DET-URL-001",
        "category": "官网地址",
        "severity": "serious",
        "original_text": "https://en.mgi-tech.com/",
        "context": "Visit https://en.mgi-tech.com/ for support.",
        "suggestion": "英文手册应使用 https://global-mgitech.com/",
        "description": "官网地址错误",
        "audit_basis": "官网规范",
        "confidence": 95,
    }

    assert review_api.is_rulebook_false_positive(issue) is False


def test_review_cache_version_tracks_false_positive_module():
    assert review_api.PROJECT_ROOT / "backend" / "app" / "review_engine" / "false_positives.py" in review_api.REVIEW_CACHE_VERSION_FILES


def test_should_reuse_cached_review_respects_force_and_provider():
    cached = SimpleNamespace(id=9)
    assert review_api._should_reuse_cached_review(cached) is True
    assert review_api._should_reuse_cached_review(cached, force=True) is False
    assert review_api._should_reuse_cached_review(cached, provider="qwen") is False
    assert review_api._should_reuse_cached_review(None) is False


def test_normalize_review_status_accepts_cancelled():
    assert review_api._normalize_review_status("cancelled") == "cancelled"
    assert review_api._normalize_review_status("running") == "running"


def test_review_cancel_flag_blocks_progress_and_clears():
    review_id = 99001
    review_api._clear_review_runtime_flags(review_id)
    review_api.request_review_cancel(review_id)
    assert review_api._is_review_cancelled(review_id) is True
    try:
        review_api.set_progress(review_id, "running", "规则审核", 25, "进行中")
        assert False, "cancelled review should reject progress updates"
    except review_api.ReviewCancelled:
        pass
    review_api.set_progress(
        review_id,
        "cancelled",
        "已停止",
        0,
        "用户已停止审核",
        allow_when_cancelled=True,
    )
    assert review_api.get_progress(review_id)["status"] == "cancelled"
    review_api._clear_review_runtime_flags(review_id)
    assert review_api._is_review_cancelled(review_id) is False


def test_run_cached_ai_chunk_review_bypasses_cache_on_force(monkeypatch):
    review_id = 99002
    cache_key = "force-chunk-key"
    review_api._clear_review_runtime_flags(review_id)
    review_api._review_force_rerun_ids.add(review_id)
    review_api._ai_review_chunk_cache[cache_key] = [{"rule": "cached"}]
    monkeypatch.setattr(review_api, "_build_ai_chunk_cache_key", lambda *args, **kwargs: cache_key)
    calls = []

    def fake_timeout(*args, **kwargs):
        calls.append(1)
        return {"issues": [{"rule": "fresh"}]}

    monkeypatch.setattr(review_api, "_call_with_timeout", fake_timeout)
    issues, cache_hit = review_api._run_cached_ai_chunk_review(
        review_id, "chunk", "cn", "basis", 10
    )
    assert cache_hit is False
    assert issues == [{"rule": "fresh", "review_id": review_id, "status": "pending"}]
    assert calls == [1]
    review_api._clear_review_runtime_flags(review_id)
    review_api._ai_review_chunk_cache.pop(cache_key, None)
