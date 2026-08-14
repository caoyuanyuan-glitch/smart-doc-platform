import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.api import translation
from app.api.auth import create_access_token
from app.database import Base, get_db
from app.models.knowledge import Folder, KnowledgeFile
from app.models.memory import MemoryBank
from app.models.user import User


class TranslationMemoryFileEntryApiTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        cls.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=cls.engine)
        Base.metadata.create_all(bind=cls.engine)

        app = FastAPI()
        app.include_router(translation.router, prefix="/api/translation")

        def override_get_db():
            db = cls.SessionLocal()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[get_db] = override_get_db
        cls.client = TestClient(app)

    def setUp(self):
        Base.metadata.drop_all(bind=self.engine)
        Base.metadata.create_all(bind=self.engine)
        with translation._memory_file_cache_lock:
            translation._memory_file_cache.clear()
        translation._translation_stats_cache.clear()
        self.admin_id = self._create_user("admin_user", "admin")
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)

    def tearDown(self):
        with translation._memory_file_cache_lock:
            translation._memory_file_cache.clear()
        translation._translation_stats_cache.clear()

    def _create_user(self, username: str, role: str) -> int:
        db = self.SessionLocal()
        try:
            user = User(username=username, password_hash="test-hash", display_name=username, role=role, status="active")
            db.add(user)
            db.commit()
            db.refresh(user)
            return user.id
        finally:
            db.close()

    def _auth_headers(self, username: str) -> dict:
        token = create_access_token({"sub": username})
        return {"Authorization": f"Bearer {token}"}

    def _create_memory_workbook(self, file_path: Path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Memory"
        sheet.append(["zh-CN", "en-US"])
        sheet.append(["旧原文", "Old Translation"])
        workbook.save(file_path)

    def test_memory_file_entry_endpoint_syncs_runtime_seed_and_memory_bank(self):
        runtime_path = Path(self.temp_dir.name) / "runtime-memory.xlsx"
        seed_path = Path(self.temp_dir.name) / "seed-memory.xlsx"
        self._create_memory_workbook(runtime_path)

        db = self.SessionLocal()
        try:
            root_folder = Folder(name="资源库", parent_id=None, created_by=self.admin_id)
            db.add(root_folder)
            db.flush()

            memory_folder = Folder(name="记忆库", parent_id=root_folder.id, created_by=self.admin_id)
            db.add(memory_folder)
            db.flush()

            memory_file = KnowledgeFile(
                folder_id=memory_folder.id,
                name="AI翻译语料写入Excel.xlsx",
                filename="runtime-memory.xlsx",
                file_path=str(runtime_path),
                file_size=runtime_path.stat().st_size,
                file_type="xlsx",
                created_by=self.admin_id,
            )
            db.add(memory_file)
            db.commit()
            db.refresh(memory_file)
            memory_file_id = memory_file.id
        finally:
            db.close()

        with patch("app.api.translation._build_memory_seed_file_path", return_value=seed_path):
            response = self.client.post(
                "/api/translation/memory/file-entry",
                headers=self._auth_headers("admin_user"),
                json={
                    "memory_file_id": memory_file_id,
                    "source_text": "MC_API_VERIFY_SOURCE_20260814",
                    "translated_text": "MC_API_VERIFY_TARGET_20260814",
                    "source_lang": "zh",
                    "target_lang": "en",
                },
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["memory_file_id"], memory_file_id)
        self.assertTrue(payload["saved_to_memory_bank"])

        runtime_rows = list(load_workbook(runtime_path).active.iter_rows(values_only=True))
        seed_rows = list(load_workbook(seed_path).active.iter_rows(values_only=True))
        self.assertEqual(runtime_rows[-1][:2], ("MC_API_VERIFY_SOURCE_20260814", "MC_API_VERIFY_TARGET_20260814"))
        self.assertEqual(seed_rows[-1][:2], ("MC_API_VERIFY_SOURCE_20260814", "MC_API_VERIFY_TARGET_20260814"))

        db = self.SessionLocal()
        try:
            entry = db.query(MemoryBank).filter(
                MemoryBank.source_text == "MC_API_VERIFY_SOURCE_20260814",
                MemoryBank.translated_text == "MC_API_VERIFY_TARGET_20260814",
            ).first()
            self.assertIsNotNone(entry)
            self.assertEqual(entry.source_lang, "zh")
            self.assertEqual(entry.target_lang, "en")
        finally:
            db.close()
