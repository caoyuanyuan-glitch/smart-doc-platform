#!/usr/bin/env python3
import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from io import StringIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.api.translation import (  # noqa: E402
    _build_memory_candidate_bundle,
    _count_text_units,
    _count_translatable_text_units,
    _get_memory_match_trace,
    _match_memory_candidates,
    _thread_locals,
)
from app.utils.file_utils import read_file_safe  # noqa: E402


def load_memory_candidates(csv_path: Path, source_col: int, target_col: int, encoding: str):
    if encoding == "auto":
        rows = list(csv.reader(StringIO(read_file_safe(str(csv_path)).lstrip("\ufeff"))))
    else:
        with csv_path.open("r", encoding=encoding, newline="") as handle:
            rows = list(csv.reader(handle))

    candidates = []
    for row in rows[1:]:
        if max(source_col, target_col) >= len(row):
            continue
        source_text = (row[source_col] or "").strip()
        translated_text = (row[target_col] or "").strip()
        if not source_text or not translated_text:
            continue
        candidates.append({
            "source_text": source_text,
            "translated_text": translated_text,
        })
    return candidates


def extract_excel_terms(xlsx_path: Path):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = []
    for sheet in workbook.worksheets:
        headers = None
        for row in sheet.iter_rows(values_only=True):
            values = [value if value is not None else "" for value in row]
            if headers is None and any(str(value).strip() in {"Document Name", "Name"} for value in values):
                headers = [str(value).strip() for value in values]
                continue
            if not headers or not any(values):
                continue
            if "Document Name" in headers:
                name_index = headers.index("Document Name")
            elif "Name" in headers:
                name_index = headers.index("Name")
            else:
                continue
            document_name = values[name_index] if name_index < len(values) else ""
            if not document_name:
                continue
            rows.append({
                "sheet": sheet.title,
                "row_no": str(values[0]) if values and values[0] else "",
                "source_text": str(document_name),
            })
    return rows


def evaluate_memory(xlsx_path: Path, csv_path: Path, source_col: int, target_col: int, encoding: str):
    items = extract_excel_terms(xlsx_path)
    bundle = _build_memory_candidate_bundle(load_memory_candidates(csv_path, source_col, target_col, encoding))

    reason_counts = Counter()
    sheet_counts = defaultdict(lambda: Counter())
    misses = []
    raw_units_total = 0
    translatable_units_total = 0
    translatable_units_matched = 0

    for item in items:
        if hasattr(_thread_locals, "memory_match_trace"):
            delattr(_thread_locals, "memory_match_trace")

        source_text = item["source_text"]
        raw_units = _count_text_units(source_text)
        translatable_units = _count_translatable_text_units(source_text)
        raw_units_total += raw_units
        translatable_units_total += translatable_units

        matched = _match_memory_candidates(source_text, bundle)
        sheet = item["sheet"]
        if matched:
            translatable_units_matched += translatable_units
            trace = _get_memory_match_trace()
            reason = trace[-1]["reason"] if trace else "matched_untraced"
            reason_counts[reason] += 1
            sheet_counts[sheet]["matched"] += 1
            sheet_counts[sheet][reason] += 1
        else:
            reason_counts["miss"] += 1
            sheet_counts[sheet]["miss"] += 1
            misses.append({
                **item,
                "translatable_units": translatable_units,
            })

    total_rows = len(items)
    matched_rows = total_rows - len(misses)
    return {
        "xlsx_path": str(xlsx_path),
        "memory_csv_path": str(csv_path),
        "rows_total": total_rows,
        "rows_matched": matched_rows,
        "row_match_rate": round(matched_rows / total_rows, 4) if total_rows else 0,
        "raw_units_total": raw_units_total,
        "translatable_units_total": translatable_units_total,
        "translatable_units_matched": translatable_units_matched,
        "translatable_unit_match_rate": round(
            translatable_units_matched / translatable_units_total, 4
        ) if translatable_units_total else 0,
        "reason_counts": dict(reason_counts),
        "sheet_summary": {
            sheet: {
                "total": counts["matched"] + counts["miss"],
                "matched": counts["matched"],
                "miss": counts["miss"],
                "rate": round(counts["matched"] / (counts["matched"] + counts["miss"]), 4)
                if (counts["matched"] + counts["miss"]) else 0,
            }
            for sheet, counts in sorted(sheet_counts.items())
        },
        "misses": misses,
    }


def print_summary(result, miss_limit: int):
    print(f"rows_total: {result['rows_total']}")
    print(f"rows_matched: {result['rows_matched']}")
    print(f"row_match_rate: {result['row_match_rate']:.2%}")
    print(f"raw_units_total: {result['raw_units_total']}")
    print(f"translatable_units_total: {result['translatable_units_total']}")
    print(f"translatable_units_matched: {result['translatable_units_matched']}")
    print(f"translatable_unit_match_rate: {result['translatable_unit_match_rate']:.2%}")
    print("reason_counts:")
    for reason, count in sorted(result["reason_counts"].items(), key=lambda item: (-item[1], item[0])):
        print(f"  {reason}: {count}")
    print("sheet_summary:")
    for sheet, summary in result["sheet_summary"].items():
        print(f"  {sheet}: matched={summary['matched']} miss={summary['miss']} rate={summary['rate']:.2%}")
    print(f"misses_top_{miss_limit}:")
    for item in result["misses"][:miss_limit]:
        print(f"  [{item['sheet']}] {item['source_text']} | row={item['row_no']} | translatable_units={item['translatable_units']}")


def main():
    parser = argparse.ArgumentParser(description="Evaluate translation memory coverage against an Excel source file.")
    parser.add_argument("xlsx", help="Path to the source Excel file")
    parser.add_argument("memory_csv", help="Path to the translation memory CSV file")
    parser.add_argument("--source-col", type=int, default=1, help="Zero-based source text column in memory CSV")
    parser.add_argument("--target-col", type=int, default=0, help="Zero-based translated text column in memory CSV")
    parser.add_argument("--encoding", default="auto", help="Memory CSV encoding, or auto")
    parser.add_argument("--json", dest="json_output", help="Optional JSON result output path")
    parser.add_argument("--miss-limit", type=int, default=20, help="How many miss rows to print")
    args = parser.parse_args()

    result = evaluate_memory(
        Path(args.xlsx),
        Path(args.memory_csv),
        source_col=args.source_col,
        target_col=args.target_col,
        encoding=args.encoding,
    )
    print_summary(result, args.miss_limit)

    if args.json_output:
        output_path = Path(args.json_output)
        output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"json_output: {output_path}")


if __name__ == "__main__":
    raise SystemExit(main())
